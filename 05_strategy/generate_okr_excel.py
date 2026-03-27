#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── 颜色常量 ──
BLUE_DARK   = "1F497D"
BLUE_MID    = "2E74B5"
BLUE_LIGHT  = "DEEAF1"
BLUE_HEADER = "BDD7EE"
ORANGE      = "F5A623"
GREEN       = "70AD47"
GRAY_LIGHT  = "F2F2F2"
WHITE       = "FFFFFF"
RED_LIGHT   = "FCE4D6"

def header_fill(color=BLUE_MID):
    return PatternFill("solid", fgColor=color)

def light_fill(color=BLUE_LIGHT):
    return PatternFill("solid", fgColor=color)

def gray_fill():
    return PatternFill("solid", fgColor=GRAY_LIGHT)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=10, color=WHITE):
    return Font(name="Arial", bold=True, size=size, color=color)

def body_font(size=9, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_cell(ws, row, col, value, fill=None, font=None, align=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill   = fill
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if border: cell.border = thin_border()
    return cell

def merge_title(ws, row, col_start, col_end, value, bg=BLUE_DARK, fg=WHITE, size=11):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", bold=True, size=size, color=fg)
    cell.alignment = center()
    cell.border    = thin_border()

# ══════════════════════════════════════════
# Sheet 1：OKR 总览
# ══════════════════════════════════════════
ws1 = wb.active
ws1.title = "OKR总览"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 6
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 42
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 10
ws1.column_dimensions["F"].width = 10

# 大标题
ws1.row_dimensions[1].height = 36
ws1.merge_cells("A1:F1")
c = ws1["A1"]
c.value    = "预测市场 GMGN — 市场总监 OKR（Q1 优化版）"
c.fill     = PatternFill("solid", fgColor=BLUE_DARK)
c.font     = Font(name="Arial", bold=True, size=14, color=WHITE)
c.alignment = center()

# 副标题
ws1.row_dimensions[2].height = 20
ws1.merge_cells("A2:F2")
c2 = ws1["A2"]
c2.value     = "获客策略：代理渠道 30% + 直客渠道 70%  |  目标市场：欧美 + 东南亚  |  用户：加密货币用户 + 美股用户"
c2.fill      = PatternFill("solid", fgColor=BLUE_MID)
c2.font      = Font(name="Arial", size=9, color=WHITE)
c2.alignment = center()

# 列标题
row = 3
ws1.row_dimensions[row].height = 22
headers = ["#", "Objective", "Key Result（关键结果）", "截止时间", "负责人", "状态"]
fills   = [BLUE_HEADER] * 6
for ci, (h, f) in enumerate(zip(headers, fills), 1):
    set_cell(ws1, row, ci, h,
             fill=header_fill(BLUE_MID),
             font=header_font(10),
             align=center())

# OKR 数据
okr_rows = [
    # (obj_label, obj_text, kr_label, kr_text, deadline, owner, status, obj_color)
    ("O1", "建立品牌认知\n覆盖欧美+东南亚核心用户圈",
     "KR1", "加密媒体渠道：发布 8 篇品牌内容，欧美+东南亚累计曝光 >50万",       "第1季度末", "", "未开始", "2E74B5"),
    ("O1", "",
     "KR2", "财经媒体渠道：发布 4 篇股票预测内容，覆盖欧美受众",               "第1季度末", "", "未开始", "2E74B5"),
    ("O1", "",
     "KR3", "Twitter 账号粉丝达到 5,000，每条内容互动率 >3%",                  "第1季度末", "", "未开始", "2E74B5"),

    ("O2", "代理渠道（30%）\n用KOL激励替代BD冷启动",
     "KR1", "签约 5-8 个欧美+东南亚加密/美股中腰部KOL（1-10万粉），提供激励+定制化活动支持", "第2个月末", "", "未开始", "ED7D31"),
    ("O2", "",
     "KR2", "KOL合作产出内容 >20 条，带来 Waitlist 注册用户 >300",             "第2个月末", "", "未开始", "ED7D31"),
    ("O2", "",
     "KR3", "举办至少 2 场预测体验定制化活动，每场参与人数 >200",               "第1季度末", "", "未开始", "ED7D31"),

    ("O3", "直客渠道（70%）\n多路并发获取真实用户",
     "KR1（积分）",  "上线积分体系（交易/邀请/持仓），首月参与 >500，裂变新用户 >200",  "上线后第1月", "", "未开始", "70AD47"),
    ("O3", "",
     "KR2（大使）",  "招募全球大使 20 人（欧美+东南亚），瓜分奖池，带来注册用户 >300",   "第1季度末",  "", "未开始", "70AD47"),
    ("O3", "",
     "KR3（SEO）",   "布局 20 个核心关键词，自然搜索流量达 >5,000/月",                 "第1季度末",  "", "未开始", "70AD47"),
    ("O3", "",
     "KR4（邮件）",  "建立流失用户召回机制，邮件打开率 >25%，召回率 >10%",             "上线后第1月", "", "未开始", "70AD47"),

    ("O4", "产品上线首月\n跑通两类用户转化路径",
     "KR1", "加密用户注册 700 + 美股用户注册 300，合计 1,000",                  "上线后第1月", "", "未开始", "7030A0"),
    ("O4", "",
     "KR2", "跟单功能首月活跃用户 200，人均跟单次数 >3次",                       "上线后第1月", "", "未开始", "7030A0"),
    ("O4", "",
     "KR3", "用户自发分享跟单战绩内容 >50 条（战绩卡设计驱动）",                 "上线后第1月", "", "未开始", "7030A0"),
]

prev_obj = None
obj_start_row = {}
row = 4
for (obj, obj_text, kr, kr_text, deadline, owner, status, obj_color) in okr_rows:
    ws1.row_dimensions[row].height = 30

    # O 列
    if obj != prev_obj:
        obj_start_row[obj] = row
        set_cell(ws1, row, 1, obj,
                 fill=PatternFill("solid", fgColor=obj_color),
                 font=Font(name="Arial", bold=True, size=10, color=WHITE),
                 align=center())
        set_cell(ws1, row, 2, obj_text,
                 fill=PatternFill("solid", fgColor=obj_color),
                 font=Font(name="Arial", bold=True, size=9, color=WHITE),
                 align=left())
        prev_obj = obj
    else:
        set_cell(ws1, row, 1, "",
                 fill=PatternFill("solid", fgColor=obj_color),
                 font=body_font(), align=center())
        set_cell(ws1, row, 2, "",
                 fill=PatternFill("solid", fgColor=obj_color),
                 font=body_font(), align=left())

    # KR 列
    row_fill = light_fill("EBF3FB") if row % 2 == 0 else light_fill(WHITE)
    set_cell(ws1, row, 3, kr_text,  fill=row_fill, font=body_font(9), align=left())
    set_cell(ws1, row, 4, deadline, fill=row_fill, font=body_font(9), align=center())
    set_cell(ws1, row, 5, owner,    fill=row_fill, font=body_font(9), align=center())
    set_cell(ws1, row, 6, status,   fill=row_fill, font=body_font(9), align=center())
    row += 1

# ══════════════════════════════════════════
# Sheet 2：冷启动执行计划
# ══════════════════════════════════════════
ws2 = wb.create_sheet("冷启动执行计划")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 36
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 16
ws2.column_dimensions["F"].width = 10

ws2.row_dimensions[1].height = 32
ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value     = "冷启动渠道执行计划"
c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
c.alignment = center()

row = 2
for h in ["渠道类型", "平台", "具体动作", "频次", "目标", "优先级"]:
    set_cell(ws2, row, ["渠道类型","平台","具体动作","频次","目标","优先级"].index(h)+1,
             h, fill=header_fill(BLUE_MID), font=header_font(), align=center())

plan = [
    ("代理渠道", "Twitter/X\nYouTube",
     "签约欧美+东南亚中腰部KOL（加密/美股）\n提供激励+定制化活动支持\n产出 >20条内容",
     "持续", "Waitlist >300\nKOL签约5-8个", "★★★"),
    ("代理渠道", "线上活动",
     "举办预测体验定制化活动\n对应代理定制主题\n参与有奖励",
     "每月1场", "每场参与 >200人", "★★"),
    ("直客-媒体", "加密媒体\n（Decrypt等）",
     "发布加密货币预测相关品牌内容\n聚焦欧美+东南亚地区\n差异化产品特性宣传",
     "每月2-3篇", "累计曝光 >50万", "★★★"),
    ("直客-媒体", "财经媒体\n（Bloomberg等）",
     "发布美股预测相关内容\n聚焦欧美市场\n财经类渠道宣发",
     "每月1-2篇", "覆盖欧美受众", "★★"),
    ("直客-增长", "产品内",
     "上线积分奖励体系（交易/邀请/持仓）\n设计邀请裂变机制\n排行榜激励",
     "持续运营", "参与 >500\n裂变新用户 >200", "★★★"),
    ("直客-增长", "全球社群",
     "招募全球大使20人（欧美+东南亚）\n瓜分奖池激励\n大使自发传播",
     "持续招募", "带来注册 >300", "★★"),
    ("直客-SEO", "Google/Bing",
     "布局20个核心关键词\nprediction market / Polymarket alternative等\n内容SEO优化",
     "持续", "自然流量 >5000/月", "★★"),
    ("直客-邮件", "邮件系统",
     "建立流失用户名单\n设计召回邮件序列\n个性化内容推送",
     "每周1次", "打开率 >25%\n召回率 >10%", "★"),
]

for ri, (ch, pl, action, freq, goal, pri) in enumerate(plan):
    r = ri + 3
    ws2.row_dimensions[r].height = 42
    rf = light_fill("EBF3FB") if ri % 2 == 0 else light_fill(WHITE)
    set_cell(ws2, r, 1, ch,     fill=rf, font=body_font(9, bold=True), align=center())
    set_cell(ws2, r, 2, pl,     fill=rf, font=body_font(9), align=center())
    set_cell(ws2, r, 3, action, fill=rf, font=body_font(9), align=left())
    set_cell(ws2, r, 4, freq,   fill=rf, font=body_font(9), align=center())
    set_cell(ws2, r, 5, goal,   fill=rf, font=body_font(9), align=left())
    set_cell(ws2, r, 6, pri,    fill=rf, font=body_font(9), align=center())

# ══════════════════════════════════════════
# Sheet 3：OKR 优化对比
# ══════════════════════════════════════════
ws3 = wb.create_sheet("OKR优化对比")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 30

ws3.row_dimensions[1].height = 32
ws3.merge_cells("A1:C1")
c = ws3["A1"]
c.value     = "OKR 优化前后对比"
c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
c.alignment = center()

row = 2
for h in ["维度", "优化前", "优化后"]:
    set_cell(ws3, row, ["维度","优化前","优化后"].index(h)+1,
             h, fill=header_fill(BLUE_MID), font=header_font(), align=center())

cmp_data = [
    ("用户分类",     "未细分",                       "加密货币用户 + 美股用户分开打"),
    ("地区策略",     "无",                           "聚焦欧美 + 东南亚"),
    ("代理渠道",     "仅提KOL合作",                  "明确KOL激励替代BD，针对前期无手续费的专项方案"),
    ("直客增长机制", "无",                           "积分体系 + 全球大使 + SEO + 邮件营销四路并发"),
    ("用户召回",     "无",                           "邮件营销覆盖流失用户名单，召回率目标 >10%"),
    ("媒体渠道",     "泛化内容输出",                 "加密媒体 + 财经媒体分开布局，匹配不同用户群"),
    ("KPI拆解",      "注册用户合计1000",             "加密用户700 + 美股用户300，分渠道追踪"),
]

for ri, (dim, before, after) in enumerate(cmp_data):
    r = ri + 3
    ws3.row_dimensions[r].height = 36
    rf = light_fill("EBF3FB") if ri % 2 == 0 else light_fill(WHITE)
    set_cell(ws3, r, 1, dim,    fill=rf, font=body_font(9, bold=True), align=center())
    set_cell(ws3, r, 2, before, fill=PatternFill("solid", fgColor="FCE4D6"),
             font=body_font(9), align=left())
    set_cell(ws3, r, 3, after,  fill=PatternFill("solid", fgColor="E2EFDA"),
             font=body_font(9), align=left())

path = "/Users/coco/agent-twitter/GMGN_市场OKR_Q1.xlsx"
wb.save(path)
print(f"Excel 已生成：{path}")
