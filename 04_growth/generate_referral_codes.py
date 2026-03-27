#!/usr/bin/env python3
"""
MoonX Waitlist Referral System
================================
Generates referral codes, tracks invitations, and manages waitlist queue.
No product code changes needed — runs standalone, backed by local CSV/Excel.

Usage:
  python generate_referral_codes.py --action register --email user@example.com --name "Alice"
  python generate_referral_codes.py --action register --email user@example.com --name "Bob" --ref MOON-ALICE-1A2B
  python generate_referral_codes.py --action confirm_login --email user@example.com
  python generate_referral_codes.py --action status --email user@example.com
  python generate_referral_codes.py --action leaderboard
  python generate_referral_codes.py --action export
  python generate_referral_codes.py --action stats
"""

import argparse
import csv
import hashlib
import json
import os
import random
import string
import sys
from datetime import datetime, timezone

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── Config ───────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "referral_data")
USERS_CSV = os.path.join(DATA_DIR, "waitlist_users.csv")
REFS_CSV  = os.path.join(DATA_DIR, "referral_links.csv")
LOG_CSV   = os.path.join(DATA_DIR, "activity_log.csv")
EXCEL_OUT = os.path.join(DATA_DIR, "MoonX_Waitlist_Dashboard.xlsx")

# Tier thresholds
TIER_EARLY_BIRD  = 3   # invites needed for Early Bird
TIER_POWER_USER  = 10  # invites needed for Power User (KOL tier)

# Points config
POINTS_SIGNUP        = 100   # base points on registration
POINTS_PER_INVITE    = 50    # per valid (logged-in) referral
POINTS_BONUS_EARLY   = 200   # bonus on hitting 3 invites
POINTS_BONUS_POWER   = 500   # bonus on hitting 10 invites

# Queue position multiplier: earlier signup = smaller queue number
# We never expose raw queue position, only relative rank tier
MOONX_REFERRAL_URL = "https://www.bydfi.com/en/moonx/markets/trending?ref="

# ─── Utilities ────────────────────────────────────────────────────────────────

def now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def generate_code(email: str, length: int = 8) -> str:
    """
    Deterministic-ish but collision-resistant referral code.
    Format: MOON-{NAME_PREFIX}-{HEX4}
    e.g.   MOON-ALICE-3F9A
    """
    name_part = email.split("@")[0].upper()[:5].replace(".", "")
    salt = "".join(random.choices(string.hexdigits.upper(), k=4))
    # Prevent same-name collision with hash suffix
    h = hashlib.sha256(f"{email}{salt}".encode()).hexdigest()[:4].upper()
    return f"MOON-{name_part}-{h}"

def ensure_dirs():
    os.makedirs(DATA_DIR, exist_ok=True)

def ensure_users_csv():
    ensure_dirs()
    if not os.path.exists(USERS_CSV):
        with open(USERS_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=user_fields())
            writer.writeheader()

def ensure_refs_csv():
    ensure_dirs()
    if not os.path.exists(REFS_CSV):
        with open(REFS_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=ref_fields())
            writer.writeheader()

def ensure_log_csv():
    ensure_dirs()
    if not os.path.exists(LOG_CSV):
        with open(LOG_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=log_fields())
            writer.writeheader()

def user_fields():
    return [
        "email", "name", "referral_code", "referred_by_code",
        "queue_position", "registered_at", "first_login_at",
        "valid_invites", "pending_invites", "points", "tier",
        "tier_upgraded_at", "status"
        # status: waitlisted | early_bird | power_user
    ]

def ref_fields():
    return [
        "referral_code", "inviter_email", "invitee_email",
        "invited_at", "login_confirmed_at", "is_valid"
        # is_valid: 0 (pending) | 1 (confirmed by login)
    ]

def log_fields():
    return ["timestamp", "email", "action", "detail"]

# ─── Data Access ──────────────────────────────────────────────────────────────

def load_users() -> dict:
    ensure_users_csv()
    users = {}
    with open(USERS_CSV, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            users[row["email"]] = row
    return users

def save_users(users: dict):
    ensure_dirs()
    with open(USERS_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=user_fields())
        writer.writeheader()
        writer.writerows(users.values())

def load_refs() -> list:
    ensure_refs_csv()
    rows = []
    with open(REFS_CSV, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    return rows

def save_refs(refs: list):
    ensure_dirs()
    with open(REFS_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ref_fields())
        writer.writeheader()
        writer.writerows(refs)

def append_log(email: str, action: str, detail: str = ""):
    ensure_log_csv()
    with open(LOG_CSV, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=log_fields())
        writer.writerow({
            "timestamp": now_utc(),
            "email": email,
            "action": action,
            "detail": detail
        })

# ─── Business Logic ───────────────────────────────────────────────────────────

def get_queue_position(users: dict) -> int:
    """Next queue position = total registered + 1."""
    return len(users) + 1

def compute_tier(valid_invites: int) -> str:
    if valid_invites >= TIER_POWER_USER:
        return "power_user"
    elif valid_invites >= TIER_EARLY_BIRD:
        return "early_bird"
    return "waitlisted"

def find_user_by_code(users: dict, code: str):
    for u in users.values():
        if u["referral_code"] == code:
            return u
    return None

def action_register(email: str, name: str, ref_code: str = None):
    """Register a new user on the waitlist."""
    users = load_users()

    if email in users:
        print(f"[ERROR] {email} is already registered.")
        sys.exit(1)

    # Validate referral code if provided
    inviter = None
    if ref_code:
        inviter = find_user_by_code(users, ref_code)
        if not inviter:
            print(f"[WARN] Referral code '{ref_code}' not found. Registering without referral.")
            ref_code = None

    # Generate unique code
    code = generate_code(email)
    while find_user_by_code(users, code):  # collision guard
        code = generate_code(email)

    queue_pos = get_queue_position(users)
    ts = now_utc()

    new_user = {
        "email": email,
        "name": name,
        "referral_code": code,
        "referred_by_code": ref_code or "",
        "queue_position": queue_pos,
        "registered_at": ts,
        "first_login_at": "",
        "valid_invites": 0,
        "pending_invites": 0,
        "points": POINTS_SIGNUP,
        "tier": "waitlisted",
        "tier_upgraded_at": "",
        "status": "waitlisted"
    }
    users[email] = new_user
    save_users(users)
    append_log(email, "register", f"code={code} queue={queue_pos} ref={ref_code or 'none'}")

    # Record referral link as pending (awaiting login confirmation)
    if inviter:
        refs = load_refs()
        refs.append({
            "referral_code": ref_code,
            "inviter_email": inviter["email"],
            "invitee_email": email,
            "invited_at": ts,
            "login_confirmed_at": "",
            "is_valid": "0"
        })
        save_refs(refs)
        # Increment inviter's pending count
        users = load_users()
        users[inviter["email"]]["pending_invites"] = int(users[inviter["email"]]["pending_invites"]) + 1
        save_users(users)
        append_log(inviter["email"], "referral_pending", f"invitee={email}")

    print(f"\n[OK] Registered: {name} <{email}>")
    print(f"     Your referral code : {code}")
    print(f"     Your share link    : {MOONX_REFERRAL_URL}{code}")
    print(f"     Queue position     : #{queue_pos}")
    print(f"     Starting points    : {POINTS_SIGNUP}")
    print(f"     Tier               : Waitlisted")
    print(f"\n     Invite 3 friends to unlock Early Bird access.")
    print(f"     Invite 10 friends for +20% points bonus + KOL Dashboard.")


def action_confirm_login(email: str):
    """
    Called when a referred user completes their first login.
    Validates the referral link and awards points to inviter.
    In production: call this via webhook from your auth system.
    """
    users = load_users()
    if email not in users:
        print(f"[ERROR] User {email} not found.")
        sys.exit(1)

    user = users[email]
    if user["first_login_at"]:
        print(f"[INFO] {email} has already confirmed login. No changes.")
        return

    ts = now_utc()
    user["first_login_at"] = ts
    save_users(users)
    append_log(email, "first_login", "")

    # Check if this user was referred by someone
    refs = load_refs()
    updated = False
    for ref in refs:
        if ref["invitee_email"] == email and ref["is_valid"] == "0":
            ref["is_valid"] = "1"
            ref["login_confirmed_at"] = ts
            updated = True

            # Award inviter
            inviter_email = ref["inviter_email"]
            users = load_users()
            if inviter_email in users:
                inviter = users[inviter_email]
                prev_valid = int(inviter["valid_invites"])
                new_valid  = prev_valid + 1
                prev_pending = max(0, int(inviter["pending_invites"]) - 1)
                pts = int(inviter["points"]) + POINTS_PER_INVITE
                old_tier = inviter["tier"]

                # Bonus milestones
                bonus = 0
                if prev_valid < TIER_EARLY_BIRD <= new_valid:
                    bonus += POINTS_BONUS_EARLY
                    print(f"[MILESTONE] {inviter_email} hit 3 invites → Early Bird unlocked!")
                if prev_valid < TIER_POWER_USER <= new_valid:
                    bonus += POINTS_BONUS_POWER
                    print(f"[MILESTONE] {inviter_email} hit 10 invites → Power User unlocked!")

                pts += bonus
                new_tier = compute_tier(new_valid)
                tier_upgraded_at = inviter["tier_upgraded_at"]
                if new_tier != old_tier:
                    tier_upgraded_at = ts

                inviter.update({
                    "valid_invites": new_valid,
                    "pending_invites": prev_pending,
                    "points": pts,
                    "tier": new_tier,
                    "tier_upgraded_at": tier_upgraded_at,
                    "status": new_tier
                })
                save_users(users)
                append_log(inviter_email, "invite_confirmed",
                           f"invitee={email} pts+{POINTS_PER_INVITE+bonus} tier={new_tier}")
                print(f"[OK] Referral confirmed: {inviter_email} gets +{POINTS_PER_INVITE + bonus} pts "
                      f"(valid_invites={new_valid}, tier={new_tier})")
            break

    if updated:
        save_refs(refs)
    else:
        print(f"[INFO] No pending referral found for {email}. Login recorded.")


def action_status(email: str):
    """Show a user's waitlist status and referral stats."""
    users = load_users()
    if email not in users:
        print(f"[ERROR] User {email} not found.")
        sys.exit(1)

    u = users[email]
    valid   = int(u["valid_invites"])
    pending = int(u["pending_invites"])
    pts     = int(u["points"])
    tier    = u["tier"]
    code    = u["referral_code"]
    queue   = u["queue_position"]

    # Points multiplier
    multiplier = "1.2x" if tier == "power_user" else "1.0x"

    # Next milestone
    if valid < TIER_EARLY_BIRD:
        next_goal = f"Invite {TIER_EARLY_BIRD - valid} more → Early Bird (+{POINTS_BONUS_EARLY} pts)"
    elif valid < TIER_POWER_USER:
        next_goal = f"Invite {TIER_POWER_USER - valid} more → Power User (+{POINTS_BONUS_POWER} pts + KOL Dashboard)"
    else:
        next_goal = "MAX tier reached. You are a MoonX Power User."

    tier_labels = {
        "waitlisted": "Waitlisted",
        "early_bird": "Early Bird",
        "power_user": "Power User"
    }

    print(f"\n{'='*50}")
    print(f"  MoonX Waitlist — {u['name']} <{email}>")
    print(f"{'='*50}")
    print(f"  Queue Position : #{queue}")
    print(f"  Tier           : {tier_labels.get(tier, tier)}")
    print(f"  Points         : {pts}  (multiplier: {multiplier})")
    print(f"  Valid Invites  : {valid}")
    print(f"  Pending Invites: {pending} (awaiting first login)")
    print(f"  Referral Code  : {code}")
    print(f"  Share Link     : {MOONX_REFERRAL_URL}{code}")
    print(f"\n  Next Goal      : {next_goal}")
    print(f"{'='*50}\n")


def action_leaderboard():
    """Print top 20 users by valid invites."""
    users = load_users()
    if not users:
        print("[INFO] No users yet.")
        return

    ranked = sorted(users.values(),
                    key=lambda u: (int(u["valid_invites"]), int(u["points"])),
                    reverse=True)[:20]

    print(f"\n{'='*60}")
    print(f"  MoonX Waitlist Leaderboard  (Top {len(ranked)})")
    print(f"{'='*60}")
    print(f"  {'#':<4} {'Name':<20} {'Invites':>7} {'Points':>8} {'Tier':<12}")
    print(f"  {'-'*56}")
    for i, u in enumerate(ranked, 1):
        tier_emoji = {"power_user": "[POW]", "early_bird": "[BIRD]"}.get(u["tier"], "[WAIT]")
        print(f"  {i:<4} {u['name'][:19]:<20} {u['valid_invites']:>7} {u['points']:>8} {tier_emoji}")
    print(f"{'='*60}\n")


def action_stats():
    """Print aggregate stats."""
    users = load_users()
    refs  = load_refs()

    total = len(users)
    waitlisted  = sum(1 for u in users.values() if u["tier"] == "waitlisted")
    early_birds = sum(1 for u in users.values() if u["tier"] == "early_bird")
    power_users = sum(1 for u in users.values() if u["tier"] == "power_user")
    total_valid = sum(int(u["valid_invites"]) for u in users.values())
    total_pts   = sum(int(u["points"]) for u in users.values())
    k_factor    = round(total_valid / total, 2) if total else 0

    print(f"\n{'='*50}")
    print(f"  MoonX Waitlist Stats  [{now_utc()}]")
    print(f"{'='*50}")
    print(f"  Total Registrations  : {total}")
    print(f"  Waitlisted           : {waitlisted}")
    print(f"  Early Bird           : {early_birds}")
    print(f"  Power User           : {power_users}")
    print(f"  Total Valid Invites  : {total_valid}")
    print(f"  Total Points Issued  : {total_pts}")
    print(f"  K-Factor (invites/user): {k_factor}")
    print(f"{'='*50}\n")


def action_export():
    """Export data to Excel dashboard (requires openpyxl)."""
    if not HAS_OPENPYXL:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    users = load_users()
    refs  = load_refs()

    wb = openpyxl.Workbook()

    # Sheet 1: Users
    ws1 = wb.active
    ws1.title = "Waitlist Users"
    ws1.append(["#", "Name", "Email", "Referral Code", "Queue Position",
                "Valid Invites", "Pending Invites", "Points", "Tier",
                "Referred By", "Registered At", "First Login At"])

    tier_labels = {"waitlisted": "Waitlisted", "early_bird": "Early Bird", "power_user": "Power User"}
    for i, u in enumerate(sorted(users.values(), key=lambda x: int(x["queue_position"])), 1):
        ws1.append([
            i, u["name"], u["email"], u["referral_code"],
            int(u["queue_position"]), int(u["valid_invites"]),
            int(u["pending_invites"]), int(u["points"]),
            tier_labels.get(u["tier"], u["tier"]),
            u["referred_by_code"],
            u["registered_at"], u["first_login_at"]
        ])

    # Sheet 2: Referral Links
    ws2 = wb.create_sheet("Referral Links")
    ws2.append(["Referral Code", "Inviter Email", "Invitee Email",
                "Invited At", "Login Confirmed At", "Is Valid"])
    for r in refs:
        ws2.append([r["referral_code"], r["inviter_email"], r["invitee_email"],
                    r["invited_at"], r["login_confirmed_at"],
                    "Yes" if r["is_valid"] == "1" else "Pending"])

    # Sheet 3: Summary Stats
    ws3 = wb.create_sheet("Stats Summary")
    total = len(users)
    waitlisted  = sum(1 for u in users.values() if u["tier"] == "waitlisted")
    early_birds = sum(1 for u in users.values() if u["tier"] == "early_bird")
    power_users = sum(1 for u in users.values() if u["tier"] == "power_user")
    total_valid = sum(int(u["valid_invites"]) for u in users.values())
    k_factor    = round(total_valid / total, 2) if total else 0

    ws3.append(["Metric", "Value"])
    ws3.append(["Total Registrations", total])
    ws3.append(["Waitlisted", waitlisted])
    ws3.append(["Early Bird", early_birds])
    ws3.append(["Power User", power_users])
    ws3.append(["Total Valid Invites", total_valid])
    ws3.append(["K-Factor", k_factor])
    ws3.append(["Export Time", now_utc()])

    wb.save(EXCEL_OUT)
    print(f"[OK] Exported to: {EXCEL_OUT}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="MoonX Waitlist Referral System",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("--action", required=True,
                        choices=["register", "confirm_login", "status",
                                 "leaderboard", "stats", "export"],
                        help="Action to perform")
    parser.add_argument("--email", help="User email address")
    parser.add_argument("--name",  help="User display name (for register)")
    parser.add_argument("--ref",   help="Referral code used during registration")

    args = parser.parse_args()

    if args.action == "register":
        if not args.email or not args.name:
            print("[ERROR] --email and --name required for register.")
            sys.exit(1)
        action_register(args.email, args.name, args.ref)

    elif args.action == "confirm_login":
        if not args.email:
            print("[ERROR] --email required for confirm_login.")
            sys.exit(1)
        action_confirm_login(args.email)

    elif args.action == "status":
        if not args.email:
            print("[ERROR] --email required for status.")
            sys.exit(1)
        action_status(args.email)

    elif args.action == "leaderboard":
        action_leaderboard()

    elif args.action == "stats":
        action_stats()

    elif args.action == "export":
        action_export()


if __name__ == "__main__":
    main()
