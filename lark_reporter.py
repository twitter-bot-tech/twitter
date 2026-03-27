#!/usr/bin/env python3
"""
MoonX Lark 报告系统
每天 18:30 BJT — 各部门 Bot 发日报，Team Lead Bot 汇总
"""
import os, json, re, urllib.request
from pathlib import Path
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env", override=True)
load_dotenv(Path(__file__).parent / ".env.outreach", override=True)

BJT  = ZoneInfo("Asia/Shanghai")
BASE = Path(__file__).parent

WEBHOOKS = {
    "lead":     os.getenv("LARK_LEAD"),
    "social":   os.getenv("LARK_SOCIAL"),
    "seo":      os.getenv("LARK_SEO"),
    "kol":      os.getenv("LARK_KOL"),
    "growth":   os.getenv("LARK_GROWTH"),
    "strategy": os.getenv("LARK_STRATEGY"),
}


def send(bot: str, text: str):
    url = WEBHOOKS.get(bot)
    if not url:
        return
    payload = json.dumps({"msg_type": "text", "content": {"text": text}}).encode()
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
    try:
        res = json.loads(urllib.request.urlopen(req, timeout=10).read())
        return res.get("code") == 0
    except Exception as e:
        print(f"Lark send error ({bot}): {e}")
        return False


def today() -> str:
    return datetime.now(BJT).strftime("%Y-%m-%d")

def yesterday() -> str:
    return (datetime.now(BJT) - timedelta(days=1)).strftime("%Y-%m-%d")


# ── 员工1：社媒运营 ──────────────────────────────────────────────────────────

def report_social() -> str:
    t = today()
    # 推文记录
    tweet_path = BASE / "01_social_media/logs/tweet_history.json"
    tweets_today = 0
    if tweet_path.exists():
        data = json.loads(tweet_path.read_text())
        tweets_today = len([x for x in data if x.get("date") == t])

    # 推文日志里的错误数
    tweet_log = BASE / "01_social_media/logs/tweet.log"
    errors = 0
    if tweet_log.exists():
        lines = tweet_log.read_text(errors="ignore").split("\n")
        errors = sum(1 for l in lines[-100:] if "[ERROR]" in l and t in l)

    # TG日志
    tg_log = BASE / "01_social_media/logs/daily_tweet_poster.log"
    tg_sent = 0
    if tg_log.exists():
        lines = tg_log.read_text(errors="ignore").split("\n")
        tg_sent = sum(1 for l in lines[-50:] if "✅" in l and t in l)

    status = "🟢 正常" if tweets_today > 0 else "🔴 今日未发"
    tw_icon = "✅" if tweets_today > 0 else "❌"
    tg_icon = "✅" if tg_sent > 0 else "⏳"
    err_icon = "🔴" if errors > 0 else "✅"

    msg = f"""📱 员工1号｜社媒运营日报 {t}
总状态：{status}

**📊 今日执行**
| 任务 | 状态 | 数量 |
|------|------|------|
| Twitter 推文 | {tw_icon} | {tweets_today} 条 |
| Telegram 推送 | {tg_icon} | {tg_sent} 次 |
| 脚本异常 | {err_icon} | {errors} 个 |

**⚙️ 定时任务**
| 时间 | 任务 |
|------|------|
| 09:00 | trending推文 |
| 14:00 | smart-money推文 |
| 20:00 | closing-soon推文 |
| 08:00/14:05/20:05 | TG早中晚报 |

**📋 待执行**
| 优先级 | 事项 |
|--------|------|
| 🟠 中 | 大V评论策略落地（5个账号/天） |
| 🟡 低 | 热点监控排入 09:00/14:00/20:00 |"""
    return msg


# ── 员工2：SEO专家 ────────────────────────────────────────────────────────────

def report_seo() -> str:
    t = today()
    pub = BASE / "02_seo/logs/published_articles.json"
    articles = 0
    if pub.exists():
        data = json.loads(pub.read_text())
        articles = len([x for x in data if x.get("date", "") == t])

    rank = BASE / "02_seo/logs/rank_history.json"
    rank_info = "暂无数据"
    if rank.exists():
        data = json.loads(rank.read_text())
        if data:
            latest = data[-1]
            ranked = len([r for r in latest.get("results", []) if r.get("rank")])
            rank_info = f"上榜词 {ranked} 个（更新于 {latest.get('date','?')}）"

    art_icon = "✅" if articles > 0 else "🟡"
    rank_icon = "✅" if "上榜词" in rank_info else "⏳"

    msg = f"""🔍 员工2号｜SEO专家日报 {t}
总状态：{"🟢 有产出" if articles > 0 else "🟡 今日无文章"}

**📊 今日执行**
| 任务 | 状态 | 数据 |
|------|------|------|
| 文章发布 | {art_icon} | {articles} 篇（累计4篇） |
| 关键词排名 | {rank_icon} | {rank_info} |
| SSR技术对齐 | ⏳ | 规格文档已发，等技术确认 |
| 外链建设 | 🔴 | 未启动 |

**📋 待执行**
| 优先级 | 事项 |
|--------|------|
| 🟠 高 | Article #5/#6 生产 |
| 🟠 高 | 推动技术团队 SSR 排期确认 |
| 🟡 中 | 进榜单外联邮件跟进 |"""
    return msg


# ── 员工3：KOL & 媒体 ──────────────────────────────────────────────────────

def report_kol() -> str:
    t = today()
    import openpyxl

    kol_dir = BASE / "03_kol_media"

    # ── 今日收集：读最新 YouTube 名单，统计今天新增行 ──
    yt_today = 0
    yt_total = 0
    yt_excels = sorted(kol_dir.glob("MoonX_YouTube_KOL名单_*.xlsx"), reverse=True)
    if yt_excels:
        try:
            wb = openpyxl.load_workbook(yt_excels[0], read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row and row[0]:
                    yt_total += 1
                    if row[11] and str(row[11]).startswith(t):  # 收集日期列
                        yt_today += 1
            wb.close()
        except: pass

    kalshi_today = 0
    kalshi_total = 0
    kalshi_excels = sorted(kol_dir.glob("MoonX_Kalshi_KOL名单_*.xlsx"), reverse=True)
    if kalshi_excels:
        try:
            wb = openpyxl.load_workbook(kalshi_excels[0], read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row and row[0] and row[1] != "KOL Demo" and row[10] != "无效":
                    kalshi_total += 1
                    if row[13] and str(row[13]).startswith(t):
                        kalshi_today += 1
            wb.close()
        except: pass

    all_total = yt_total + kalshi_total

    # ── 今日发送：读邮件发送日志 ──
    send_log = kol_dir / "logs" / "scheduled_send.log"
    email_sent_today = 0
    email_failed_today = 0
    email_skip_today = 0
    email_sent_total = 0
    if send_log.exists():
        for line in send_log.read_text(errors="ignore").split("\n"):
            if "✓ 已发送" in line:
                email_sent_total += 1
                if t in line:
                    email_sent_today += 1
            elif "✗ 失败" in line and t in line:
                email_failed_today += 1
        # 跳过数从 "跳过 N 条" 提取
        import re
        for line in reversed(send_log.read_text(errors="ignore").split("\n")):
            if t in line and "跳过" in line:
                m = re.search(r"跳过\s*(\d+)", line)
                if m:
                    email_skip_today = int(m.group(1))
                break

    # ── 分级统计（YouTube 名单） ──
    tiers = {"Kalshi": 0, "A级": 0, "B级": 0, "C级": 0, "D级": 0}
    tiers["Kalshi"] = kalshi_total
    if yt_excels:
        try:
            wb = openpyxl.load_workbook(yt_excels[0], read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row and row[0] and row[5]:
                    tier = str(row[5]).strip()
                    if tier in tiers:
                        tiers[tier] += 1
            wb.close()
        except: pass

    # 当前主攻：有效名单里C级优先
    if tiers["C级"] > 0:
        focus = f"C级（{tiers['C级']}个，1万~10万粉）"
    elif tiers["D级"] > 0:
        focus = f"D级（{tiers['D级']}个，1千~1万粉）"
    else:
        focus = "B级"

    # ── 自动化状态：读各脚本日志是否今天运行 ──
    def ran_today(log_name):
        log = kol_dir / "logs" / log_name
        if not log.exists():
            return False
        return t in log.read_text(errors="ignore")[-2000:]

    auto_yt     = "✅" if ran_today("youtube_kol.log")  else "❌"
    auto_kalshi = "✅" if ran_today("kalshi_kol.log")   else "❌"
    auto_enrich = "✅" if ran_today("kalshi_enrich.log") else "❌"
    auto_send   = "✅" if email_sent_today > 0          else ("❌" if t in (send_log.read_text(errors="ignore")[-500:] if send_log.exists() else "") else "⏳ 22:00待跑")

    # ── 待执行 ──
    pending = []
    if kalshi_total > 0 and tiers["Kalshi"] > 0:
        pending.append("🔴 Twitter DM权限待升级（Kalshi前KOL主力渠道）")
    pending.append("🟡 Substack渠道待接入")
    pending.append("🟡 官网contact页爬取（邮箱命中率提升）")
    pending_str = "\n".join(pending)

    send_status = "🟢" if email_sent_today >= 8 else ("🟡" if email_sent_today > 0 else "🔴")

    msg = f"""🤝 员工3号｜KOL媒体日报 {t}

**📥 今日收集**
| 渠道 | 今日新增 | 累计总量 |
|------|---------|---------|
| YouTube | +{yt_today} 个 | {yt_total} 个 |
| Kalshi专项 | +{kalshi_today} 个 | {kalshi_total} 个 |
| **合计** | **+{yt_today + kalshi_today} 个** | **{all_total} 个** |

**{send_status} 今日发送**
| 状态 | 数量 |
|------|------|
| 发出 | {email_sent_today} 封 |
| 失败 | {email_failed_today} 封 |
| 跳过（无邮箱） | {email_skip_today} 条 |
| 累计总发送 | {email_sent_total} 封 |

**📊 名单分级**
| 级别 | 数量 | 主攻 |
|------|------|------|
| Kalshi前KOL | {tiers['Kalshi']} 个 | 待Twitter DM开通 |
| C级（1万~10万） | {tiers['C级']} 个 | ← 当前主攻 |
| D级（1千~1万） | {tiers['D级']} 个 | 次优先 |
| B级（10万~100万） | {tiers['B级']} 个 | 数据内容合作 |
| A级（100万+） | {tiers['A级']} 个 | 暂缓 |

**🎯 当前主攻**：{focus}
邮件钩子：工具体验 + 返佣65%/50%/40%（行业标准3倍）

**⚙️ 自动化运行**
| 时间 | 任务 | 状态 |
|------|------|------|
| 10:00 | YouTube KOL收集 | {auto_yt} |
| 10:30 | Kalshi专项收集 | {auto_kalshi} |
| 11:00 | Twitter Handle补全 | {auto_enrich} |
| 22:00 | 邮件发送（上限10封） | {auto_send} |

**📋 待执行**
| 优先级 | 事项 |
|--------|------|
| 🔴 高 | Twitter DM权限升级（Kalshi前KOL主力渠道） |
| 🟡 中 | Substack渠道接入 |
| 🟡 中 | 官网contact页爬取（提升邮箱命中率）|"""
    return msg


# ── 员工4：增长运营 ──────────────────────────────────────────────────────────

def report_growth() -> str:
    t = today()
    msg = f"""📈 员工4号｜增长运营日报 {t}
总状态：🔴 待启动

**📊 模块状态**
| 模块 | 状态 | 进展 |
|------|------|------|
| 积分体系 | 🔴 未启动 | 方案未立项 |
| 大使招募 | 🔴 未启动 | 0 / 20 人 |
| 裂变机制 | 🔴 未启动 | — |
| 邮件召回 | 🔴 未启动 | — |

**📋 待执行**
| 优先级 | 事项 |
|--------|------|
| 🟠 高 | 积分体系方案设计（参考 Blur/Uniswap 模型） |
| 🟡 中 | 首批大使招募名单梳理 |"""
    return msg


# ── 员工5：策略数据 ──────────────────────────────────────────────────────────

def report_strategy() -> str:
    t = today()
    okr_file = BASE / "GMGN_市场OKR_Q1.xlsx"
    okr_status = "✅ 文件存在" if okr_file.exists() else "❌ 文件缺失"

    competitor_data = BASE / "01_social_media/logs/competitor_data.json"
    competitor_info = "暂无数据"
    if competitor_data.exists():
        try:
            data = json.loads(competitor_data.read_text())
            if data:
                latest = data[-1]
                competitor_info = f"追踪 {len(latest.get('data',[]))} 个账号（更新 {latest.get('date','?')}）"
        except:
            pass

    okr_icon = "✅" if "存在" in okr_status else "❌"
    comp_icon = "✅" if "追踪" in competitor_info else "⏳"

    msg = f"""📊 员工5号｜策略数据日报 {t}
总状态：🟡 部分运行

**📊 今日执行**
| 任务 | 状态 | 数据 |
|------|------|------|
| Q1 OKR追踪 | {okr_icon} | {okr_status} |
| 竞品监控 | {comp_icon} | {competitor_info} |
| 平台数据接入 | ⏳ | GSC待gsc_credentials.json |
| 周度复盘 | ⏳ | 待输出 |

**📋 待执行**
| 优先级 | 事项 |
|--------|------|
| 🟠 高 | Q1 OKR 完成率本周更新 |
| 🟡 中 | GSC接入（需 gsc_credentials.json） |
| 🟡 中 | 本周竞品动态整理输出 |"""
    return msg


# ── 卡片工具函数 ──────────────────────────────────────────────────────────────

DASH_URL    = "https://moonx-lark-server.fly.dev/dash"
TWITTER_URL = "https://twitter.com/moonx_bydfi"
KOL_SHEET   = f"https://docs.google.com/spreadsheets/d/{os.getenv('GOOGLE_SHEET_ID_KOL','')}"


def _btn(text: str, url: str, btn_type: str = "default") -> dict:
    return {"tag": "button", "text": {"tag": "plain_text", "content": text},
            "type": btn_type, "url": url}


def _actions(*buttons) -> dict:
    return {"tag": "action", "actions": list(buttons)}


def _col_stat(label: str, value: str, note: str = "") -> dict:
    """单个 stat block 列：大数字 + 标签 + 备注"""
    body = f"**{value}**\n{label}"
    if note:
        body += f"\n<font color='grey'>{note}</font>"
    return {
        "tag": "column",
        "width": "weighted",
        "weight": 1,
        "vertical_align": "center",
        "elements": [{"tag": "markdown", "content": body}],
    }


def _col_set(columns: list) -> dict:
    return {
        "tag": "column_set",
        "flex_mode": "stretch",
        "background_style": "grey",
        "columns": columns,
    }


def build_social_card() -> dict:
    t = today()
    tweet_path = BASE / "01_social_media/logs/tweet_history.json"
    tweets_today = 0
    if tweet_path.exists():
        try:
            tweets_today = len([x for x in json.loads(tweet_path.read_text()) if x.get("date") == t])
        except Exception: pass

    tweet_log = BASE / "01_social_media/logs/tweet.log"
    errors = 0
    if tweet_log.exists():
        lines = tweet_log.read_text(errors="ignore").split("\n")
        errors = sum(1 for l in lines[-100:] if "[ERROR]" in l and t in l)

    tg_log = BASE / "01_social_media/logs/daily_tweet_poster.log"
    tg_sent = 0
    if tg_log.exists():
        tg_sent = sum(1 for l in tg_log.read_text(errors="ignore").split("\n")[-50:] if "✅" in l and t in l)

    status_text = "🟢 正常运行" if tweets_today > 0 else "🔴 今日未发"
    err_color   = "red" if errors > 0 else "green"

    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"📱 社媒运营日报  {t}"},
            "template": "wathet",
        },
        "elements": [
            _md(f"总状态：{status_text}"),
            _col_set([
                _col_stat("Twitter推文", str(tweets_today), "今日已发条数"),
                _col_stat("TG推送", str(tg_sent), "今日推送次数"),
                _col_stat("脚本异常", f"<font color='{err_color}'>{errors}</font>", "ERROR日志数"),
            ]),
            _hr(),
            _md("**📅 定时任务**"),
            _tbl(
                [("time", "时间"), ("job", "任务"), ("status", "状态")],
                [
                    {"time": "09:00", "job": "trending 推文",     "status": "✅ 已执行" if tweets_today > 0 else "❌ 未执行"},
                    {"time": "14:00", "job": "smart-money 推文",  "status": "⏳ 待执行"},
                    {"time": "20:00", "job": "closing-soon 推文", "status": "⏳ 待执行"},
                    {"time": "08/14/20:05", "job": "TG 早中晚报", "status": "✅ 已执行" if tg_sent > 0 else "⏳ 待执行"},
                ]
            ),
            _hr(),
            _md("**📌 待执行**\n🟠 大V评论策略落地（5个账号/天）\n🟡 热点监控排入 09:00/14:00/20:00 时段"),
            _actions(
                _btn("查看推文 →", TWITTER_URL, "primary"),
                _btn("今日数据", DASH_URL),
            ),
        ],
    }


def build_seo_card() -> dict:
    t = today()
    pub = BASE / "02_seo/logs/published_articles.json"
    articles = 0
    if pub.exists():
        try:
            articles = len([x for x in json.loads(pub.read_text()) if x.get("date", "") == t])
        except Exception: pass

    rank_file = BASE / "02_seo/logs/rank_history.json"
    ranked_kw, rank_date = 0, "—"
    if rank_file.exists():
        try:
            data = json.loads(rank_file.read_text())
            if data:
                latest = data[-1]
                ranked_kw = len([r for r in latest.get("results", []) if r.get("rank")])
                rank_date = latest.get("date", "?")
        except Exception: pass

    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"🔍 SEO 专家日报  {t}"},
            "template": "turquoise",
        },
        "elements": [
            _md(f"总状态：{'🟢 有产出' if articles > 0 else '🟡 今日无文章'}"),
            _col_set([
                _col_stat("今日文章", str(articles), "累计 4 篇"),
                _col_stat("上榜关键词", str(ranked_kw), f"更新 {rank_date}"),
                _col_stat("外链建设", "🔴 未启动", "待规划"),
            ]),
            _hr(),
            _md("**📊 任务状态**"),
            _tbl(
                [("task", "任务"), ("status", "状态"), ("note", "备注")],
                [
                    {"task": "文章生产",    "status": "✅ 有产出" if articles > 0 else "🟡 今日无", "note": f"{articles} 篇"},
                    {"task": "关键词追踪",  "status": "🔄 运行中", "note": f"上榜 {ranked_kw} 词 ({rank_date})"},
                    {"task": "SSR 技术对齐","status": "⏳ 待排期", "note": "规格文档已发，等技术确认"},
                    {"task": "外链建设",    "status": "🔴 未启动", "note": "—"},
                ]
            ),
            _hr(),
            _md("**📌 待执行**\n🟠 Article #5/#6 生产\n🟠 推动技术团队 SSR 排期确认\n🟡 进榜单外联邮件跟进"),
            _actions(
                _btn("查看排名数据", DASH_URL, "primary"),
                _btn("今日数据", DASH_URL),
            ),
        ],
    }


def build_kol_card() -> dict:
    t = today()
    import openpyxl

    kol_dir = BASE / "03_kol_media"
    yt_today, yt_total = 0, 0
    yt_excels = sorted(kol_dir.glob("MoonX_YouTube_KOL名单_*.xlsx"), reverse=True)
    if yt_excels:
        try:
            wb = openpyxl.load_workbook(yt_excels[0], read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row and row[0]:
                    yt_total += 1
                    if row[11] and str(row[11]).startswith(t):
                        yt_today += 1
            wb.close()
        except Exception: pass

    send_log = kol_dir / "logs" / "scheduled_send.log"
    email_sent_today, email_sent_total = 0, 0
    if send_log.exists():
        for line in send_log.read_text(errors="ignore").split("\n"):
            if "✓ 已发送" in line:
                email_sent_total += 1
                if t in line:
                    email_sent_today += 1

    send_status = "🟢" if email_sent_today >= 8 else ("🟡" if email_sent_today > 0 else "🔴")

    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"🤝 KOL 媒体日报  {t}"},
            "template": "green",
        },
        "elements": [
            _md(f"邮件状态：{send_status}  今日发送 {email_sent_today} 封"),
            _col_set([
                _col_stat("今日新增 KOL", f"+{yt_today}", "YouTube 渠道"),
                _col_stat("名单累计",     str(yt_total), "有效 KOL 总数"),
                _col_stat("今日邮件",     str(email_sent_today), f"累计 {email_sent_total} 封"),
            ]),
            _hr(),
            _md("**📊 自动化状态**"),
            _tbl(
                [("time", "时间"), ("job", "任务"), ("status", "状态")],
                [
                    {"time": "10:00", "job": "YouTube KOL 收集",  "status": "✅ 已跑" if yt_today > 0 else "⏳"},
                    {"time": "22:00", "job": "邮件发送（上限10封）","status": "✅ 已发" if email_sent_today > 0 else "⏳ 待跑"},
                    {"time": "常态",  "job": "Twitter DM",         "status": "❌ 401 权限受阻"},
                ]
            ),
            _hr(),
            _md("**📌 待执行**\n🔴 Twitter DM 权限升级（developer.twitter.com）\n🟡 Substack 渠道接入\n🟡 官网 contact 页爬取（提升邮箱命中率）"),
            _actions(
                _btn("查看 KOL 名单 →", KOL_SHEET, "primary"),
                _btn("今日数据", DASH_URL),
            ),
        ],
    }


def build_growth_card() -> dict:
    t = today()
    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"📈 增长运营日报  {t}"},
            "template": "red",
        },
        "elements": [
            _md("总状态：🔴 全模块待启动"),
            _col_set([
                _col_stat("积分体系", "🔴", "方案未立项"),
                _col_stat("大使招募", "0 / 20", "目标 20 人"),
                _col_stat("裂变机制", "🔴", "未启动"),
            ]),
            _hr(),
            _md("**📊 模块状态**"),
            _tbl(
                [("module", "模块"), ("status", "状态"), ("note", "进展")],
                [
                    {"module": "积分体系", "status": "🔴 未启动", "note": "方案未立项"},
                    {"module": "大使招募", "status": "🔴 未启动", "note": "0 / 20 人"},
                    {"module": "裂变机制", "status": "🔴 未启动", "note": "—"},
                    {"module": "邮件召回", "status": "🔴 未启动", "note": "—"},
                ]
            ),
            _hr(),
            _md("**📌 待执行**\n🟠 积分体系方案设计（参考 Blur/Uniswap 模型）\n🟡 首批大使招募名单梳理"),
            _actions(
                _btn("今日数据", DASH_URL),
            ),
        ],
    }


def build_strategy_card() -> dict:
    t = today()
    okr_file = BASE / "GMGN_市场OKR_Q1.xlsx"
    okr_ok   = okr_file.exists()

    competitor_data = BASE / "01_social_media/logs/competitor_data.json"
    comp_n, comp_date = 0, "—"
    if competitor_data.exists():
        try:
            data = json.loads(competitor_data.read_text())
            if data:
                latest = data[-1]
                comp_n    = len(latest.get("data", []))
                comp_date = latest.get("date", "?")
        except Exception: pass

    return {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"📊 策略数据日报  {t}"},
            "template": "indigo",
        },
        "elements": [
            _md(f"总状态：{'🟢 部分运行' if okr_ok else '🟡 部分缺失'}"),
            _col_set([
                _col_stat("Q1 OKR", "✅ 就绪" if okr_ok else "❌ 缺失", "GMGN_市场OKR_Q1.xlsx"),
                _col_stat("竞品追踪", str(comp_n), f"账号数 ({comp_date})"),
                _col_stat("GSC 接入", "⏳", "待 credentials.json"),
            ]),
            _hr(),
            _md("**📊 任务状态**"),
            _tbl(
                [("task", "任务"), ("status", "状态"), ("note", "备注")],
                [
                    {"task": "Q1 OKR 追踪",  "status": "✅ 运行中" if okr_ok else "❌ 文件缺失", "note": "完成率本周待更新"},
                    {"task": "竞品监控",      "status": "✅ 运行中" if comp_n > 0 else "⏳",       "note": f"追踪 {comp_n} 账号"},
                    {"task": "平台数据接入",  "status": "⏳ 待配置", "note": "GSC 待配置"},
                    {"task": "周度复盘",      "status": "⏳ 待输出", "note": "—"},
                ]
            ),
            _hr(),
            _md("**📌 待执行**\n🟠 Q1 OKR 完成率本周更新\n🟡 GSC 接入（需 gsc_credentials.json）\n🟡 本周竞品动态整理输出"),
            _actions(
                _btn("今日数据", DASH_URL, "primary"),
            ),
        ],
    }


def send_card(bot: str, card: dict) -> bool:
    url = WEBHOOKS.get(bot)
    if not url:
        return False
    payload = json.dumps({"msg_type": "interactive", "card": card}, ensure_ascii=False).encode()
    req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"})
    try:
        res = json.loads(urllib.request.urlopen(req, timeout=10).read())
        ok = res.get("code") == 0
        if not ok:
            print(f"  card error ({bot}): {res.get('msg','')}")
        return ok
    except Exception as e:
        print(f"  card error ({bot}): {e}")
        return False


def _md(content: str) -> dict:
    return {"tag": "markdown", "content": content}

def _hr() -> dict:
    return {"tag": "hr"}

def _tbl(cols: list, rows: list) -> dict:
    """
    cols: [(name, display_name), ...]
    rows: [{name: str, ...}, ...]
    """
    return {
        "tag": "table",
        "page_size": 20,
        "row_height": "low",
        "header_style": {
            "text_align": "left",
            "background_color": "grey",
            "bold": True,
        },
        "columns": [
            {"tag": "column", "name": n, "display_name": d,
             "data_type": "text", "width": "auto", "horizontal_align": "left"}
            for n, d in cols
        ],
        "rows": [
            {k: str(v) for k, v in row.items()}
            for row in rows
        ],
    }

def _dept_section(title: str, rows: list, todos: list) -> list:
    """返回一个部门的卡片 elements 列表：标题 + 表格 + 待做"""
    elems = [_md(f"**{title}**")]
    elems.append(_tbl(
        [("item", "工作项"), ("status", "状态"), ("result", "今日结果")],
        [{"item": r[0], "status": r[1], "result": r[2]} for r in rows]
    ))
    if todos:
        todo_text = "　".join(
            f"{'🔴' if i == 0 else '🟡'} {td}" for i, td in enumerate(todos)
        )
        elems.append(_md(f"📌 待做：{todo_text}"))
    elems.append(_hr())
    return elems


# ── Team Lead：卡片日报 ────────────────────────────────────────────────────────

def build_lead_cards() -> tuple:
    """返回 (report_card, exception_card) 两张 Lark 互动卡片"""
    t   = today()
    now = datetime.now(BJT).strftime("%H:%M")

    ctx = {}
    ctx_file = BASE / "loop_context.json"
    if ctx_file.exists():
        try:
            ctx = json.loads(ctx_file.read_text())
        except Exception: pass

    # ── 读取各部门数据 ──────────────────────────────────────────────────────────

    # 社媒
    tweet_path = BASE / "01_social_media/logs/tweet_history.json"
    tweets_today = 0
    if tweet_path.exists():
        try:
            tweets_today = len([x for x in json.loads(tweet_path.read_text()) if x.get("date") == t])
        except Exception: pass
    tg_log = BASE / "01_social_media/logs/daily_tweet_poster.log"
    tg_sent = 0
    if tg_log.exists():
        tg_sent = sum(1 for l in tg_log.read_text(errors="ignore").split("\n")[-50:] if "✅" in l and t in l)
    tweet_log = BASE / "01_social_media/logs/tweet.log"
    tw_errors = 0
    if tweet_log.exists():
        tw_errors = sum(1 for l in tweet_log.read_text(errors="ignore").split("\n")[-100:] if "[ERROR]" in l and t in l)
    social_ok = tweets_today > 0

    # SEO
    pub = BASE / "02_seo/logs/published_articles.json"
    articles_n = 0
    if pub.exists():
        try:
            articles_n = len([x for x in json.loads(pub.read_text()) if x.get("date", "") == t])
        except Exception: pass
    rank_file = BASE / "02_seo/logs/rank_history.json"
    rank = "暂无数据"
    if rank_file.exists():
        try:
            data = json.loads(rank_file.read_text())
            if data:
                latest = data[-1]
                ranked = len([r for r in latest.get("results", []) if r.get("rank")])
                rank = f"上榜词 {ranked} 个（{latest.get('date','?')}）"
        except Exception: pass

    # KOL
    email_sent = 0
    email_sent_total = 0
    send_log = BASE / "03_kol_media/logs/scheduled_send.log"
    if send_log.exists():
        for line in send_log.read_text(errors="ignore").split("\n"):
            if "✓ 已发送" in line:
                email_sent_total += 1
                if t in line:
                    email_sent += 1
    kol_dir = BASE / "03_kol_media"
    yt_excels = sorted(kol_dir.glob("MoonX_YouTube_KOL名单_*.xlsx"), reverse=True)
    yt_total_n = 0
    if yt_excels:
        try:
            import openpyxl as _ox
            wb = _ox.load_workbook(yt_excels[0], read_only=True)
            yt_total_n = sum(1 for row in wb.active.iter_rows(min_row=4, values_only=True) if row and row[0])
            wb.close()
        except Exception: pass

    # 策略
    okr_file = BASE / "GMGN_市场OKR_Q1.xlsx"
    okr_ok = okr_file.exists()
    okr_status = "✅ 已就绪" if okr_ok else "❌ 文件缺失"
    competitor_data = BASE / "01_social_media/logs/competitor_data.json"
    competitor = "暂无数据"
    if competitor_data.exists():
        try:
            data = json.loads(competitor_data.read_text())
            if data:
                latest = data[-1]
                competitor = f"追踪 {len(latest.get('data',[]))} 个账号（{latest.get('date','?')}）"
        except Exception: pass

    # ── 卡片1：全团队日报 ───────────────────────────────────────────────────────
    elements = []

    # 全局概览
    elements.append(_md("**🗂️ 全局概览**"))
    elements.append(_tbl(
        [("dept", "部门"), ("metrics", "核心指标"), ("st", "状态"), ("top", "最高优先事项")],
        [
            {"dept": "📱 社媒", "metrics": f"推文 {tweets_today} 条 / TG {tg_sent} 次",
             "st": "🟢" if social_ok else "🔴", "top": "大V评论策略落地"},
            {"dept": "🔍 SEO", "metrics": f"文章 {articles_n} 篇 / {rank}",
             "st": "🟢" if articles_n > 0 else "🟡", "top": "Article #5/#6 生产"},
            {"dept": "🤝 KOL", "metrics": f"今日发 {email_sent} 封 / 名单 {yt_total_n} 个",
             "st": "🟢" if email_sent > 0 else "🟡", "top": "Twitter DM 权限升级"},
            {"dept": "📈 增长", "metrics": "积分/大使/裂变 全未启动",
             "st": "🔴", "top": "积分体系方案立项"},
            {"dept": "📊 策略", "metrics": f"{okr_status} / {competitor}",
             "st": "🟢" if okr_ok else "🟡", "top": "Q1 OKR 完成率更新"},
        ]
    ))
    elements.append(_hr())

    # 各部门详情
    elements += _dept_section("📱 社媒运营", [
        ("Twitter推文", "✅ 已执行" if social_ok else "❌ 未执行", f"{tweets_today} 条"),
        ("TG推送",      "✅ 已执行" if tg_sent > 0 else "⏳ 待确认", f"{tg_sent} 次"),
        ("大V评论",     "⏳ 待建立", "0条 / 每日5条目标"),
        ("热点监控",    "⏳ 待建立", "未排入日程"),
    ], ["大V评论策略落地（5个账号/天）", "热点监控排入 09:00/14:00/20:00"])

    elements += _dept_section("🔍 SEO专家", [
        ("文章生产",    "✅ 有产出" if articles_n > 0 else "🟡 今日无文章", f"{articles_n} 篇（累计4篇）"),
        ("关键词追踪",  "🔄 运行中", rank),
        ("SSR技术对齐", "⏳ 待排期", "规格文档已发，等技术确认"),
        ("外链建设",    "⏳ 待启动", "—"),
    ], ["Article #5/#6 生产", "推动技术团队 SSR 排期确认", "进榜单外联邮件跟进"])

    elements += _dept_section("🤝 KOL媒体", [
        ("YouTube收集", "✅ 每日运行", f"名单 {yt_total_n} 个，今日自动收集"),
        ("邮件外联",    "✅ 已发送" if email_sent > 0 else "⏳ 22:00待跑", f"今日 {email_sent} 封 / 累计 {email_sent_total} 封"),
        ("Twitter DM",  "❌ 权限受阻", "401错误，待开通 Read+Write+DM 权限"),
        ("媒体报道",    "🔄 跟进中",  "13封已发，待回复"),
    ], ["Twitter DM权限升级（developer.twitter.com）", "C级KOL回复跟进，推动首批合作确认", "Substack渠道接入"])

    elements += _dept_section("📈 增长运营", [
        ("积分体系", "🔴 未启动", "方案未立项"),
        ("大使招募", "🔴 未启动", "0 / 20人"),
        ("裂变机制", "🔴 未启动", "—"),
        ("邮件召回", "🔴 未启动", "—"),
    ], ["积分体系方案设计（参考 Blur/Uniswap 模型）", "首批大使招募名单梳理"])

    elements += _dept_section("📊 策略数据", [
        ("竞品监控",    "✅ 运行中", competitor),
        ("Q1 OKR追踪",  "✅ 运行中" if okr_ok else "❌ 文件缺失", okr_status),
        ("平台数据接入","⏳ 待配置", "GSC待gsc_credentials.json"),
        ("周度复盘",    "⏳ 待输出", "—"),
    ], ["Q1 OKR 完成率本周更新", "GSC接入（需 gsc_credentials.json）", "本周竞品动态整理输出"])

    # 决策
    if not ctx.get("today_decided"):
        option_a = ctx.get("today_option_a", "")
        option_b = ctx.get("today_option_b", "")
        if option_a and option_b:
            elements.append(_md(f"**需要 Kelly 决策**\nA：{option_a}\nB：{option_b}\n回复「我选A」或「我选B」"))
        # 无待决策则不显示该 block
    else:
        elements.append(_md(f"✅ 今日已执行方案 {ctx.get('today_choice', '')}"))

    # Lark 单卡片最多 3 个 table，拆成上下两张
    # 上半：全局概览 + 社媒 + SEO（3 tables）
    # 下半：KOL + 增长 + 策略（3 tables）
    split = next(
        i for i, el in enumerate(elements)
        if el.get("tag") == "markdown" and "🤝" in el.get("content", "")
    )
    elements_top    = elements[:split]
    elements_bottom = elements[split:]

    # Lead 卡片顶部加全局 stat 总览
    all_ok   = sum([social_ok, articles_n > 0, email_sent > 0, okr_ok])
    health   = ["🔴", "🟡", "🟡", "🟢", "🟢"][all_ok]
    elements_top = [
        _col_set([
            _col_stat("社媒推文", str(tweets_today), "今日"),
            _col_stat("SEO文章",  str(articles_n),   "今日"),
            _col_stat("KOL邮件",  str(email_sent),    "今日发出"),
            _col_stat("团队健康", health,              f"{all_ok}/4 正常"),
        ]),
        _hr(),
    ] + elements_top

    elements_bottom = elements_bottom + [
        _hr(),
        _actions(
            _btn("查看实时数据 →", DASH_URL, "primary"),
            _btn("查看推文",       TWITTER_URL),
            _btn("KOL 名单",       KOL_SHEET),
        ),
    ]

    report_card = {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"👔 全团队日报  {t}  {now} BJT（上）"},
            "template": "blue",
        },
        "elements": elements_top,
    }
    report_card2 = {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": f"👔 全团队日报  {t}（下）"},
            "template": "blue",
        },
        "elements": elements_bottom,
    }

    # ── 卡片2：异常提示 ─────────────────────────────────────────────────────────
    executions = ctx.get("today_executions", [])
    failed = [e for e in executions if "❌" in e.get("status", "") or "⚠️" in e.get("status", "")]
    now_hour = datetime.now(BJT).hour
    auto_alerts = []
    if now_hour >= 10 and not social_ok:
        auto_alerts.append(f"📱 社媒：09:00 trending推文未执行")
    if now_hour >= 23 and email_sent == 0:
        auto_alerts.append("🤝 KOL：22:00 邮件外联未执行（0封）")
    if now_hour >= 19 and tw_errors > 0:
        auto_alerts.append(f"📱 社媒：脚本异常 {tw_errors} 个")

    exc_elements = []
    if failed or auto_alerts:
        if failed:
            exc_elements.append(_md("**脚本执行失败**"))
            for e in failed:
                detail = f"\n　└ {e['detail'][:80]}" if e.get("detail") else ""
                exc_elements.append(_md(f"• {e['time']} {e['job']}: {e['status']}{detail}"))
        if auto_alerts:
            exc_elements.append(_md("**自动化任务未执行**"))
            for a in auto_alerts:
                exc_elements.append(_md(f"• {a}"))
        exc_template, exc_title = "red", f"⚠️ 异常提示  {t}"
    else:
        exc_elements.append(_md("所有脚本执行正常，无待处理异常。"))
        exc_template, exc_title = "green", f"✅ 今日无异常  {t}"

    exception_card = {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": exc_title},
            "template": exc_template,
        },
        "elements": exc_elements,
    }

    return report_card, report_card2, exception_card


# ── 主函数 ────────────────────────────────────────────────────────────────────

def run():
    print(f"🚀 Lark 日报发送中 — {today()}")

    # 各部门卡片
    dept_cards = [
        ("social",   "社媒",   build_social_card),
        ("seo",      "SEO",    build_seo_card),
        ("kol",      "KOL",    build_kol_card),
        ("growth",   "增长",   build_growth_card),
        ("strategy", "策略",   build_strategy_card),
    ]
    for bot, name, builder in dept_cards:
        try:
            ok = send_card(bot, builder())
            print(f"  {'✅' if ok else '❌'} {name} 日报")
        except Exception as e:
            print(f"  ❌ {name} 日报构建失败: {e}")

    # Lead 汇总卡片
    card1, card2, exc_card = build_lead_cards()
    ok1 = send_card("lead", card1)
    print(f"  {'✅' if ok1 else '❌'} lead (日报上)")
    ok2 = send_card("lead", card2)
    print(f"  {'✅' if ok2 else '❌'} lead (日报下)")
    ok3 = send_card("lead", exc_card)
    print(f"  {'✅' if ok3 else '❌'} lead (异常)")


if __name__ == "__main__":
    import sys as _sys; _sys.path.insert(0, str(Path(__file__).parent))
    from scripts.pid_lock import acquire_lock; acquire_lock()
    run()
