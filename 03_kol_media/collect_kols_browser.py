#!/usr/bin/env python3
"""
BYDFi MoonX — 浏览器爬取 Twitter KOL（Playwright + Chrome CDP）
不使用 Twitter API，直接走浏览器页面，绕过 API 限制
目标：收集 Meme币 / 加密货币 KOL，提取邮箱，无邮箱的记录待 DM

启动方式：
  1. Chrome 以调试模式运行：已在 port 9222
  2. python3 collect_kols_browser.py
"""

import re, time, json, logging
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

TODAY      = datetime.now().strftime("%Y-%m-%d")
SCRIPT_DIR = Path(__file__).parent
LOG_DIR    = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "browser_kol.log"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

CDP_URL = "http://localhost:9222"

# ── 搜索关键词 ──
SEARCH_QUERIES = [
    "meme coin calls",
    "solana meme alpha",
    "pump fun trading",
    "GMGN crypto",
    "dexscreener gems",
    "on-chain alpha crypto",
    "meme coin 100x",
    "crypto KOL influencer",
    "bnb chain meme",
    "altcoin gems 2026",
]

# 每个关键词最多抓多少个 profile
MAX_PER_QUERY = 20

# 最低粉丝数
MIN_FOLLOWERS = 1000

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
EMAIL_BLOCKLIST = {
    "example.com", "domain.com", "email.com", "twitter.com", "t.co",
    "sentry.io", "cloudflare.com", "wixpress.com", "squarespace.com",
    "gmail.com", "outlook.com",  # 通用邮箱，不是业务邮箱
}
EMAIL_EXACT_BLOCKLIST = {
    "johnappleseed@gmail.com", "example@gmail.com", "test@gmail.com",
}


def extract_email(text: str) -> str:
    if not text:
        return ""
    normalized = re.sub(r'\s*\(at\)\s*', '@', text, flags=re.IGNORECASE)
    normalized = re.sub(r'\s*\(dot\)\s*', '.', normalized, flags=re.IGNORECASE)
    for email in EMAIL_RE.findall(normalized):
        domain = email.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST and "." in domain and email.lower() not in EMAIL_EXACT_BLOCKLIST:
            return email.lower()
    return ""


def parse_follower_count(text: str) -> int:
    """将 '12.3K'、'1.2M' 等转为整数。"""
    if not text:
        return 0
    text = text.replace(",", "").strip()
    try:
        if text.endswith("K"):
            return int(float(text[:-1]) * 1000)
        elif text.endswith("M"):
            return int(float(text[:-1]) * 1_000_000)
        return int(text)
    except Exception:
        return 0


def classify_tier(followers: int) -> str:
    if followers >= 1_000_000:
        return "A级"
    elif followers >= 100_000:
        return "B级"
    elif followers >= 10_000:
        return "C级"
    return "D级"


def load_existing_handles() -> set:
    """加载所有已收集的 handle，用于去重。"""
    seen = set()
    for fp in SCRIPT_DIR.glob("MoonX_KOL名单_*.xlsx"):
        try:
            wb = load_workbook(fp, read_only=True)
            for row in wb.active.iter_rows(min_row=3, values_only=True):
                if row and row[1]:
                    seen.add(str(row[1]).lstrip("@").lower())
            wb.close()
        except Exception:
            pass
    for fp in SCRIPT_DIR.glob("MoonX_Browser_KOL名单_*.xlsx"):
        try:
            wb = load_workbook(fp, read_only=True)
            for row in wb.active.iter_rows(min_row=3, values_only=True):
                if row and row[1]:
                    seen.add(str(row[1]).lstrip("@").lower())
            wb.close()
        except Exception:
            pass
    return seen


def scrape_profile(page, handle: str):
    """访问 twitter.com/@handle，提取 profile 信息。"""
    url = f"https://x.com/{handle}"
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(2000)

        # 名字
        name = ""
        try:
            name_el = page.query_selector('[data-testid="UserName"] span span')
            if name_el:
                name = name_el.inner_text().strip()
        except Exception:
            pass

        # 粉丝数
        followers = 0
        try:
            # 找 followers 链接
            followers_el = page.query_selector('a[href$="/followers"] span span')
            if followers_el:
                followers = parse_follower_count(followers_el.inner_text())
        except Exception:
            pass

        # Bio
        bio = ""
        try:
            bio_el = page.query_selector('[data-testid="UserDescription"]')
            if bio_el:
                bio = bio_el.inner_text().strip()
        except Exception:
            pass

        # 外链 URL
        website = ""
        try:
            link_el = page.query_selector('[data-testid="UserUrl"] a')
            if link_el:
                website = link_el.get_attribute("href") or ""
        except Exception:
            pass

        # 先从 bio 提取邮箱
        email = extract_email(bio)

        # bio 没有的话，尝试访问 website
        if not email and website and "twitter.com" not in website and "x.com" not in website:
            try:
                wp = page.context.new_page()
                wp.goto(website, wait_until="domcontentloaded", timeout=10000)
                wp.wait_for_timeout(1500)
                page_text = wp.inner_text("body") if wp.query_selector("body") else ""
                email = extract_email(page_text)
                wp.close()
            except Exception:
                pass

        if followers < MIN_FOLLOWERS:
            return None

        return {
            "handle":    "@" + handle.lstrip("@"),
            "name":      name or handle,
            "followers": followers,
            "tier":      classify_tier(followers),
            "bio":       bio[:200],
            "website":   website,
            "email":     email,
            "date":      TODAY,
        }
    except Exception as e:
        logger.warning(f"抓取 @{handle} 失败: {e}")
        return None


def search_handles_on_twitter(page, query: str, max_results: int = 20) -> list[str]:
    """在 Twitter 搜索页抓出 handle 列表。"""
    handles = []
    try:
        encoded = query.replace(" ", "%20")
        page.goto(
            f"https://x.com/search?q={encoded}&src=typed_query&f=user",
            wait_until="domcontentloaded",
            timeout=15000,
        )
        page.wait_for_timeout(3000)

        # 滚动加载更多结果
        for _ in range(3):
            page.keyboard.press("End")
            page.wait_for_timeout(1500)

        # 提取用户卡片中的 handle
        cells = page.query_selector_all('[data-testid="UserCell"]')
        for cell in cells[:max_results]:
            try:
                link = cell.query_selector('a[href^="/"]')
                if link:
                    href = link.get_attribute("href") or ""
                    handle = href.strip("/").split("/")[0]
                    if handle and not handle.startswith("i/") and handle not in ("home", "explore", "notifications"):
                        handles.append(handle)
            except Exception:
                pass

        # 去重保序
        seen = set()
        unique = []
        for h in handles:
            if h.lower() not in seen:
                seen.add(h.lower())
                unique.append(h)
        return unique

    except Exception as e:
        logger.warning(f"搜索 '{query}' 失败: {e}")
        return []


def build_excel(kols: list[dict], output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Twitter KOL（浏览器）"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill  = PatternFill("solid", fgColor="2E75B6")
    title_font  = Font(bold=True, color="FFFFFF", size=13)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"BYDFi MoonX — Twitter KOL 名单（浏览器爬取，{TODAY}）"
    c.fill, c.font, c.alignment = title_fill, title_font, center
    ws.row_dimensions[1].height = 28

    headers   = ["Twitter Handle", "名字", "粉丝数", "分级", "Bio摘要", "官网", "邮箱", "联系状态", "备注", "收集日期"]
    col_widths = [22, 20, 12, 8, 50, 35, 35, 12, 15, 14]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, thin
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w
    ws.row_dimensions[2].height = 22

    tier_colors = {"A级": "FFD700", "B级": "C6EFCE", "C级": "DDEBF7", "D级": "F2F2F2"}
    for i, kol in enumerate(kols, 3):
        row_data = [
            kol["handle"], kol["name"], kol["followers"], kol["tier"],
            kol["bio"], kol["website"], kol["email"],
            "", "", kol["date"],
        ]
        fill = PatternFill("solid", fgColor=tier_colors.get(kol["tier"], "FFFFFF"))
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill, cell.border = fill, thin
            cell.alignment = Alignment(vertical="center", wrap_text=(j == 5))
        ws.row_dimensions[i].height = 18

    ws.freeze_panes = "A3"
    wb.save(output_path)
    logger.info(f"✅ 已保存: {output_path.name}（{len(kols)} 条）")


def run():
    existing = load_existing_handles()
    logger.info(f"{'='*60}")
    logger.info(f"🐦 Twitter 浏览器爬取开始 — {TODAY}")
    logger.info(f"   已有记录: {len(existing)} 个 handle（自动去重）")
    logger.info(f"{'='*60}")

    all_kols = []

    with sync_playwright() as pw:
        # 连接到已运行的 Chrome
        try:
            browser = pw.chromium.connect_over_cdp(CDP_URL)
            logger.info(f"✅ 已连接 Chrome (CDP: {CDP_URL})")
        except Exception as e:
            logger.error(f"❌ 无法连接 Chrome: {e}")
            logger.error("   请确保 Chrome 以 --remote-debugging-port=9222 启动")
            return

        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = context.new_page()

        # 确认是否已登录 Twitter
        page.goto("https://x.com/home", wait_until="domcontentloaded", timeout=15000)
        page.wait_for_timeout(2000)
        if "login" in page.url or page.query_selector('[data-testid="loginButton"]'):
            logger.warning("⚠️  Twitter 未登录！请在浏览器中手动登录后重新运行。")
            page.close()
            return
        logger.info("✅ Twitter 已登录")

        # 逐个关键词搜索
        discovered_handles: list[str] = []
        for query in SEARCH_QUERIES:
            logger.info(f"\n🔍 搜索用户: {query}")
            handles = search_handles_on_twitter(page, query, MAX_PER_QUERY)
            new_handles = [h for h in handles if h.lower() not in existing and h.lower() not in [x.lower() for x in discovered_handles]]
            logger.info(f"   发现 {len(handles)} 个，新增 {len(new_handles)} 个")
            discovered_handles.extend(new_handles)
            time.sleep(1)

        # 去重
        seen = set()
        unique_handles = []
        for h in discovered_handles:
            if h.lower() not in seen:
                seen.add(h.lower())
                unique_handles.append(h)

        logger.info(f"\n📊 待抓取 profile: {len(unique_handles)} 个")

        # 逐个访问 profile
        for i, handle in enumerate(unique_handles, 1):
            logger.info(f"[{i}/{len(unique_handles)}] 抓取 @{handle} ...")
            kol = scrape_profile(page, handle)
            if kol:
                all_kols.append(kol)
                email_info = f"邮箱: {kol['email']}" if kol["email"] else "无邮箱"
                logger.info(f"   ✓ {kol['name']} | {kol['followers']:,}粉 | {kol['tier']} | {email_info}")
            else:
                logger.info(f"   ✗ 跳过（粉丝不足或抓取失败）")
            time.sleep(1.5)  # 避免触发限速

        page.close()

    if not all_kols:
        logger.info("⚠️  没有新 KOL，结束")
        return

    all_kols.sort(key=lambda x: x["followers"], reverse=True)

    # 统计
    tiers = {"A级": 0, "B级": 0, "C级": 0, "D级": 0}
    with_email, no_email = 0, 0
    for kol in all_kols:
        tiers[kol["tier"]] += 1
        if kol["email"]: with_email += 1
        else: no_email += 1

    logger.info(f"\n📈 结果统计:")
    for t, n in tiers.items():
        logger.info(f"   {t}: {n} 个")
    logger.info(f"   有邮箱: {with_email} 个")
    logger.info(f"   无邮箱（待发DM）: {no_email} 个")

    # 保存 Excel
    out = SCRIPT_DIR / f"MoonX_Browser_KOL名单_{TODAY}.xlsx"
    build_excel(all_kols, out)

    # 输出待 DM 名单
    no_email_list = [k for k in all_kols if not k["email"]]
    if no_email_list:
        dm_out = SCRIPT_DIR / f"MoonX_待发DM_{TODAY}.json"
        with open(dm_out, "w") as f:
            json.dump([{"handle": k["handle"], "name": k["name"], "followers": k["followers"]} for k in no_email_list], f, ensure_ascii=False, indent=2)
        logger.info(f"📋 待发DM名单已保存: {dm_out.name}")

    logger.info(f"\n{'='*60}")
    logger.info(f"完成！共收集 {len(all_kols)} 个 Twitter KOL（浏览器爬取）")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    run()
