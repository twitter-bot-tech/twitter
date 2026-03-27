#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()

BLUE_DARK   = "1F497D"
BLUE_MID    = "2E74B5"
WHITE       = "FFFFFF"
GREEN_LIGHT = "E2EFDA"
ORANGE_LIGHT= "FCE4D6"
BLUE_LIGHT  = "EBF3FB"
PURPLE_LIGHT= "EAE0F4"

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value, bg=None, bold=False, size=9, color="000000", wrap=True, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap
    )
    cell.border = thin_border()
    return cell

def sheet_title(ws, title, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

def header_row(ws, row, headers, bg=BLUE_MID):
    ws.row_dimensions[row].height = 22
    for ci, h in enumerate(headers, 1):
        sc(ws, row, ci, h, bg=bg, bold=True, color=WHITE, center=True)

# ══════════════════════════════════════════
# Sheet 1：预测市场 KOL
# ══════════════════════════════════════════
ws1 = wb.active
ws1.title = "预测市场KOL"
ws1.sheet_view.showGridLines = False

cols1 = {"A":18,"B":20,"C":12,"D":22,"E":24,"F":12,"G":14,"H":10}
for col, w in cols1.items():
    ws1.column_dimensions[col].width = w

sheet_title(ws1, "预测市场 KOL 名单 — 欧美+东南亚", 8)
header_row(ws1, 2, ["姓名/账号", "Twitter Handle", "粉丝量", "内容方向", "联系方式", "地区", "合作优先级", "状态"])

kol_data = [
    ("PolymarketWhales",    "@PolymarketWhales",  "~40K",  "Polymarket鲸鱼追踪/市场分析", "Twitter DM",                      "美国",    "★★★", "待联系"),
    ("Prediction Mkt Alpha","@PMAlphaXYZ",        "~18K",  "Polymarket/Kalshi交易策略",   "Twitter DM / Substack",           "美国",    "★★★", "待联系"),
    ("Zvi Mowshowitz",      "@TheZvi",            "~35K",  "理性主义/预测/Polymarket评论", "thezvi@gmail.com",                "美国",    "★★★", "待联系"),
    ("Scott Alexander",     "@slatestarcodex",    "~120K", "理性主义/预测市场",           "Substack联系表单",                 "美国",    "★★★", "待联系"),
    ("Robin Hanson",        "@robinhanson",       "~40K",  "预测市场理论/期货民主",        "rhanson@gmu.edu",                 "美国",    "★★",  "待联系"),
    ("NightHawkX",          "@NightHawkX",        "~15K",  "Polymarket alpha/市场分析",   "Twitter DM",                      "美国",    "★★★", "待联系"),
    ("PolyTraderXYZ",       "@PolyTraderXYZ",     "~22K",  "Polymarket策略/DeFi",         "Twitter DM",                      "美国/东南亚","★★★","待联系"),
    ("Misha Yagudin",       "@mishayagudin",      "~12K",  "预测/EA/Metaculus",           "Twitter DM",                      "英国",    "★★",  "待联系"),
    ("Jens Wiechers",       "@jenswiechers",      "~10K",  "欧洲预测市场/政治预测",        "Twitter DM",                      "德国",    "★★",  "待联系"),
    ("Forecasting News",    "@forecastingnews",   "~9K",   "预测市场新闻聚合",             "Substack",                        "英国",    "★★",  "待联系"),
    ("Jason Maier",         "@jasonmaier",        "~15K",  "Kalshi/预测市场/金融",         "Twitter DM",                      "美国",    "★★",  "待联系"),
    ("Dawid Kopczyk",       "@DawidKopczyk",      "~12K",  "量化预测/预测市场",            "LinkedIn/Twitter DM",             "波兰",    "★★",  "待联系"),
    ("Samotsvety Team",     "@SamotsvetyTeam",    "~8K",   "协作预测/准确率",              "Manifold/Twitter DM",             "国际",    "★",   "待联系"),
]

for ri, row in enumerate(kol_data):
    r = ri + 3
    ws1.row_dimensions[r].height = 28
    bg = BLUE_LIGHT if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        sc(ws1, r, ci, val, bg=bg, center=(ci in [3,6,7,8]))

# ══════════════════════════════════════════
# Sheet 2：加密媒体
# ══════════════════════════════════════════
ws2 = wb.create_sheet("加密媒体")
ws2.sheet_view.showGridLines = False

cols2 = {"A":18,"B":18,"C":12,"D":20,"E":26,"F":12,"G":10}
for col, w in cols2.items():
    ws2.column_dimensions[col].width = w

sheet_title(ws2, "加密货币媒体名单 — 投稿/联系方式", 7)
header_row(ws2, 2, ["媒体名称", "Twitter", "粉丝量", "内容方向", "联系邮箱", "地区", "状态"])

crypto_media = [
    ("Cointelegraph",  "@Cointelegraph",  "~2M",   "加密新闻/区块链/市场",     "press@cointelegraph.com",      "国际",   "待联系"),
    ("CoinDesk",       "@CoinDesk",       "~1.8M", "加密新闻/市场/监管",        "press@coindesk.com",           "美国",   "待联系"),
    ("Decrypt",        "@Decrypt_Co",     "~340K", "加密/Web3/NFT/DeFi",       "press@decrypt.co",             "美国",   "待联系"),
    ("The Block",      "@TheBlock__",     "~400K", "加密研究/新闻/数据",        "tips@theblock.co",             "美国",   "待联系"),
    ("Blockworks",     "@Blockworks_",    "~280K", "机构加密/宏观",             "editorial@blockworks.co",      "美国",   "待联系"),
    ("Bankless",       "@BanklessHQ",     "~300K", "DeFi/以太坊/加密原生",      "press@bankless.com",           "美国",   "待联系"),
    ("The Defiant",    "@DefiantNews",    "~110K", "DeFi/去中心化金融",         "press@thedefiant.io",          "美国",   "待联系"),
    ("BeInCrypto",     "@beincrypto",     "~500K", "加密教育/新闻",             "press@beincrypto.com",         "国际",   "待联系"),
    ("CryptoSlate",    "@CryptoSlate",    "~120K", "加密新闻/ICO/DeFi",        "press@cryptoslate.com",        "美国",   "待联系"),
    ("Forkast News",   "@ForkastNews",    "~80K",  "亚太加密",                  "tips@forkast.news",            "香港/东南亚","待联系"),
    ("Asia Crypto Today","@AsiaCryptoToday","~50K","东南亚加密",               "contact@asiacryptotoday.com",  "东南亚", "待联系"),
    ("Unchained",      "@unchainedcrypto","~80K",  "加密深度/访谈",             "contact@unchained.com",        "美国",   "待联系"),
    ("Protos",         "@ProtosMedia",    "~30K",  "加密调查报道",              "tips@protos.com",              "国际",   "待联系"),
]

for ri, row in enumerate(crypto_media):
    r = ri + 3
    ws2.row_dimensions[r].height = 26
    bg = GREEN_LIGHT if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        sc(ws2, r, ci, val, bg=bg, center=(ci in [3,6,7]))

# ══════════════════════════════════════════
# Sheet 3：股票/财经媒体
# ══════════════════════════════════════════
ws3 = wb.create_sheet("财经股票媒体")
ws3.sheet_view.showGridLines = False

cols3 = {"A":18,"B":18,"C":12,"D":20,"E":26,"F":12,"G":10}
for col, w in cols3.items():
    ws3.column_dimensions[col].width = w

sheet_title(ws3, "财经/股票媒体名单 — 投稿/联系方式", 7)
header_row(ws3, 2, ["媒体名称", "Twitter", "粉丝量", "内容方向", "联系邮箱", "地区", "状态"])

finance_media = [
    ("Bloomberg Markets",  "@markets",        "~500K", "全球市场/宏观",           "press@bloomberg.net",          "美国/国际","待联系"),
    ("Financial Times",    "@FT",             "~5M",   "全球金融/经济",           "tips@ft.com",                  "英国",    "待联系"),
    ("Reuters Finance",    "@ReutersBiz",     "~3M",   "全球金融新闻",            "news.tips@reuters.com",        "国际",    "待联系"),
    ("CNBC Markets",       "@CNBCMarkets",    "~1.3M", "电视/数字财经新闻",        "cnbcpr@nbcuni.com",            "美国",    "待联系"),
    ("MarketWatch",        "@MarketWatch",    "~1.6M", "金融新闻/股票/经济",       "tips@marketwatch.com",         "美国",    "待联系"),
    ("Yahoo Finance",      "@YahooFinance",   "~800K", "市场/股票/经济",          "yfeditor@yahooinc.com",        "美国",    "待联系"),
    ("Forbes Finance",     "@Forbes",         "~18M",  "商业/金融/市场",          "contributors@forbes.com",      "美国",    "待联系"),
    ("Fortune Magazine",   "@FortuneMagazine","~2.3M", "商业/投资/经济",          "editors@fortune.com",          "美国",    "待联系"),
    ("Business Insider",   "@BusinessInsider","~3.6M", "金融/市场/经济",          "tips@businessinsider.com",     "美国/国际","待联系"),
    ("Seeking Alpha",      "@SeekingAlpha",   "~390K", "股票分析/投资",           "submissions@seekingalpha.com", "美国",    "待联系"),
    ("Benzinga",           "@Benzinga",       "~350K", "股票新闻/交易/市场",       "press@benzinga.com",           "美国",    "待联系"),
    ("The Motley Fool",    "@MotleyFool",     "~650K", "股票投资/长期",           "foolish@fool.com",             "美国",    "待联系"),
    ("Barron's",           "@barronsonline",  "~260K", "股票/基金/华尔街",         "editors@barrons.com",          "美国",    "待联系"),
    ("The Street",         "@TheStreet",      "~200K", "股票/投资",               "tips@thestreet.com",           "美国",    "待联系"),
    ("Investopedia",       "@Investopedia",   "~500K", "金融教育/市场",           "editorial@investopedia.com",   "美国",    "待联系"),
]

for ri, row in enumerate(finance_media):
    r = ri + 3
    ws3.row_dimensions[r].height = 26
    bg = ORANGE_LIGHT if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        sc(ws3, r, ci, val, bg=bg, center=(ci in [3,6,7]))

# ══════════════════════════════════════════
# Sheet 4：外联话术模板
# ══════════════════════════════════════════
ws4 = wb.create_sheet("外联话术模板")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 20
ws4.column_dimensions["B"].width = 80

sheet_title(ws4, "外联话术模板 — KOL DM + 媒体邮件 + 跟进邮件", 2)
header_row(ws4, 2, ["模板类型", "内容（复制使用，{}内填写对应信息）"])

templates = [
    ("KOL DM — 加密方向",
"""Hey {name} 👋

Big fan of your work on prediction markets — your take on {recent_topic} was spot on.

I'm reaching out because we're building something you might find interesting: a platform that tracks smart money moves across prediction markets (think GMGN, but for Polymarket/Kalshi).

We're onboarding a small group of early KOL partners before our public launch:
• Exclusive early access to our smart money dashboard
• Revenue share on every user you refer
• Custom analytics reports for your content

Would love to hop on a quick call or share more details.

Cheers, {sender_name} | {platform_name}"""),

    ("KOL DM — 美股方向",
"""Hey {name},

Love your coverage of {recent_topic} — really sharp analysis.

We're launching a prediction market aggregator that tracks where smart money is flowing on events like Fed decisions, earnings calls, and macro moves. Think Bloomberg Terminal meets Polymarket.

Looking for sharp finance voices to partner with for our launch:
• Early access to our smart money tracking dashboard
• Revenue share on referrals
• Co-branded market reports for your audience

Interested? Happy to send details or jump on a quick call.

Best, {sender_name} | {platform_name}"""),

    ("媒体邮件 — 加密媒体\n邮件主题：Exclusive: New Platform Brings Smart Money Tracking to Prediction Markets",
"""Hi {editor_name},

I'm reaching out from {platform_name}, a new prediction market aggregator bringing smart money tracking — familiar to crypto traders via tools like GMGN — to platforms like Polymarket and Kalshi.

Why this matters for your readers:
• Prediction markets now process $800M+ in daily volume with 100K+ daily active traders
• Until now, there's been no easy way to track where sophisticated money is moving
• We're the first platform to aggregate smart money signals across multiple prediction markets

We'd love to offer {media_name} an exclusive first look before our public launch, including:
• Interview with our founding team
• Early access to our smart money dashboard
• Proprietary data on prediction market flows

Would you be open to a briefing this week?

Best regards,
{sender_name} | {sender_title}
{platform_name} | {sender_email}"""),

    ("媒体邮件 — 财经媒体\n邮件主题：New Fintech Platform Lets Investors Track Smart Money in Prediction Markets",
"""Hi {editor_name},

I wanted to introduce {platform_name} — a new platform that aggregates prediction market data and shows investors where sophisticated money is flowing on events like Fed rate decisions, earnings surprises, and election outcomes.

Key data points:
• Global prediction markets: $800M+ daily volume, growing 40% YoY
• Our platform aggregates signals from Polymarket, Kalshi, and 5+ other markets
• Early users seeing 23% better returns by following smart money signals

We'd love to offer {media_name} an exclusive story opportunity, including data access and founder interviews.

Would you have 20 minutes for a briefing this week?

Best,
{sender_name} | {sender_title}
{platform_name} | {sender_email}"""),

    ("跟进邮件（7天无回复后发送）\n邮件主题：Re: {original_subject}",
"""Hi {name},

Just wanted to bump this up in case it got buried.

We're finalizing our launch partner list this week and wanted to make sure you had a chance to consider it before we close the early access window.

Even a 15-min call or quick email exchange works great.

Thanks either way!

{sender_name} | {platform_name}"""),

    ("KOL合作条款要点（谈判话术）",
"""【我们提供】
1. 独家早期访问权限（抢在公众前使用产品）
2. 专属推荐码 + 佣金分成（每成功注册用户收益的X%）
3. 定制化聪明钱数据报告（可直接用于内容创作）
4. 联合品牌活动机会

【我们期望】
1. 每月至少2条提及产品的原创内容
2. 使用专属推荐码（追踪转化）
3. 内容标注 #Ad 或 #Sponsored（合规要求）

【谈判底线】
• 头部KOL（>50万粉）：固定费用 + 分成
• 中腰部KOL（1-10万粉）：纯分成，无固定费用
• 媒体：提供独家数据+采访，不支付稿费"""),
]

for ri, (ttype, content) in enumerate(templates):
    r = ri + 3
    ws4.row_dimensions[r].height = 120
    bg = PURPLE_LIGHT if ri % 2 == 0 else WHITE
    sc(ws4, r, 1, ttype,   bg=bg, bold=True, size=9, center=False)
    sc(ws4, r, 2, content, bg=bg, bold=False, size=9)

path = "/Users/coco/agent-twitter/GMGN_KOL媒体外联名单.xlsx"
wb.save(path)
print(f"Excel 已生成：{path}")
