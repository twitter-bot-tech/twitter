#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()

BLUE_DARK    = "1F497D"
BLUE_MID     = "2E74B5"
ORANGE       = "C55A11"
ORANGE_LIGHT = "FCE4D6"
GREEN_DARK   = "375623"
GREEN_MID    = "70AD47"
GREEN_LIGHT  = "E2EFDA"
PURPLE_MID   = "7030A0"
PURPLE_LIGHT = "EAE0F4"
BLUE_LIGHT   = "EBF3FB"
WHITE        = "FFFFFF"

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value, bg=None, bold=False, size=9, color="000000", center=False):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border = thin_border()
    return cell

def sheet_title(ws, title, cols, bg=BLUE_DARK):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

def header_row(ws, row, headers, bg=BLUE_MID):
    ws.row_dimensions[row].height = 22
    for ci, h in enumerate(headers, 1):
        sc(ws, row, ci, h, bg=bg, bold=True, color=WHITE, center=True)

# ══════════════════════════════════════════
# Sheet 1：Meme 币 KOL 名单
# ══════════════════════════════════════════
ws1 = wb.active
ws1.title = "Meme币KOL"
ws1.sheet_view.showGridLines = False
for col, w in {"A":20,"B":20,"C":12,"D":22,"E":26,"F":14,"G":12,"H":10}.items():
    ws1.column_dimensions[col].width = w

sheet_title(ws1, "Meme 币交易 KOL 名单 — 欧美+东南亚", 8, bg=ORANGE)
header_row(ws1, 2, ["姓名/账号","Twitter Handle","粉丝量","内容方向","联系方式","地区","合作优先级","状态"], bg=ORANGE)

kol_data = [
    ("Murad Mahmudov",    "@MustStopMurad",    "~500K", "Meme币/Solana/加密周期",      "Twitter DM",                   "美国",    "★★★", "待联系"),
    ("Ansem",             "@blknoiz06",         "~550K", "Solana Meme币/链上交易",      "Twitter DM",                   "美国",    "★★★", "待联系"),
    ("Gainzy",            "@gainzy222",         "~180K", "Meme币交易/pump.fun",         "Twitter DM",                   "美国",    "★★★", "待联系"),
    ("Hsaka",             "@HsakaTrades",       "~350K", "加密交易/Meme币",             "Twitter DM",                   "美国",    "★★★", "待联系"),
    ("Ledger Status",     "@ledgerstatus",      "~120K", "Solana链上分析/Meme币",       "Twitter DM",                   "美国",    "★★★", "待联系"),
    ("Meme Insider",      "@memecoinsider",     "~90K",  "Meme币早期发现/alpha",        "Twitter DM",                   "美国",    "★★★", "待联系"),
    ("Defi Ignas",        "@DefiIgnas",         "~140K", "DeFi/Meme币/链上数据",        "Twitter DM",                   "欧洲",    "★★",  "待联系"),
    ("Crypto Banter",     "@cryptobanter",      "~400K", "加密/Meme币/YouTube",         "contact@cryptobanter.com",     "南非/欧洲","★★★","待联系"),
    ("Altcoin Daily",     "@AltcoinDailyio",    "~350K", "山寨币/Meme币/YouTube",       "contact@altcoindaily.co",      "美国",    "★★",  "待联系"),
    ("Crypto Daku",       "@crypto_daku",       "~85K",  "Meme币/东南亚加密",           "Twitter DM",                   "东南亚",  "★★★", "待联系"),
    ("Solana Floor",      "@solanafloor",       "~75K",  "Solana NFT/Meme币生态",       "Twitter DM",                   "美国",    "★★",  "待联系"),
    ("Nathan Worsley",    "@nathanworsley_",    "~60K",  "pump.fun/Solana Meme",        "Twitter DM",                   "英国",    "★★",  "待联系"),
    ("Koroush AK",        "@KoroushAK",         "~220K", "加密交易/Meme币策略",          "Twitter DM",                   "欧洲",    "★★",  "待联系"),
    ("Crypto Dog",        "@TheCryptoDog",      "~280K", "加密/Meme币",                 "Twitter DM",                   "美国",    "★★",  "待联系"),
    ("Jakub Dziadkowiec", "@jakubdz",           "~55K",  "Solana/Meme币/DEX数据",       "Twitter DM",                   "欧洲(波兰)","★★","待联系"),
]

for ri, row in enumerate(kol_data):
    r = ri + 3
    ws1.row_dimensions[r].height = 28
    bg = ORANGE_LIGHT if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        sc(ws1, r, ci, val, bg=bg, center=(ci in [3,6,7,8]))

# ══════════════════════════════════════════
# Sheet 2：Meme 币社群
# ══════════════════════════════════════════
ws2 = wb.create_sheet("Meme币社群")
ws2.sheet_view.showGridLines = False
for col, w in {"A":24,"B":12,"C":14,"D":30,"E":28,"F":14,"G":10}.items():
    ws2.column_dimensions[col].width = w

sheet_title(ws2, "Meme 币交易社群名单 — Telegram + Discord", 7, bg=GREEN_DARK)
header_row(ws2, 2, ["社群名称","平台","成员数","简介","链接/联系方式","地区","状态"], bg=GREEN_MID)

community_data = [
    ("Solana Meme Coins",         "Telegram", "~50K+",  "Solana链Meme币讨论/alpha分享",          "搜索 @solanamemecoins",          "国际",   "待联系"),
    ("pump.fun Traders",          "Telegram", "~30K+",  "pump.fun新币发现/早期交易",              "搜索 @pumpfuntraders",            "国际",   "待联系"),
    ("GMGN Alpha",                "Telegram", "~25K+",  "GMGN平台用户/聪明钱追踪",               "搜索 @gmgnalpha",                 "国际",   "待联系"),
    ("Meme Coin Gems",            "Telegram", "~40K+",  "早期Meme币发现/100x猎手",               "搜索 @memecoin_gems",             "国际",   "待联系"),
    ("DEX Screener Community",    "Telegram", "~35K+",  "DEX Screener用户/新币追踪",             "搜索 @dexscreener_community",     "国际",   "待联系"),
    ("Solana Trading Alpha",      "Discord",  "~60K+",  "Solana链交易/Meme币/NFT",               "discord.gg/solana-trading",       "国际",   "待联系"),
    ("Meme Coin Millionaires",    "Telegram", "~45K+",  "Meme币暴富策略/案例",                   "搜索 @memecoins_mm",              "国际",   "待联系"),
    ("Crypto Gems Underground",   "Telegram", "~20K+",  "低市值Meme币/早期进场",                 "搜索 @cryptogemsug",              "欧美",   "待联系"),
    ("SEA Crypto Traders",        "Telegram", "~30K+",  "东南亚加密/Meme币社群",                 "搜索 @seacryptofam",              "东南亚", "待联系"),
    ("Moonshot Hunters",          "Discord",  "~25K+",  "Meme币/小市值100x猎手",                 "discord.gg/moonshot",             "欧美",   "待联系"),
    ("Degen Trading Hub",         "Telegram", "~55K+",  "高风险Meme币/链上操盘",                 "搜索 @degentradinghub",           "国际",   "待联系"),
    ("Solana Snipers",            "Telegram", "~28K+",  "Solana新币狙击/机器人交易",             "搜索 @solanasnipers",             "国际",   "待联系"),
    ("Crypto Alpha Calls",        "Telegram", "~70K+",  "综合加密alpha/Meme币信号",              "搜索 @cryptoalphacalls",          "国际",   "待联系"),
    ("Base Chain Memes",          "Telegram", "~22K+",  "Base链Meme币/早期发现",                 "搜索 @basechainmemes",            "国际",   "待联系"),
    ("Meme Coin Research DAO",    "Discord",  "~18K+",  "Meme币研究/社区治理",                   "discord.gg/memeresearch",         "欧美",   "待联系"),
]

for ri, row in enumerate(community_data):
    r = ri + 3
    ws2.row_dimensions[r].height = 28
    bg = GREEN_LIGHT if ri % 2 == 0 else WHITE
    for ci, val in enumerate(row, 1):
        sc(ws2, r, ci, val, bg=bg, center=(ci in [2,3,6,7]))

# ══════════════════════════════════════════
# Sheet 3：推广消息模板
# ══════════════════════════════════════════
ws3 = wb.create_sheet("推广消息模板")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 78

sheet_title(ws3, "MoonX 推广消息模板 — KOL邮件 + 社群消息", 2, bg=PURPLE_MID)
header_row(ws3, 2, ["模板类型","内容"], bg=PURPLE_MID)

MOONX = "https://www.bydfi.com/en/moonx/markets/trending"

templates = [
    ("KOL 合作邮件\n（主题：Collab Opportunity: MoonX — The Fastest Meme Coin Trading Platform）",
f"""Hey {{name}} 👋

Love your meme coin content — you're one of the sharpest voices in the space.

I'm Kelly, Head of Marketing at BYDFi. We just launched MoonX ({MOONX}), and I think your audience would love it.

What makes MoonX different:
• Real-time meme coin discovery — catch pumps before they go viral
• Smart money tracking — see where whales are moving before you miss the move
• One-click trading directly from the trend feed
• Aggregated from pump.fun, DEX Screener and more

We'd love to have you as a launch partner:
✅ Exclusive early access + referral dashboard
✅ Revenue share on every trader you bring in
✅ Custom meme coin data feeds for your content
✅ Co-marketing (we'll amplify your posts)

Interested? Let's chat on Telegram or jump on a call.

Kelly | Head of Marketing | BYDFi MoonX
{MOONX}
Email: ppmworker@gmail.com | TG: @BDkelly"""),

    ("社群消息 — 通用版",
f"""👀 Meme coin traders — check this out

Just found MoonX by BYDFi and it's solid for finding early meme plays:

🔥 Real-time trending meme coins
🐋 Smart money / whale tracking
⚡ One-click trade straight from the feed
📊 Data from pump.fun, DEX Screener & more

Worth bookmarking if you're hunting early pumps 👇
{MOONX}

(Not financial advice — always DYOR)"""),

    ("社群消息 — Solana 社群版",
f"""Solana meme hunters 🚨

If you're not using MoonX yet you're probably missing moves.

Shows you what's trending in real-time + where smart money is going before it pumps. Built for speed — one-click trade from the discovery feed.

Try it: {MOONX}

Drop your thoughts below 👇"""),

    ("社群消息 — GMGN 用户版",
f"""For everyone who uses GMGN to track meme coins 👇

MoonX ({MOONX}) is worth checking as a companion:

→ Aggregates meme coin signals across chains
→ Smart money / whale alerts
→ Integrated trading (no wallet switching needed)

Free to use: {MOONX}"""),

    ("社群消息 — pump.fun 版",
f"""pump.fun traders 🎯

Tired of missing early pumps?

MoonX tracks what's trending on pump.fun + DEX Screener in real-time, shows you whale moves, and lets you trade in one click.

Check it: {MOONX}

Anyone else using this? 👀"""),
]

for ri, (ttype, content) in enumerate(templates):
    r = ri + 3
    ws3.row_dimensions[r].height = 130
    bg = PURPLE_LIGHT if ri % 2 == 0 else WHITE
    sc(ws3, r, 1, ttype,   bg=bg, bold=True, size=9)
    sc(ws3, r, 2, content, bg=bg, size=9)

path = "/Users/coco/agent-twitter/BYDFi_MoonX_Meme推广名单.xlsx"
wb.save(path)
print(f"Excel 已生成：{path}")
