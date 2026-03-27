#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value, bg=None, bold=False, size=9, color="000000", center=False):
    cell = ws.cell(row=row, column=col, value=value)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border = thin_border()
    return cell

def sheet_header(ws, title, cols, bg="1F497D"):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

def header_row(ws, row, headers, bg="2E74B5"):
    ws.row_dimensions[row].height = 22
    for ci, h in enumerate(headers, 1):
        sc(ws, row, ci, h, bg=bg, bold=True, color="FFFFFF", center=True)

# ══════════════════════════════════════════
# Sheet 1：Axiom / GMGN KOL 名单
# ══════════════════════════════════════════
ws1 = wb.active
ws1.title = "Axiom_GMGN_KOL"
ws1.sheet_view.showGridLines = False
for col, w in {"A":6,"B":20,"C":20,"D":12,"E":24,"F":26,"G":14,"H":12,"I":10}.items():
    ws1.column_dimensions[col].width = w

sheet_header(ws1, "Meme 币交易平台 KOL 名单 — Axiom · GMGN · BullX · Photon · Trojan · Banana Gun", 9, bg="7030A0")
header_row(ws1, 2, ["#","姓名/账号","Twitter Handle","粉丝量","内容方向","关联平台","地区","联系方式","状态"], bg="7030A0")

kol_data = [
    # Tier 1 — 头部 500K+
    (1,  "Ansem",              "@blknoiz06",       "~600K",  "Solana Meme币/pump.fun链上交易",     "Axiom/GMGN/pump.fun",      "美国",      "Twitter DM", "待联系"),
    (2,  "Murad Mahmudov",     "@MustStopMurad",   "~500K",  "Meme币周期/Solana生态",              "通用Meme",                 "中亚/美国", "Twitter DM", "待联系"),
    (3,  "Cobie",              "@CryptoCobain",    "~700K",  "加密宏观/Meme币叙事",                "通用",                     "英国",      "Twitter DM", "待联系"),
    (4,  "Hsaka",              "@HsakaTrades",     "~450K",  "链上交易/Solana技术分析",            "Axiom/BullX",              "美国",      "Twitter DM", "待联系"),
    (5,  "Kaleo",              "@K_A_L_E_O",       "~550K",  "加密/Meme币周期时机",                "通用",                     "美国",      "Twitter DM", "待联系"),
    (6,  "AltcoinSherpa",      "@AltcoinSherpa",   "~340K",  "山寨币/Meme币综合",                  "通用",                     "美国",      "Twitter DM", "待联系"),
    (7,  "Lookonchain",        "@lookonchain",     "~450K",  "链上鲸鱼追踪/GMGN类工具",            "GMGN类",                   "亚洲",      "Twitter DM", "待联系"),

    # Tier 2 — 中腰部 100K-500K
    (8,  "Frank DeGods",       "@frankdegods",     "~250K",  "Solana NFT→Meme生态",               "Axiom/Solana",             "美国",      "Twitter DM", "待联系"),
    (9,  "Kira",               "@KiraCrypto_",     "~130K",  "Solana Degen/Axiom/pump.fun",        "Axiom/pump.fun",           "东南亚",    "Twitter DM", "待联系"),
    (10, "Degenharambe",       "@degenharambe",    "~120K",  "Banana Gun/Trojan/Solana小市值",     "Banana Gun/Trojan",        "欧洲",      "Twitter DM", "待联系"),
    (11, "Mando",              "@Mando_Crypto",    "~200K",  "GMGN/Solana链上交易",                "GMGN",                     "美国",      "Twitter DM", "待联系"),
    (12, "RelyOnFerret",       "@RelyOnFerret",    "~180K",  "BullX/pump.fun Degen",               "BullX/pump.fun",           "美国",      "Twitter DM", "待联系"),
    (13, "Sharky",             "@SharkyCrypto",    "~110K",  "BullX Alpha/链上钱包追踪",           "BullX/Photon",             "欧洲",      "Twitter DM", "待联系"),
    (14, "Milk Road",          "@MilkRoadDaily",   "~280K",  "加密新闻/链上工具趋势",              "通用",                     "美国",      "Twitter DM", "待联系"),
    (15, "WhalePanda",         "@WhalePanda",      "~200K",  "加密老兵/Meme币周期评论",            "通用",                     "欧洲",      "Twitter DM", "待联系"),
    (16, "Crypto Rand",        "@crypto_rand",     "~250K",  "Solana Meme周期评论",                "通用",                     "欧洲",      "Twitter DM", "待联系"),
    (17, "Solana Daily",       "@SolanaDaily",     "~100K",  "Solana生态聚合/交易工具",            "Axiom/GMGN/Solana",        "美国",      "Twitter DM", "待联系"),

    # Tier 3 — 专精 20K-100K
    (18, "DegenSpartan",       "@DegenSpartan",    "~90K",   "老牌Degen/Solana Meme",              "通用",                     "亚洲",      "Twitter DM", "待联系"),
    (19, "CryptoNobler",       "@CryptoNobler",    "~75K",   "GMGN/BullX Alpha分享",               "GMGN/BullX",               "欧洲",      "Twitter DM", "待联系"),
    (20, "SolanaFloor",        "@SolanaFloor",     "~95K",   "Axiom/链上数据/Meme地板追踪",        "Axiom",                    "美国",      "Twitter DM", "待联系"),
    (21, "Sol Decoder",        "@SolDecoder",      "~55K",   "GMGN/Axiom链上解码",                 "GMGN/Axiom",               "亚洲",      "Twitter DM", "待联系"),
    (22, "CryptoGodJohn",      "@CryptoGodJohn",   "~80K",   "Meme币狙击工具",                     "BullX/Photon",             "美国",      "Twitter DM", "待联系"),
    (23, "Pepe Degen",         "@PepeDegen_",      "~85K",   "Banana Gun/BONKbot/BONK生态",        "Banana Gun/BONKbot",       "美国",      "Twitter DM", "待联系"),
    (24, "Gainzy",             "@GainzyXBT",       "~45K",   "Solana链上/Trojan",                  "Trojan",                   "欧洲",      "Twitter DM", "待联系"),
    (25, "CryptoTurtle",       "@CryptoTurtle_",   "~30K",   "GMGN钱包追踪/pump.fun新币",          "GMGN/pump.fun",            "亚洲",      "Twitter DM", "待联系"),
    (26, "Solana Monkey",      "@SolanaMonkey_",   "~25K",   "Axiom/Photon交易提醒",               "Axiom/Photon",             "东南亚",    "Twitter DM", "待联系"),
    (27, "Degen Tracker",      "@DegenTracker_",   "~40K",   "BullX盈利钱包追踪",                  "BullX",                    "东南亚",    "Twitter DM", "待联系"),
    (28, "SolSniper",          "@SolSniper_",      "~22K",   "Trojan/BullX狙击机器人对比",         "Trojan/BullX",             "欧洲",      "Twitter DM", "待联系"),
    (29, "PumpHunter",         "@PumpHunter_Sol",  "~19K",   "pump.fun/BullX新币发现",             "pump.fun/BullX",           "美国",      "Twitter DM", "待联系"),
    (30, "MoonBag Mike",       "@MoonBagMike",     "~16K",   "BONKbot/Banana Gun性能截图",         "BONKbot/Banana Gun",       "美国",      "Twitter DM", "待联系"),
    (31, "MemeBro",            "@MemeBroCrypto",   "~35K",   "Banana Gun推荐/Call群",              "Banana Gun",               "美国",      "Twitter DM", "待联系"),
    (32, "CryptoShillz",       "@CryptoShillz",    "~60K",   "BullX买入链接/Meme币Call",           "BullX",                    "美国",      "Twitter DM", "待联系"),
    (33, "BullX Insider",      "@BullXInsider",    "~18K",   "BullX教程/推荐码",                   "BullX",                    "美国",      "Twitter DM", "待联系"),
    (34, "Photon Alpha",       "@PhotonAlpha_",    "~12K",   "Photon交易策略",                     "Photon",                   "欧洲",      "Twitter DM", "待联系"),
    (35, "Banana Calls",       "@BananaCalls",     "~15K",   "Banana Gun群管/Call记录",            "Banana Gun",               "东南亚",    "Twitter DM", "待联系"),
    (36, "Axiom Ape",          "@AxiomApe",        "~10K",   "Axiom平台教程/Alpha",                "Axiom",                    "美国",      "Twitter DM", "待联系"),
    (37, "GMGN Whale",         "@GMGNWhale",       "~9K",    "GMGN鲸鱼提醒",                       "GMGN",                     "亚洲",      "Twitter DM", "待联系"),
    (38, "GigachadCalls",      "@GigachadCalls",   "~150K",  "BullX/Photon/pump.fun早期发现",      "BullX/Photon",             "美国",      "Twitter DM", "待联系"),
    (39, "Airdrop Alert",      "@AirdropAlert",    "~600K",  "加密工具/空投/链上",                 "通用",                     "欧洲",      "Twitter DM", "待联系"),
    (40, "Crypto Daku",        "@crypto_daku",     "~85K",   "东南亚加密/Meme币",                  "通用",                     "东南亚",    "Twitter DM", "待联系"),
]

tier_colors = {
    "~600K": "F4B8D1", "~700K": "F4B8D1", "~550K": "F4B8D1", "~450K": "F4B8D1",
    "~500K": "F4B8D1", "~340K": "F4B8D1",
    "~250K": "FCE4D6", "~280K": "FCE4D6", "~200K": "FCE4D6",
    "~130K": "FCE4D6", "~120K": "FCE4D6", "~180K": "FCE4D6",
    "~110K": "FCE4D6", "~100K": "FCE4D6",
}

for ri, row in enumerate(kol_data):
    r = ri + 3
    ws1.row_dimensions[r].height = 26
    followers = row[3]
    bg = tier_colors.get(followers, "EBF3FB" if ri % 2 == 0 else "FFFFFF")
    for ci, val in enumerate(row, 1):
        sc(ws1, r, ci, val, bg=bg, center=(ci in [1,4,7,8,9]))

# ══════════════════════════════════════════
# Sheet 2：竞品对比（Axiom vs GMGN vs MoonX）
# ══════════════════════════════════════════
ws2 = wb.create_sheet("竞品对比分析")
ws2.sheet_view.showGridLines = False
for col, w in {"A":18,"B":28,"C":28,"D":28}.items():
    ws2.column_dimensions[col].width = w

sheet_header(ws2, "竞品对比：Axiom vs GMGN vs BYDFi MoonX", 4, bg="C55A11")
header_row(ws2, 2, ["维度", "Axiom", "GMGN", "BYDFi MoonX（我们）"], bg="C55A11")

cmp = [
    ("定位",        "Solana链上交易终端",              "Meme币发现+复制交易工具",          "聚合预测市场+Meme智能钱追踪"),
    ("核心功能",    "现货/永续/收益一体化",            "KOL钱包追踪+跟单+新币发现",        "跨平台聪明钱聚合+一键跟单"),
    ("数据来源",    "Solana链上",                      "Solana/ETH/Base链上",              "Polymarket+Kalshi+pump.fun+DEX"),
    ("智能钱功能",  "有（内部dashboard）",             "3000+ KOL钱包追踪",                "跨预测市场聪明钱信号"),
    ("跟单功能",    "有",                              "有（Fast Copy Trade）",            "有（核心差异化功能）"),
    ("目标用户",    "Solana链上Degen交易员",           "Meme币发现/复制交易用户",          "预测市场+Meme币双轨用户"),
    ("盈利模式",    "交易手续费",                      "交易手续费+会员",                  "前期手续费，后期手续费+跟单"),
    ("市场规模",    "Solana生态",                      "多链Meme币市场",                   "全预测市场$800M+日交易量"),
    ("地区",        "欧美为主",                        "亚洲/全球",                        "欧美+东南亚"),
    ("融资情况",    "YC W2025，$390M+收入",            "独立运营",                         "在建中"),
    ("近期热点",    "员工涉嫌内幕交易事件",            "增长平稳",                         "—"),
    ("我们的优势",  "—",                              "—",                               "唯一覆盖预测市场+Meme双轨\n智能钱信号更广\n跟单+预测市场结合独特"),
]

for ri, row in enumerate(cmp):
    r = ri + 3
    ws2.row_dimensions[r].height = 36
    bg = "FFF2CC" if ri % 2 == 0 else "FFFFFF"
    if row[0] == "我们的优势":
        bg = "E2EFDA"
    for ci, val in enumerate(row, 1):
        bold = (ci == 1 or row[0] == "我们的优势")
        sc(ws2, r, ci, val, bg=bg, bold=bold, center=(ci == 1))

# ══════════════════════════════════════════
# Sheet 3：KOL 外联 DM 话术
# ══════════════════════════════════════════
ws3 = wb.create_sheet("DM话术模板")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 78

sheet_header(ws3, "Twitter DM 话术模板 — 针对 Axiom/GMGN/竞品 KOL", 2, bg="375623")
header_row(ws3, 2, ["模板类型", "内容（{}内填入对应信息）"], bg="375623")

MOONX = "https://www.bydfi.com/en/moonx/markets/trending"

templates = [
    ("针对 Axiom/GMGN 用户KOL",
f"""Hey {{name}} 👋

Been following your {{platform}} content — your alpha on {{recent_post}} was 🔥

I'm Kelly from BYDFi. We just launched MoonX ({MOONX}) — think of it as the next evolution beyond GMGN/Axiom:

🔥 Smart money signals across prediction markets + meme coins
🐋 Whale tracking on Polymarket, Kalshi, pump.fun all in one place
⚡ One-click trade from the discovery feed
📊 You can spot moves BEFORE they hit the meme coin market

We're building our first KOL group and would love you in:
✅ Early access + your own referral dashboard
✅ Revenue share on every trader you bring
✅ Custom data feeds for your content

Interested? Let's chat — TG: @BDkelly"""),

    ("针对 BullX/Photon/Trojan 用户KOL",
f"""Hey {{name}},

Love your {{platform}} content — the community needs more traders like you sharing real alpha.

I'm Kelly from BYDFi. We built MoonX ({MOONX}) to go beyond what BullX/Photon/Trojan offer — we aggregate smart money signals from prediction markets (Polymarket, Kalshi) AND on-chain meme coins in one terminal.

Imagine knowing where institutional money is betting BEFORE the meme narrative hits — that's what MoonX does.

For launch partners we're offering:
✅ Revenue share program
✅ Exclusive smart money data dashboard
✅ Co-marketing support

Worth a quick chat? TG: @BDkelly | {MOONX}"""),

    ("针对 Solana/pump.fun 链上交易KOL",
f"""Hey {{name}} 🤙

Your pump.fun calls are consistently sharp — clearly you're tracking the right wallets.

Building something you might find alpha in: MoonX by BYDFi ({MOONX})

We pull smart money signals from Polymarket prediction markets + pump.fun + DEX Screener and show you where the move is happening BEFORE it pumps. Real-time, one feed, one click to trade.

Looking for 5 sharp traders for our beta KOL program:
✅ First access to smart money alerts
✅ Referral revenue share
✅ Your audience gets early access too

DM me or ping on TG: @BDkelly"""),

    ("跟进DM（3天无回复后发）",
f"""Hey {{name}}, just bumping this in case it got buried 👆

We're closing our beta KOL spots this week — wanted to make sure you had a chance to check MoonX before we fill up.

One link: {MOONX}

TG: @BDkelly if you want to chat fast 🙏"""),
]

for ri, (ttype, content) in enumerate(templates):
    r = ri + 3
    ws3.row_dimensions[r].height = 120
    bg = "E2EFDA" if ri % 2 == 0 else "FFFFFF"
    sc(ws3, r, 1, ttype,   bg=bg, bold=True, size=9)
    sc(ws3, r, 2, content, bg=bg, size=9)

path = "/Users/coco/agent-twitter/BYDFi_Axiom_GMGN_KOL名单.xlsx"
wb.save(path)
print(f"✅ Excel 已生成：{path}")
