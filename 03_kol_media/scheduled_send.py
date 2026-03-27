#!/usr/bin/env python3
"""
定时批量发送 — BYDFi MoonX KOL 合作邮件
北京时间 22:00 自动发送（= 美东 09:00）
每日上限 10 封，4套差异化模板，按KOL分级路由
"""

import smtplib, os, time, logging, shutil, json, urllib.request
from datetime import datetime
from zoneinfo import ZoneInfo
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from typing import Optional
from dotenv import load_dotenv
from openpyxl import load_workbook
import random

DASH_PUSH_URL = "https://moonx-lark-server.fly.dev/push/stats"


def _push_kol_stats(sent_today: int, sent_total: int, kol_total: int):
    """发送完毕后把数据推送到 Fly.io server，供 /dash 页面展示"""
    BJT = ZoneInfo("Asia/Shanghai")
    payload = json.dumps({
        "date":        datetime.now(BJT).strftime("%Y-%m-%d"),
        "sent_today":  sent_today,
        "sent_total":  sent_total,
        "kol_total":   kol_total,
    }).encode()
    try:
        req = urllib.request.Request(
            DASH_PUSH_URL, data=payload,
            headers={"Content-Type": "application/json"}, method="POST"
        )
        urllib.request.urlopen(req, timeout=6)
    except Exception as e:
        logging.getLogger(__name__).warning(f"push_stats 失败: {e}")

# ── 数据目录：Fly.io 用持久化 volume，本地用脚本目录 ─────────────────────────
_SCRIPT_DIR = Path(__file__).parent
_ON_FLY     = bool(os.getenv("FLY_APP_NAME"))
KOL_DATA_DIR = Path("/data") if _ON_FLY else _SCRIPT_DIR
_LOG_DIR     = KOL_DATA_DIR / "logs"
_LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(_LOG_DIR / "scheduled_send.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


def _ensure_data_dir():
    """Fly.io 首次启动时把 Excel 从镜像目录复制到持久化 volume。"""
    if not _ON_FLY:
        return
    for pattern in ["MoonX_YouTube_KOL名单_*.xlsx", "MoonX_Kalshi_KOL名单_*.xlsx"]:
        for src in sorted(_SCRIPT_DIR.glob(pattern)):
            dest = KOL_DATA_DIR / src.name
            if not dest.exists():
                shutil.copy2(src, dest)
                logger.info(f"初始化：已复制 {src.name} → /data/")

load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

SMTP_HOST    = os.getenv("BYDFI_SMTP_HOST")
SMTP_PORT    = int(os.getenv("BYDFI_SMTP_PORT", "465"))
SMTP_USER    = os.getenv("BYDFI_EMAIL", "kelly@bydfi.com")
SMTP_PASS    = os.getenv("BYDFI_EMAIL_PASSWORD")
SENDER_NAME  = os.getenv("SENDER_NAME", "Kelly")
SENDER_TITLE = os.getenv("SENDER_TITLE", "Head of Marketing")
SENDER_TG    = os.getenv("SENDER_TG", "@BDkelly")
MOONX_URL    = "https://www.bydfi.com/en/moonx/markets/trending"
REBATE_URL   = "https://www.bydfi.com/zh/moonx/account/my-rebate?type=my-rebate"
UTM_BASE     = "https://www.bydfi.com/en/moonx/markets/trending"


def _make_utm_url(utm_code: str) -> str:
    """为 KOL 生成专属 UTM 追踪链接"""
    if not utm_code:
        return MOONX_URL
    return (
        f"{UTM_BASE}?utm_source=kol&utm_medium=email"
        f"&utm_campaign=outreach&utm_content={utm_code}"
    )
BJT          = ZoneInfo("Asia/Shanghai")

# 返佣比例（已确认）
REV_FUTURES     = 65
REV_MEME        = 50
REV_PREDICTION  = 40


# ══════════════════════════════════════════════════════════════════════════════
# 底层发送（传入已生成的 subject + body）
# ══════════════════════════════════════════════════════════════════════════════

def _send_raw(to_email: str, to_name: str, subject: str, body: str) -> bool:
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"{SENDER_NAME} <{SMTP_USER}>"
        msg["To"]      = to_email
        msg.attach(MIMEText(body, "plain", "utf-8"))
        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=30) as server:
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_USER, to_email, msg.as_string())
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_USER, to_email, msg.as_string())
        logger.info(f"  ✓ 已发送 → {to_name} <{to_email}>")
        return True
    except Exception as e:
        logger.warning(f"  ✗ 失败 → {to_email}：{e}")
        return False


# ══════════════════════════════════════════════════════════════════════════════
# 模板生成（按 KOL 分级）
# ══════════════════════════════════════════════════════════════════════════════

def _make_kalshi_email(name: str, desc_snippet: str, utm_url: str = "") -> tuple:
    """Kalshi 前KOL / PM级：直接点明返佣数字，founding partner 角度"""
    subjects = [
        f"Founding partner offer — {REV_FUTURES}% rev-share on futures",
        f"We offer {REV_FUTURES}% rev-share on futures — not a typo",
        f"Founding partner terms: up to {REV_FUTURES}% revenue share",
    ]
    hook = f'Saw your content on "{desc_snippet}..." — and reached out specifically because of it.' \
        if desc_snippet else "I came across your work on prediction markets and wanted to reach out directly."
    body = (
        f"Hi {name},\n\n"
        f"{hook}\n\n"
        f"I run marketing at BYDFi MoonX — a prediction market aggregator pulling live data "
        f"from Polymarket, Kalshi, and on-chain markets into one feed.\n\n"
        f"We're opening founding partner slots, and the terms are different from standard:\n\n"
        f"  Futures trading        → {REV_FUTURES}% rev-share to you\n"
        f"  Meme/on-chain trading  → {REV_MEME}%\n"
        f"  Prediction markets     → {REV_PREDICTION}%\n\n"
        f"Industry standard is 20~30%. We can offer this because we want the right partners "
        f"early, not the most partners.\n\n"
        f"One question: what market are you covering next? I can pull the smart money flow "
        f"data on that topic and send it to you before you publish.\n\n"
        f"{SENDER_NAME} · {SENDER_TITLE} · BYDFi MoonX\n"
        f"{utm_url or MOONX_URL}\n"
        f"TG: {SENDER_TG}"
    )
    return random.choice(subjects), body


def _make_b_tier_email(name: str, desc_snippet: str, utm_url: str = "") -> tuple:
    """B级（10~100万）：独家数据内容合作角度"""
    subjects = [
        "Your audience is already trading these markets — I have the data",
        "Exclusive data angle for your next video",
        "Smart money data your viewers haven't seen",
    ]
    intro = f'Saw your recent content — "{desc_snippet}..." — and thought this was directly relevant.' \
        if desc_snippet else "Been following your channel — your audience is exactly who we built this for."
    body = (
        f"Hi {name},\n\n"
        f"{intro}\n\n"
        f"We built MoonX at BYDFi — it aggregates real-time smart money flows across "
        f"Polymarket, Kalshi, and on-chain prediction markets into one feed. We can see "
        f"exactly where sophisticated capital is moving before it hits the odds.\n\n"
        f"I can pull live data on any market your audience is watching right now — and share "
        f'it exclusively with you before it goes public. Strong content angle: '
        f'"here\'s what smart money actually thinks about X."\n\n'
        f"If you want to share the platform with your audience, founding partner rev-share: "
        f"{REV_PREDICTION}% on prediction trades, {REV_FUTURES}% on futures.\n\n"
        f"Are you covering prediction markets or on-chain alpha this week? "
        f"I'll pull the relevant data.\n\n"
        f"{SENDER_NAME} · {SENDER_TITLE} · BYDFi MoonX\n"
        f"{utm_url or MOONX_URL}\n"
        f"TG: {SENDER_TG}"
    )
    return random.choice(subjects), body


def _make_c_tier_email(name: str, desc_snippet: str, utm_url: str = "") -> tuple:
    """C/D级（1千~10万）：工具体验 + 返佣数字，选择题CTA"""
    subjects = [
        "Tool your audience would actually use — free access",
        "Prediction market data most crypto traders don't see",
        "Early access: smart money tracker for crypto traders",
    ]
    intro = f'Noticed you cover "{desc_snippet}..." — this is directly relevant.' \
        if desc_snippet else "Your content resonates with exactly the audience we're building for."
    body = (
        f"Hi {name},\n\n"
        f"{intro}\n\n"
        f"We just launched MoonX — real-time aggregator for prediction markets "
        f"(Polymarket, Kalshi) and on-chain meme coin smart money. One feed, all the signals.\n\n"
        f"Free access for you, no strings.\n\n"
        f"If your audience finds it useful, founding partner rev-share:\n"
        f"  Futures  → {REV_FUTURES}% to you\n"
        f"  Meme     → {REV_MEME}%\n"
        f"  Prediction markets → {REV_PREDICTION}%\n\n"
        f"Most platforms offer 20~30%. We offer more because we want quality partners, "
        f"not volume.\n\n"
        f"Which fits your audience better — prediction markets or meme/on-chain? "
        f"I'll send you the relevant onboarding link.\n\n"
        f"Platform: {utm_url or MOONX_URL}\n"
        f"Rev-share: {REBATE_URL}\n\n"
        f"{SENDER_NAME} · {SENDER_TITLE} · BYDFi MoonX\n"
        f"TG: {SENDER_TG}"
    )
    return random.choice(subjects), body


def _make_meme_email(name: str, desc_snippet: str, utm_url: str = "") -> tuple:
    """Meme KOL：链上钱包追踪角度 + 返佣数字"""
    subjects = [
        "Saw a token 10x before CT caught it — sharing how",
        "The Solana scanner I've been using lately",
        f"{REV_MEME}% rev-share on meme trades — want in?",
    ]
    intro = f'Been following your content on "{desc_snippet}..." — you clearly know how to spot moves early.' \
        if desc_snippet else "Been following your content — you clearly know how to spot moves early."
    body = (
        f"Hi {name},\n\n"
        f"{intro}\n\n"
        f"MoonX tracks smart wallet activity across Solana and BNB Chain in real time. "
        f"Filters out 97% of noise, surfaces only tokens where smart money is moving — "
        f"before CT catches it. Integrated with pump.fun, Raydium, PancakeSwap.\n\n"
        f"Free access, no strings.\n\n"
        f"Rev-share for your audience:\n"
        f"  Meme/on-chain trades  → {REV_MEME}% to you\n"
        f"  Futures               → {REV_FUTURES}%\n\n"
        f"Your audience is already trading these markets. They just need the right tool "
        f"and you get paid on every trade they make.\n\n"
        f"Are you more focused on Solana or BNB Chain right now? "
        f"I'll customize the onboarding for your audience.\n\n"
        f"Platform: {utm_url or MOONX_URL}\n"
        f"Rev-share: {REBATE_URL}\n\n"
        f"{SENDER_NAME} · {SENDER_TITLE} · BYDFi MoonX\n"
        f"TG: {SENDER_TG}"
    )
    return random.choice(subjects), body


# ══════════════════════════════════════════════════════════════════════════════
# 按分级路由模板（YouTube KOL 名单）
# ══════════════════════════════════════════════════════════════════════════════

def _get_youtube_subject_and_body(name: str, tier: str, description: str,
                                   utm_url: str = "") -> tuple:
    desc_snippet = (description or "")[:80].strip()
    tier_upper = tier.upper()

    if "KALSHI" in tier_upper or "PM" in tier_upper:
        return _make_kalshi_email(name, desc_snippet, utm_url)
    elif "B" in tier_upper:
        return _make_b_tier_email(name, desc_snippet, utm_url)
    elif "D" in tier_upper:
        return _make_meme_email(name, desc_snippet, utm_url)
    else:
        return _make_c_tier_email(name, desc_snippet, utm_url)


# ══════════════════════════════════════════════════════════════════════════════
# 旧版 batch_send 兼容（Twitter KOL 名单用，niche 分类）
# ══════════════════════════════════════════════════════════════════════════════

def detect_niche(category: str) -> str:
    cat = category.lower()
    if "meme" in cat or "pump" in cat or "solana" in cat or "gmgn" in cat or "degen" in cat:
        return "meme"
    if "defi" in cat or "yield" in cat or "链上" in cat or "on-chain" in cat:
        return "onchain"
    if "bitcoin" in cat or "btc" in cat or "macro" in cat:
        return "macro"
    if "kalshi" in cat or "polymarket" in cat or "prediction" in cat or "pm" in cat:
        return "kalshi"
    return "default"


def get_subject_and_body(name: str, category: str, description: str = "") -> tuple:
    niche = detect_niche(category)
    desc_snippet = (description or "")[:80].strip()

    if niche == "kalshi":
        return _make_kalshi_email(name, desc_snippet)
    elif niche == "meme":
        return _make_meme_email(name, desc_snippet)
    elif niche in ("onchain", "macro"):
        return _make_b_tier_email(name, desc_snippet)
    else:
        return _make_c_tier_email(name, desc_snippet)


def send_email(to_email: str, to_name: str, category: str = "", description: str = "") -> bool:
    subject, body = get_subject_and_body(to_name, category, description)
    return _send_raw(to_email, to_name, subject, body)


# ══════════════════════════════════════════════════════════════════════════════
# 定时等待
# ══════════════════════════════════════════════════════════════════════════════

def wait_until_10pm():
    now = datetime.now(BJT)
    target = now.replace(hour=22, minute=0, second=0, microsecond=0)
    wait_sec = (target - now).total_seconds()
    if wait_sec > 0:
        print(f"⏰ 当前北京时间：{now.strftime('%H:%M:%S')}")
        print(f"   等待至 22:00 发送，还剩 {int(wait_sec//60)} 分 {int(wait_sec%60)} 秒...")
        time.sleep(wait_sec)


# ══════════════════════════════════════════════════════════════════════════════
# 旧版 Twitter KOL 名单批量发送
# ══════════════════════════════════════════════════════════════════════════════

def find_latest_kol_excel() -> Optional[Path]:
    kol_dir = Path(__file__).parent
    files = sorted(
        list(kol_dir.glob("MoonX_KOL名单_*.xlsx")) +
        list(kol_dir.glob("MoonX_YouTube_KOL名单_*.xlsx")),
        key=lambda f: f.name,
        reverse=True,
    )
    return files[0] if files else None


def get_col_indices(ws):
    headers = [cell.value for cell in ws[2]]
    def find(name):
        for i, h in enumerate(headers):
            if h and name in str(h):
                return i
        return None
    return {
        "name":        find("名字")    if find("名字")    is not None else 1,
        "email":       find("邮箱")    if find("邮箱")    is not None else 8,
        "status":      find("联系状态") if find("联系状态") is not None else 9,
        "category":    find("分类")    if find("分类")    is not None else 7,
        "description": find("描述")    if find("描述")    is not None else -1,
    }


def batch_send(excel_path: Optional[Path] = None):
    if excel_path is None:
        excel_path = find_latest_kol_excel()
    if not excel_path:
        print("❌ 未找到 KOL 名单 Excel，请先运行 collect_kols.py")
        return

    print(f"📂 读取名单：{excel_path.name}")
    wb = load_workbook(excel_path)
    ws = wb.active
    cols = get_col_indices(ws)
    sent, skip = 0, 0

    for row in ws.iter_rows(min_row=3, values_only=False):
        name     = row[cols["name"]].value or ""
        email    = row[cols["email"]].value or ""
        status   = row[cols["status"]].value or ""
        category = row[cols["category"]].value or ""
        desc     = row[cols["description"]].value if cols["description"] >= 0 else ""

        if not name or name.startswith("▌"):
            continue
        if not email or "@" not in str(email) or email == "—":
            skip += 1
            continue
        if status in ("已发送", "已回复", "已签约"):
            skip += 1
            continue

        success = send_email(str(email), str(name), str(category), str(desc or ""))
        if success:
            row[cols["status"]].value = "已发送"
            sent += 1
            wait = random.randint(45, 120)
            print(f"  ⏱ 等待 {wait} 秒...")
            time.sleep(wait)
        else:
            skip += 1

    wb.save(excel_path)
    print(f"\n✅ 完成！发送 {sent} 封，跳过 {skip} 条")
    print(f"📊 Excel 状态已更新：{excel_path.name}")
    # 统计累计发送总数
    total = sum(1 for r in ws.iter_rows(min_row=3, values_only=True)
                if r and (r[cols["status"]] if len(r) > cols["status"] else None) in ("已发送", "已回复", "已签约"))
    _push_kol_stats(sent_today=sent, sent_total=total, kol_total=sent + skip)


# ══════════════════════════════════════════════════════════════════════════════
# YouTube KOL 自动外联（22:00 BJT，按策略分层，每天最多 10 封）
# 优先级：Kalshi前KOL > PM级 > C级 > D级 > B级，跳过 A级
# ══════════════════════════════════════════════════════════════════════════════

def send_youtube_kol_emails(daily_limit: int = 10) -> int:
    _ensure_data_dir()

    # 优先从 SQLite 读取发送队列
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from kol_db import get_kols_to_send, mark_sent as db_mark_sent, init_db, get_db as _get_db, DB_PATH
        init_db()
        candidates_db = get_kols_to_send(daily_limit=daily_limit * 5)  # 多拿一些，按优先级排好序的
        use_sqlite = True
        logger.info(f"   [SQLite] 待发队列: {len(candidates_db)} 个")
    except Exception as e:
        logger.warning(f"   SQLite 读取失败，降级到 Excel 模式: {e}")
        use_sqlite = False
        candidates_db = []

    # 降级：从 Excel 读取（SQLite 不可用时）
    if not use_sqlite:
        return _send_youtube_from_excel(daily_limit)

    sent_count = 0
    for kol in candidates_db:
        if sent_count >= daily_limit:
            break

        kol_id   = kol["id"]
        name     = kol["name"]
        email    = kol["email"]
        tier     = kol["tier"]
        desc     = kol.get("description", "")
        subs     = kol.get("subscribers", 0)
        utm_code = kol.get("utm_code") or ""
        utm_url  = _make_utm_url(utm_code)

        subject, body = _get_youtube_subject_and_body(name, tier, desc, utm_url)
        success = _send_raw(email, name, subject, body)
        if success:
            db_mark_sent(kol_id, subject=subject, template=tier)
            _update_excel_status_by_email(email, "已发送")
            sent_count += 1
            logger.info(f"  ✓ 已发送 → [{tier}] {name} <{email}>  (subs={subs:,})")
            if sent_count < daily_limit:
                wait = random.randint(60, 180)
                logger.info(f"  ⏱ 等待 {wait} 秒...")
                time.sleep(wait)
        else:
            logger.warning(f"  ✗ 失败 → [{tier}] {name} <{email}>")

    logger.info(f"\n✅ YouTube KOL 外联完成：发送 {sent_count} 封 / 今日上限 {daily_limit} 封")
    remaining = len(candidates_db) - sent_count
    logger.info(f"📋 剩余待发：{remaining} 个")

    # 统计并推送
    try:
        with _get_db() as conn:
            total_sent = conn.execute(
                "SELECT COUNT(*) FROM kols WHERE status IN ('已发送','已回复','已签约','谈判中')"
            ).fetchone()[0]
            total_kols = conn.execute("SELECT COUNT(*) FROM kols").fetchone()[0]
    except Exception:
        total_sent, total_kols = sent_count, sent_count

    _push_kol_stats(sent_today=sent_count, sent_total=total_sent, kol_total=total_kols)
    return sent_count


def _update_excel_status_by_email(email: str, status: str):
    """在所有 Excel 中查找邮箱并更新状态（保持 Excel 与 SQLite 同步）"""
    try:
        from openpyxl import load_workbook as _lw
        for f in sorted(KOL_DATA_DIR.glob("MoonX_*KOL名单_*.xlsx"), reverse=True):
            try:
                wb = _lw(f)
                ws = wb.active
                changed = False
                for row in ws.iter_rows(min_row=3, values_only=False):
                    cell_email  = row[7].value if len(row) > 7 else None
                    cell_status = row[9].value if len(row) > 9 else None
                    if str(cell_email or "").strip().lower() == email.lower():
                        if str(cell_status or "").strip() not in ("已发送", "已回复", "已签约"):
                            row[9].value = status
                            changed = True
                if changed:
                    wb.save(f)
                wb.close()
            except Exception:
                pass
    except Exception:
        pass


def _send_youtube_from_excel(daily_limit: int = 10) -> int:
    """降级模式：直接读 Excel 发送（SQLite 不可用时）"""
    kol_dir = KOL_DATA_DIR
    all_files = (
        sorted(kol_dir.glob("MoonX_YouTube_KOL名单_*.xlsx"), reverse=True) +
        sorted(kol_dir.glob("MoonX_Kalshi_KOL名单_*.xlsx"), reverse=True)
    )
    if not all_files:
        logger.error("❌ 未找到 KOL 名单 Excel")
        return 0

    TIER_PRIORITY = {"KALSHI": 0, "PM": 1, "C": 2, "D": 3, "B": 4, "A": 99}
    candidates = []
    for f in all_files:
        wb = load_workbook(f)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
            name   = str(row[1].value or "").strip()
            email  = str(row[7].value or "").strip()
            tier   = str(row[5].value or "").strip()
            status = str(row[9].value or "").strip()
            desc   = str(row[8].value or "").strip()
            subs   = row[4].value or 0
            if not name or not email or "@" not in email:
                continue
            if status in ("已发送", "已回复", "已签约"):
                continue
            tier_key = tier.upper().replace("级", "").strip()
            priority = TIER_PRIORITY.get(tier_key, 5)
            if priority == 99:
                continue
            candidates.append((priority, f, i, name, email, tier, desc, subs))
    candidates.sort(key=lambda x: (x[0], -x[7]))

    sent_count = 0
    for priority, f, row_idx, name, email, tier, desc, subs in candidates:
        if sent_count >= daily_limit:
            break
        subject, body = _get_youtube_subject_and_body(name, tier, desc)
        success = _send_raw(email, name, subject, body)
        if success:
            wb = load_workbook(f)
            ws = wb.active
            ws.cell(row=row_idx, column=10).value = "已发送"
            wb.save(f)
            sent_count += 1
            if sent_count < daily_limit:
                time.sleep(random.randint(60, 180))
    _push_kol_stats(sent_today=sent_count, sent_total=sent_count, kol_total=len(candidates))
    return sent_count


# ══════════════════════════════════════════════════════════════════════════════
# 入口
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys
    from pathlib import Path as _Path
    import sys as _sys; _sys.path.insert(0, str(_Path(__file__).parent.parent))
    from scripts.pid_lock import acquire_lock; acquire_lock()

    if len(sys.argv) > 1 and sys.argv[1] == "--youtube":
        limit = int(sys.argv[2]) if len(sys.argv) > 2 else 10
        logger.info(f"\n🚀 YouTube KOL 外联开始 {datetime.now(BJT).strftime('%Y-%m-%d %H:%M')} BJT\n")
        send_youtube_kol_emails(daily_limit=limit)
        sys.exit(0)

    if len(sys.argv) > 1 and sys.argv[1] == "--all":
        kol_dir = Path(__file__).parent
        files = sorted(kol_dir.glob("MoonX_KOL名单_*.xlsx"))
        for f in files:
            print(f"\n{'='*50}")
            print(f"🚀 开始发送：{f.name}")
            batch_send(f)
    elif len(sys.argv) > 1:
        batch_send(Path(sys.argv[1]))
    else:
        wait_until_10pm()
        print(f"\n🚀 北京时间 {datetime.now(BJT).strftime('%H:%M:%S')} 开始批量发送...\n")
        batch_send()
