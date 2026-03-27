#!/usr/bin/env python3
"""
BYDFi MoonX — Kalshi 前KOL 专项收集脚本
目标：找到曾经宣传/使用/合作过 Kalshi 的 YouTube 创作者
不限粉丝量，全部标记为 Kalshi 级（最高优先级）
"""

import os, re, time, logging, subprocess, json
import requests
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

load_dotenv(Path(__file__).parent.parent / ".env", override=True)
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")
TODAY      = datetime.now().strftime("%Y-%m-%d")
OUTPUT_DIR = Path(__file__).parent
LOG_DIR    = OUTPUT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "kalshi_kol.log"),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)

# ── Kalshi 专用搜索词（精准定向） ────────────────────────────────────
KALSHI_QUERIES = [
    "kalshi tutorial",
    "kalshi trading",
    "kalshi prediction market",
    "kalshi how to trade",
    "kalshi review 2024",
    "kalshi review 2025",
    "kalshi vs polymarket",
    "kalshi sports betting",
    "kalshi election market",
    "kalshi crypto market",
    "kalshi finance",
    "prediction market kalshi",
    "trade kalshi",
    "kalshi affiliate",
    "kalshi referral",
]

# Kalshi 关键词检测（用于频道描述/标题中验证相关性）
KALSHI_KEYWORDS = {
    "kalshi", "prediction market", "event contract",
    "polymarket", "manifold", "metaculus"
}

MIN_SUBSCRIBERS = 500   # Kalshi KOL 不设高门槛，500粉以上都要

EMAIL_BLOCKLIST = {
    "example.com", "domain.com", "gmail.com.fake", "sentry.io",
    "cloudflare.com", "wixpress.com", "squarespace.com",
    "youtube.com", "youtu.be", "google.com",
}
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


# ── 工具函数 ──────────────────────────────────────────────────────

def extract_email(text: str) -> str:
    if not text:
        return ""
    t = re.sub(r'\s*\(at\)\s*', '@', text, flags=re.IGNORECASE)
    t = re.sub(r'\s*\(dot\)\s*', '.', t, flags=re.IGNORECASE)
    t = re.sub(r'\s*\[at\]\s*', '@', t, flags=re.IGNORECASE)
    t = re.sub(r'\s*\[dot\]\s*', '.', t, flags=re.IGNORECASE)
    for email in EMAIL_RE.findall(t):
        domain = email.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST and "." in domain:
            return email.lower()
    return ""


def scrape_channel_email(channel_url: str) -> str:
    """yt-dlp 抓取频道 About 页完整描述，提取邮箱"""
    try:
        result = subprocess.run(
            ["python3", "-m", "yt_dlp", "--dump-single-json",
             "--no-warnings", "--no-playlist", channel_url + "/about"],
            capture_output=True, text=True, timeout=15,
        )
        if result.returncode != 0 or not result.stdout.strip():
            return ""
        data = json.loads(result.stdout)
        return extract_email(data.get("description", ""))
    except Exception:
        return ""


def is_kalshi_related(title: str, description: str) -> bool:
    """二次验证：频道名/描述是否真的跟 Kalshi/预测市场相关"""
    text = (title + " " + description).lower()
    return any(kw in text for kw in KALSHI_KEYWORDS)


def search_channels(youtube, query: str, max_results: int = 50) -> list:
    try:
        resp = youtube.search().list(
            q=query, type="channel", part="snippet",
            maxResults=max_results, relevanceLanguage="en", order="relevance",
        ).execute()
        return [
            {"channel_id": item["snippet"]["channelId"], "title": item["snippet"]["title"]}
            for item in resp.get("items", [])
        ]
    except HttpError as e:
        logger.error(f"搜索失败 ({query}): {e}")
        return []


def get_channel_details(youtube, channel_ids: list) -> list:
    if not channel_ids:
        return []
    try:
        resp = youtube.channels().list(
            id=",".join(channel_ids),
            part="snippet,statistics",
        ).execute()
        results = []
        for item in resp.get("items", []):
            snippet = item.get("snippet", {})
            stats   = item.get("statistics", {})

            handle = snippet.get("customUrl", "")
            channel_url = (
                f"https://www.youtube.com/{handle}"
                if handle else f"https://www.youtube.com/channel/{item['id']}"
            )
            description = snippet.get("description", "")
            sub_str = stats.get("subscriberCount", "0")
            subs = int(sub_str) if sub_str.isdigit() else 0

            # 验证相关性（排除误搜）
            if not is_kalshi_related(snippet.get("title", ""), description):
                continue

            # 邮箱提取：description → yt-dlp（所有频道都抓，Kalshi KOL 价值高）
            email = extract_email(description)
            if not email:
                email = scrape_channel_email(channel_url)

            # Twitter handle
            twitter = ""
            m = re.search(r'twitter\.com/([A-Za-z0-9_]{1,50})|x\.com/([A-Za-z0-9_]{1,50})', description)
            if m:
                twitter = "@" + (m.group(1) or m.group(2))

            results.append({
                "channel_id":  item["id"],
                "name":        snippet.get("title", ""),
                "handle":      handle,
                "channel_url": channel_url,
                "subscribers": subs,
                "tier":        "Kalshi",   # 全部标记为 Kalshi 级
                "description": description[:200],
                "email":       email,
                "twitter":     twitter,
                "country":     snippet.get("country", ""),
            })
        return results
    except HttpError as e:
        logger.error(f"获取频道详情失败: {e}")
        return []


def load_existing_ids(output_dir: Path) -> set:
    seen = set()
    for fp in output_dir.glob("MoonX_Kalshi_KOL名单_*.xlsx"):
        try:
            wb = load_workbook(fp, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row and row[0]:
                    seen.add(str(row[0]).strip())
            wb.close()
        except Exception as e:
            logger.warning(f"读取 {fp.name} 失败: {e}")
    return seen


# ── Excel 输出 ────────────────────────────────────────────────────

def build_excel(channels: list, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kalshi前KOL名单"

    NAVY  = "1F4E79"
    LGOLD = "FFF2CC"
    GOLD  = "B8860B"
    RED   = "C00000"
    WHITE = "FFFFFF"
    GRAY  = "F2F2F2"

    thin = Border(*[Side(style="thin", color="AAAAAA")] * 4)
    def bdr(c): c.border = Border(
        left=Side(style="thin", color="AAAAAA"), right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),  bottom=Side(style="thin", color="AAAAAA"))

    # 行1：总标题
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = f"BYDFi MoonX — Kalshi 前KOL 专项名单（{TODAY}）"
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.font  = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    bdr(c); ws.row_dimensions[1].height = 30

    # 行2：说明
    ws.merge_cells("A2:N2")
    c = ws["A2"]
    c.value = (
        "⚡ 最高优先级 | 返佣：合约65% · Meme50% · 预测市场40% | "
        "Kalshi于2026年2月砍掉X平台合作 | 发送前请在【个性化钩子】列填入引用内容"
    )
    c.fill = PatternFill("solid", fgColor=LGOLD)
    c.font = Font(bold=True, color=GOLD, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    bdr(c); ws.row_dimensions[2].height = 22

    # 行3：列头
    headers = [
        ("Channel ID", 26), ("KOL名称", 22), ("YouTube链接", 38), ("Twitter Handle", 20),
        ("订阅数", 12), ("分级", 8), ("国家", 10), ("邮箱", 35),
        ("内容方向", 40), ("个性化钩子 ✏️", 45), ("联系状态", 12),
        ("跟进轮次", 12), ("备注", 20), ("收集日期", 14),
    ]
    for col_idx, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.fill = PatternFill("solid", fgColor=(RED if h.startswith("个性化") else NAVY))
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        bdr(c)
        ws.column_dimensions[chr(64 + col_idx)].width = w
    ws.row_dimensions[3].height = 24

    # 数据行
    emails_found = 0
    for r_idx, ch in enumerate(channels, 4):
        bg = WHITE if r_idx % 2 == 0 else GRAY
        row_data = [
            ch["channel_id"], ch["name"], ch["channel_url"], ch["twitter"],
            ch["subscribers"], ch["tier"], ch["country"], ch["email"],
            ch["description"], "",   # 个性化钩子（空，待填）
            "待发送", "Day0", "", TODAY,
        ]
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            if col_idx == 10:  # 个性化钩子列：淡红底提醒
                c.fill = PatternFill("solid", fgColor="FFF0F0")
            else:
                c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(size=9)
            c.alignment = Alignment(vertical="center", wrap_text=(col_idx in [9, 10]))
            bdr(c)
        ws.row_dimensions[r_idx].height = 40
        if ch["email"]:
            emails_found += 1

    ws.freeze_panes = "A4"
    wb.save(output_path)
    logger.info(f"✅ Excel 已保存：{output_path.name}（{len(channels)} 条，邮箱 {emails_found} 个）")
    return emails_found


# ── 主流程 ────────────────────────────────────────────────────────

def run():
    if not YOUTUBE_API_KEY:
        logger.error("❌ 未找到 YOUTUBE_API_KEY")
        return

    logger.info("=" * 60)
    logger.info(f"🎯 Kalshi 前KOL 收集开始 — {TODAY}")
    logger.info(f"   搜索词：{len(KALSHI_QUERIES)} 个")
    logger.info("=" * 60)

    youtube = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)
    existing_ids = load_existing_ids(OUTPUT_DIR)
    logger.info(f"📋 已有记录：{len(existing_ids)} 个（自动去重）")

    # 搜索所有关键词
    all_channel_ids: dict = {}
    for query in KALSHI_QUERIES:
        logger.info(f"🔍 搜索：{query}")
        for ch in search_channels(youtube, query, 50):
            cid = ch["channel_id"]
            if cid not in existing_ids:
                all_channel_ids[cid] = ch["title"]
        time.sleep(0.5)

    logger.info(f"\n📊 去重后待查询：{len(all_channel_ids)} 个频道")
    if not all_channel_ids:
        logger.info("⚠ 没有新频道，结束")
        return

    # 批量获取详情（每次50个）
    all_details = []
    id_list = list(all_channel_ids.keys())
    for i in range(0, len(id_list), 50):
        batch = id_list[i:i+50]
        details = get_channel_details(youtube, batch)
        all_details.extend(details)
        logger.info(f"  已处理 {min(i+50, len(id_list))}/{len(id_list)} 个，有效 {len(all_details)} 个")
        time.sleep(0.5)

    # 过滤低粉丝
    filtered = [ch for ch in all_details if ch["subscribers"] >= MIN_SUBSCRIBERS]
    # 按订阅数降序
    filtered.sort(key=lambda x: x["subscribers"], reverse=True)

    logger.info(f"\n✅ 过滤后（≥{MIN_SUBSCRIBERS}粉）：{len(filtered)} 个 Kalshi KOL")

    if not filtered:
        logger.info("⚠ 没有符合条件的频道")
        return

    # 输出 Excel
    output_path = OUTPUT_DIR / f"MoonX_Kalshi_KOL名单_{TODAY}.xlsx"

    # 如果今天已有文件，追加而不是覆盖
    if output_path.exists():
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            existing_in_file = set()
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row and row[0]:
                    existing_in_file.add(str(row[0]).strip())
            new_channels = [ch for ch in filtered if ch["channel_id"] not in existing_in_file]
            if new_channels:
                last_row = ws.max_row
                for ch in new_channels:
                    last_row += 1
                    row_data = [
                        ch["channel_id"], ch["name"], ch["channel_url"], ch["twitter"],
                        ch["subscribers"], ch["tier"], ch["country"], ch["email"],
                        ch["description"], "", "待发送", "Day0", "", TODAY,
                    ]
                    for col_idx, val in enumerate(row_data, 1):
                        ws.cell(row=last_row, column=col_idx, value=val)
                    ws.row_dimensions[last_row].height = 40
                wb.save(output_path)
                logger.info(f"✅ 追加 {len(new_channels)} 条到现有文件")
            else:
                logger.info("⚠ 无新增记录（全部已存在）")
            return
        except Exception as e:
            logger.warning(f"追加失败，重新生成：{e}")

    emails_found = build_excel(filtered, output_path)

    logger.info(f"\n{'='*60}")
    logger.info(f"完成！收集 {len(filtered)} 个 Kalshi KOL")
    logger.info(f"邮箱命中：{emails_found} 个（{emails_found/len(filtered)*100:.1f}%）")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    run()
