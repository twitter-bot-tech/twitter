#!/usr/bin/env python3
"""
BYDFi MoonX — YouTube KOL 自动收集脚本
使用 YouTube Data API v3（免费，10,000 units/day）
收集 Meme币 / 链上交易 / 加密货币 YouTube 频道
每次运行目标：20~50 个新 KOL
"""

import os
import re
import json
import time
import logging
import requests
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

load_dotenv(Path(__file__).parent.parent / ".env", override=True)
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY")
TODAY = datetime.now().strftime("%Y-%m-%d")
_SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = Path("/data") if os.getenv("FLY_APP_NAME") else _SCRIPT_DIR
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR = OUTPUT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "youtube_kol.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ── 搜索关键词（方案C：按月滚动，自动更新年份）──
def get_search_queries() -> list[str]:
    """动态生成搜索关键词：年份自动跟随当前年，每月轮换额外词组。"""
    now = datetime.now()
    year = now.year
    month = now.month

    base = [
        f"meme coin trading {year}",
        f"solana meme coins alpha {year}",
        f"crypto on-chain trading {year}",
        f"altcoin gems {year}",
        f"memecoin 100x {year}",
        f"solana dex sniper {year}",
        f"crypto alpha calls {year}",
        "pump.fun trading tutorial",
        "GMGN crypto trading",
        "dexscreener trading",
        "smart money crypto on-chain",
        "birdeye solana trading",
        "raydium dex trading",
        "meme coin calls crypto",
        "bnb chain meme coins",
        "pancakeswap trading",
        "crypto wallet tracking",
        "defi trading tutorial",
    ]

    # 每月额外词组（季节性热点）
    monthly_extras = {
        1:  [f"new year crypto picks {year}", f"january altcoin {year}"],
        2:  [f"february crypto rally {year}", f"altcoin season {year}"],
        3:  [f"crypto q1 review {year}", f"march memecoin {year}"],
        4:  [f"q2 crypto outlook {year}", f"solana ecosystem {year}"],
        5:  [f"may crypto gems {year}", f"altcoin may picks {year}"],
        6:  [f"mid year crypto review {year}", f"june crypto rally {year}"],
        7:  [f"q3 crypto outlook {year}", f"july memecoin {year}"],
        8:  [f"august crypto alpha {year}", f"summer crypto rally {year}"],
        9:  [f"september crypto gems {year}", f"q3 crypto review {year}"],
        10: [f"october crypto rally {year}", f"q4 crypto outlook {year}"],
        11: [f"november crypto picks {year}", f"bull run crypto {year}"],
        12: [f"year end crypto review {year}", f"december crypto gems {year}"],
    }

    return base + monthly_extras.get(month, [])


# 每页结果数（最大50）
RESULTS_PER_QUERY = 50

# 每日最低收集目标，不足则自动升级策略
TARGET_DAILY = 20

# ── 分级策略配置（标准 → 扩展 → 兜底）──
# active_days=0 表示不限时间
ESCALATION_CONFIGS = [
    {"name": "标准策略", "active_days": 180, "min_subscribers": 1000},
    {"name": "扩展策略", "active_days": 365, "min_subscribers": 500},
    {"name": "兜底策略", "active_days": 0,   "min_subscribers": 100},
]


def get_fallback_queries() -> list[list[str]]:
    """备用词组，仅在标准策略不足 TARGET_DAILY 时触发，每级使用不同词避免重复 API 调用。"""
    year = datetime.now().year
    return [
        # Level 2：扩展链和赛道（365天内活跃，订阅≥500）
        [
            f"ethereum meme coins {year}",
            f"base chain dex trading {year}",
            f"ton blockchain meme {year}",
            f"sui network trading {year}",
            f"avalanche avax meme {year}",
            f"arbitrum trading {year}",
            f"crypto technical analysis {year}",
            f"web3 trading signals {year}",
            f"altcoin portfolio {year}",
            f"crypto trading beginner {year}",
        ],
        # Level 3：兜底（不限时间，订阅≥100）
        [
            f"bitcoin trading signals {year}",
            f"crypto futures trading {year}",
            f"nft trading crypto {year}",
            "crypto yield farming tutorial",
            "dogecoin shiba inu trading",
            "crypto arbitrage tutorial",
            "crypto portfolio management",
            "how to trade crypto",
            "crypto passive income",
            "crypto day trading strategy",
        ],
    ]

QUERIES_POOL_FILE = _SCRIPT_DIR / "queries_pool.json"

# 加密领域核心词，用于识别有效词组
_CRYPTO_SEEDS = {
    "solana", "ethereum", "bitcoin", "bnb", "sui", "ton", "base", "arbitrum",
    "meme", "defi", "nft", "trading", "alpha", "whale", "onchain", "dex",
    "pump", "gem", "airdrop", "staking", "yield", "prediction", "polymarket",
    "kalshi", "futures", "leverage", "memecoin", "altcoin", "crypto", "web3",
    "blockchain", "token", "wallet", "smartmoney", "sniper", "launchpad",
}

def load_queries_pool() -> tuple[set[str], list[str]]:
    """读取持久化词库，返回 (已用词集合, 待用词列表)。"""
    if QUERIES_POOL_FILE.exists():
        data = json.loads(QUERIES_POOL_FILE.read_text(encoding="utf-8"))
        return set(data.get("used", [])), data.get("available", [])
    return set(), []


def save_queries_pool(used: set[str], available: list[str]):
    QUERIES_POOL_FILE.write_text(
        json.dumps({"used": sorted(used), "available": available},
                   ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def mine_queries_from_existing(n: int = 30) -> list[str]:
    """
    从已有 Excel 的频道名称 + 描述里挖高频词组，返回 n 个新搜索词。
    只返回包含加密核心词、且未在当前词库里出现过的词组。
    """
    used, _ = load_queries_pool()
    all_static = set(get_search_queries()) | {
        q for group in get_fallback_queries() for q in group
    }
    already_known = used | all_static

    texts: list[str] = []
    for fp in sorted(_SCRIPT_DIR.glob("MoonX_YouTube_KOL名单_*.xlsx")):
        try:
            wb = load_workbook(fp, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=3, values_only=True):
                name = str(row[1] or "")
                desc = str(row[8] or "")
                texts.append(f"{name} {desc}".lower())
            wb.close()
        except Exception:
            pass

    _STOPWORDS = {
        "the", "and", "for", "with", "about", "this", "that", "from", "our",
        "your", "all", "are", "have", "has", "how", "get", "can", "will",
        "not", "but", "more", "also", "new", "best", "top", "why", "what",
        "welcome", "here", "into", "use", "world", "content", "channel",
        "join", "like", "just", "you", "out", "its", "see", "one", "make",
    }

    phrase_counter: Counter = Counter()
    word_re = re.compile(r'\b[a-z][a-z0-9]{2,}\b')
    for text in texts:
        clean = re.sub(r'[^a-z0-9 ]', ' ', text)
        words = word_re.findall(clean)
        for i, w in enumerate(words):
            if w not in _CRYPTO_SEEDS:
                continue
            # 只和非停用词组合
            neighbors = []
            if i > 0 and words[i-1] not in _STOPWORDS:
                neighbors.append((words[i-1], w))
            if i + 1 < len(words) and words[i+1] not in _STOPWORDS:
                neighbors.append((w, words[i+1]))
            if (i + 2 < len(words)
                    and words[i+1] not in _STOPWORDS
                    and words[i+2] not in _STOPWORDS):
                neighbors.append((w, words[i+1], words[i+2]))
            for parts in neighbors:
                phrase_counter[" ".join(parts)] += 1

    new_queries: list[str] = []
    for phrase, cnt in phrase_counter.most_common(200):
        if cnt < 2:
            break
        parts = phrase.split()
        if phrase not in already_known and len(phrase) >= 8 and len(parts) == len(set(parts)):
            new_queries.append(phrase)
        if len(new_queries) >= n:
            break

    logger.info(f"🔍 自动扩词：从 {len(texts)} 个频道描述中挖掘到 {len(new_queries)} 个新词组")
    return new_queries


# 邮件黑名单域名
EMAIL_BLOCKLIST = {
    "example.com", "domain.com", "email.com", "gmail.com.fake",
    "sentry.io", "cloudflare.com", "wixpress.com", "squarespace.com",
    "youtube.com", "youtu.be", "google.com",
}
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
}


def extract_email(text: str) -> str:
    if not text:
        return ""
    # 还原混淆格式：partners(at)domain(dot)com → partners@domain.com
    normalized = re.sub(r'\s*\(at\)\s*', '@', text, flags=re.IGNORECASE)
    normalized = re.sub(r'\s*\(dot\)\s*', '.', normalized, flags=re.IGNORECASE)
    normalized = re.sub(r'\s*\[at\]\s*', '@', normalized, flags=re.IGNORECASE)
    normalized = re.sub(r'\s*\[dot\]\s*', '.', normalized, flags=re.IGNORECASE)
    for email in EMAIL_RE.findall(normalized):
        domain = email.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST and "." in domain:
            return email.lower()
    return ""


def scrape_channel_email(channel_url: str) -> str:
    """用 yt-dlp 获取频道完整 description，提取邮箱（支持混淆格式）"""
    try:
        import subprocess, json as _json
        result = subprocess.run(
            ["python3", "-m", "yt_dlp", "--dump-single-json", "--no-warnings", "--no-playlist",
             channel_url + "/about"],
            capture_output=True, text=True, timeout=15,
        )
        if result.returncode != 0 or not result.stdout.strip():
            return ""
        data = _json.loads(result.stdout)
        desc = data.get("description", "")
        return extract_email(desc)
    except Exception:
        return ""


def classify_tier(subscribers: int) -> str:
    if subscribers >= 1_000_000:
        return "A级"
    elif subscribers >= 100_000:
        return "B级"
    elif subscribers >= 10_000:
        return "C级"
    else:
        return "D级"


def search_channels_via_videos(youtube, query: str, max_results: int = 50,
                               published_after: str = None) -> list[dict]:
    """方案B：搜索近期视频，从视频反查频道，只收集近期活跃的创作者。"""
    try:
        params = dict(
            q=query,
            type="video",
            part="snippet",
            maxResults=max_results,
            relevanceLanguage="en",
            order="relevance",
        )
        if published_after:
            params["publishedAfter"] = published_after
        resp = youtube.search().list(**params).execute()
        seen = {}
        for item in resp.get("items", []):
            cid = item["snippet"]["channelId"]
            title = item["snippet"]["channelTitle"]
            if cid not in seen:
                seen[cid] = title
        return [{"channel_id": cid, "title": t} for cid, t in seen.items()]
    except HttpError as e:
        logger.error(f"视频搜索失败 ({query}): {e}")
        return []


def get_channel_details(youtube, channel_ids: list[str]) -> list[dict]:
    """批量获取频道详情（订阅数、描述、链接等）。"""
    if not channel_ids:
        return []
    try:
        resp = youtube.channels().list(
            id=",".join(channel_ids),
            part="snippet,statistics,brandingSettings",
        ).execute()
        results = []
        for item in resp.get("items", []):
            snippet = item.get("snippet", {})
            stats = item.get("statistics", {})
            branding = item.get("brandingSettings", {}).get("channel", {})

            channel_url = f"https://www.youtube.com/channel/{item['id']}"
            handle = snippet.get("customUrl", "")  # e.g. "@channelname"
            if handle:
                channel_url = f"https://www.youtube.com/{handle}"

            description = snippet.get("description", "")
            sub_count_str = stats.get("subscriberCount", "0")
            sub_count = int(sub_count_str) if sub_count_str.isdigit() else 0

            # 邮件：先从 description 提取，没有则抓 About 页面（仅 B级+ 值得抓）
            email = extract_email(description)
            if not email and sub_count >= 100_000:
                email = scrape_channel_email(channel_url)

            # 如有自定义链接 also check keywords in description for socials
            twitter_handle = ""
            twitter_match = re.search(
                r'twitter\.com/([A-Za-z0-9_]{1,50})|x\.com/([A-Za-z0-9_]{1,50})',
                description
            )
            if twitter_match:
                twitter_handle = "@" + (twitter_match.group(1) or twitter_match.group(2))

            results.append({
                "channel_id": item["id"],
                "name": snippet.get("title", ""),
                "handle": handle,
                "channel_url": channel_url,
                "subscribers": sub_count,
                "tier": classify_tier(sub_count),
                "description": description[:200],
                "email": email,
                "twitter": twitter_handle,
                "country": snippet.get("country", ""),
            })
        return results
    except HttpError as e:
        logger.error(f"获取频道详情失败: {e}")
        return []


def load_existing_channel_ids(output_dir: Path) -> set:
    """从所有已有的 YouTube KOL Excel 文件中读取已收集的 channel_id，去重用。"""
    seen = set()
    # 在 Fly.io 上同时扫描 /data/ 和 /app/ 以防止重复收集
    search_dirs = {output_dir, _SCRIPT_DIR} if os.getenv("FLY_APP_NAME") else {output_dir}
    all_files = [fp for d in search_dirs for fp in d.glob("MoonX_YouTube_KOL名单_*.xlsx")]
    for fp in all_files:
        try:
            wb = load_workbook(fp, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and row[0]:
                    seen.add(str(row[0]).strip())
            wb.close()
        except Exception as e:
            logger.warning(f"读取 {fp.name} 失败: {e}")
    return seen


def build_excel(channels: list[dict], output_path: Path):
    """生成 Excel 文件，格式与 Twitter KOL 名单一致。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube KOL名单"

    # 样式
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill("solid", fgColor="2E75B6")
    title_font = Font(bold=True, color="FFFFFF", size=13)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 第1行：标题
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"BYDFi MoonX — YouTube KOL 名单（{TODAY}）"
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # 第2行：表头
    headers = [
        "Channel ID", "频道名称", "YouTube 链接", "Twitter Handle",
        "订阅数", "分级", "国家/地区", "邮箱", "描述摘要",
        "联系状态", "备注", "收集日期"
    ]
    col_widths = [28, 25, 40, 22, 12, 8, 12, 35, 45, 12, 15, 14]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[chr(64 + col_idx)].width = w
    ws.row_dimensions[2].height = 22

    # 数据行
    tier_colors = {
        "A级": "FFD700",
        "B级": "C6EFCE",
        "C级": "DDEBF7",
        "D级": "F2F2F2",
    }
    for row_idx, ch in enumerate(channels, 3):
        row_data = [
            ch["channel_id"],
            ch["name"],
            ch["channel_url"],
            ch["twitter"],
            ch["subscribers"],
            ch["tier"],
            ch["country"],
            ch["email"],
            ch["description"],
            "",   # 联系状态
            "",   # 备注
            TODAY,
        ]
        fill_color = tier_colors.get(ch["tier"], "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 9))
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A3"
    wb.save(output_path)
    logger.info(f"✅ Excel 已保存: {output_path.name}（共 {len(channels)} 条）")


def run():
    if not YOUTUBE_API_KEY:
        logger.error("❌ 未找到 YOUTUBE_API_KEY，请在 .env.outreach 中添加")
        return

    logger.info("=" * 60)
    logger.info(f"🎬 YouTube KOL 收集开始 — {TODAY}  目标: ≥{TARGET_DAILY} 个")
    logger.info("=" * 60)

    youtube = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)

    existing_ids = load_existing_channel_ids(OUTPUT_DIR)
    logger.info(f"📋 已有记录: {len(existing_ids)} 个频道（自动去重）")

    # seen_ids 在整次运行中持续更新，防止跨策略重复
    seen_ids: set[str] = set(existing_ids)

    # 每级查询词：Level0 = 月度滚动词，Level1/2 = 备用词
    all_query_groups = [get_search_queries()] + get_fallback_queries()

    def _search_batch(queries: list[str], published_after: str | None) -> dict[str, str]:
        """搜索一批关键词，返回 {channel_id: title}（已过滤 seen_ids）。"""
        found = {}
        for query in queries:
            logger.info(f"   🔍 {query}")
            results = search_channels_via_videos(youtube, query, RESULTS_PER_QUERY, published_after)
            for ch in results:
                cid = ch["channel_id"]
                if cid not in seen_ids and cid not in found:
                    found[cid] = ch["title"]
            time.sleep(0.3)
        return found

    def _fetch_and_filter(channel_ids: dict[str, str], min_subscribers: int) -> list[dict]:
        """批量查频道详情并按订阅数过滤，同时把所有查询过的 ID 加入 seen_ids。"""
        details = []
        id_list = list(channel_ids.keys())
        for i in range(0, len(id_list), 50):
            batch = id_list[i:i + 50]
            details.extend(get_channel_details(youtube, batch))
            time.sleep(0.3)
        for ch in details:
            seen_ids.add(ch["channel_id"])
        return [ch for ch in details if ch["subscribers"] >= min_subscribers]

    # ── 分级执行，直到达到 TARGET_DAILY ──
    all_collected: list[dict] = []
    used_strategy = ESCALATION_CONFIGS[0]["name"]

    for level, (cfg, queries) in enumerate(zip(ESCALATION_CONFIGS, all_query_groups)):
        if level > 0:
            logger.info(
                f"\n⚡ 当前已收集 {len(all_collected)} 个，不足 {TARGET_DAILY}，"
                f"自动升级至【{cfg['name']}】"
            )
            time_range = "不限" if cfg["active_days"] == 0 else f"{cfg['active_days']}天内"
            logger.info(
                f"   订阅门槛: ≥{cfg['min_subscribers']:,}  时间范围: {time_range}"
            )
            used_strategy = cfg["name"]
        else:
            logger.info(f"\n▶ 【{cfg['name']}】— 订阅≥{cfg['min_subscribers']:,}，"
                        f"{cfg['active_days']}天内活跃，{len(queries)}个关键词")

        published_after = None
        if cfg["active_days"] > 0:
            published_after = (
                datetime.now() - timedelta(days=cfg["active_days"])
            ).strftime("%Y-%m-%dT00:00:00Z")

        batch_ids = _search_batch(queries, published_after)
        logger.info(f"   发现新频道: {len(batch_ids)} 个")

        if batch_ids:
            new_channels = _fetch_and_filter(batch_ids, cfg["min_subscribers"])
            all_collected.extend(new_channels)
            logger.info(f"   本级新增: {len(new_channels)} 个（累计 {len(all_collected)} 个）")

        if len(all_collected) >= TARGET_DAILY:
            logger.info(f"   ✅ 达到目标 {TARGET_DAILY}，停止升级")
            break

    if len(all_collected) < TARGET_DAILY:
        # Level 4：自动扩词
        logger.info(
            f"\n🤖 【自动扩词】当前 {len(all_collected)} 个，启动关键词自动挖掘..."
        )
        used_pool, available_pool = load_queries_pool()

        # 补充可用词库（如果不够则重新挖掘）
        if len(available_pool) < 10:
            available_pool = mine_queries_from_existing(n=30)

        # 取前10个未用过的词
        batch_queries = []
        remaining_pool = []
        for q in available_pool:
            if q not in used_pool and len(batch_queries) < 10:
                batch_queries.append(q)
            else:
                remaining_pool.append(q)

        if batch_queries:
            logger.info(f"   使用自动词组: {batch_queries}")
            batch_ids = _search_batch(batch_queries, None)
            logger.info(f"   发现新频道: {len(batch_ids)} 个")
            if batch_ids:
                new_channels = _fetch_and_filter(batch_ids, 100)
                all_collected.extend(new_channels)
                logger.info(f"   本级新增: {len(new_channels)} 个（累计 {len(all_collected)} 个）")
            used_pool.update(batch_queries)
            save_queries_pool(used_pool, remaining_pool)
            used_strategy = "自动扩词策略"

    if len(all_collected) < TARGET_DAILY:
        logger.warning(
            f"⚠ 四级策略均已执行，最终收集 {len(all_collected)} 个"
            f"（低于目标 {TARGET_DAILY}）"
        )

    if not all_collected:
        logger.info("⚠ 没有符合条件的频道，结束")
        return

    # 运行内去重（同一频道可能出现在多个策略级别）
    seen_in_run: set[str] = set()
    filtered: list[dict] = []
    for ch in all_collected:
        if ch["channel_id"] not in seen_in_run:
            seen_in_run.add(ch["channel_id"])
            filtered.append(ch)

    filtered.sort(key=lambda x: x["subscribers"], reverse=True)

    # 统计
    tiers: dict[str, int] = {"A级": 0, "B级": 0, "C级": 0, "D级": 0}
    emails_found = 0
    for ch in filtered:
        tiers[ch["tier"]] = tiers.get(ch["tier"], 0) + 1
        if ch["email"]:
            emails_found += 1

    logger.info(f"\n📈 分级统计（最终策略：{used_strategy}）:")
    for tier, count in tiers.items():
        logger.info(f"   {tier}: {count} 个")
    logger.info(f"   邮箱已找到: {emails_found} 个")

    # 保存 Excel（保留以向后兼容）
    output_path = OUTPUT_DIR / f"MoonX_YouTube_KOL名单_{TODAY}.xlsx"
    build_excel(filtered, output_path)

    # 同步写入 SQLite
    try:
        from kol_db import upsert_kol, init_db
        init_db()
        db_count = 0
        for ch in filtered:
            upsert_kol({
                "source":       "youtube",
                "channel_id":   ch["channel_id"],
                "channel_url":  ch["channel_url"],
                "name":         ch["name"],
                "platform":     "YouTube",
                "subscribers":  ch["subscribers"],
                "tier":         ch["tier"],
                "email":        ch.get("email", ""),
                "description":  ch.get("description", ""),
                "twitter":      ch.get("twitter", ""),
                "country":      ch.get("country", ""),
                "collect_date": TODAY,
            })
            db_count += 1
        logger.info(f"   SQLite 已同步: {db_count} 条")
    except Exception as e:
        logger.warning(f"   SQLite 写入失败（不影响 Excel）: {e}")

    logger.info(f"\n{'='*60}")
    logger.info(f"完成！共收集 {len(filtered)} 个 YouTube KOL（策略：{used_strategy}）")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    import sys as _sys; _sys.path.insert(0, str(Path(__file__).parent.parent))
    from scripts.pid_lock import acquire_lock; acquire_lock()
    run()
