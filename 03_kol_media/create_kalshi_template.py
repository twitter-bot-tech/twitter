#!/usr/bin/env python3
"""生成 Kalshi 前KOL 专项名单 Excel 模板"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from datetime import datetime

TODAY = datetime.now().strftime("%Y-%m-%d")
OUTPUT = Path(__file__).parent / f"MoonX_Kalshi_KOL名单_{TODAY}.xlsx"

# ── 颜色 ─────────────────────────────────────────────────────────────
NAVY    = "1F4E79"
GOLD    = "B8860B"
LGOLD   = "FFF2CC"
RED     = "C00000"
LRED    = "FFCCCC"
GREEN   = "375623"
LGREEN  = "E2EFDA"
GRAY    = "F2F2F2"
WHITE   = "FFFFFF"
ORANGE  = "E26B0A"

def make_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, value, bg, fg="FFFFFF", size=10, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = make_border()
    return c

def data_cell(ws, row, col, value="", bg=WHITE, color="000000", bold=False, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(color=color, size=9, name="Arial", bold=bold)
    c.alignment = Alignment(vertical="center", wrap_text=wrap)
    c.border = make_border()
    return c


wb = Workbook()
ws = wb.active
ws.title = "Kalshi前KOL名单"

# ── 第1行：标题 ────────────────────────────────────────────────────
ws.merge_cells("A1:N1")
title = ws["A1"]
title.value = f"BYDFi MoonX — Kalshi 前KOL 专项名单（{TODAY}）"
title.fill = PatternFill("solid", fgColor=NAVY)
title.font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
title.alignment = Alignment(horizontal="center", vertical="center")
title.border = make_border()
ws.row_dimensions[1].height = 30

# ── 第2行：说明 ────────────────────────────────────────────────────
ws.merge_cells("A2:N2")
note = ws["A2"]
note.value = (
    "⚡ 最高优先级名单 | 返佣：合约65% · Meme50% · 预测市场40% | "
    "Kalshi于2026年2月砍掉X平台合作，这批KOL正在寻找新平台 | "
    "发送前必须手动填写【个性化钩子】列"
)
note.fill = PatternFill("solid", fgColor=LGOLD)
note.font = Font(color=GOLD, size=9, name="Arial", bold=True)
note.alignment = Alignment(horizontal="left", vertical="center")
note.border = make_border()
ws.row_dimensions[2].height = 22

# ── 第3行：列头 ────────────────────────────────────────────────────
HEADERS = [
    ("A", 6,  "#",              NAVY),
    ("B", 22, "KOL 名称",        NAVY),
    ("C", 16, "平台",            NAVY),
    ("D", 38, "主页链接",         NAVY),
    ("E", 14, "粉丝数",          NAVY),
    ("F", 10, "分级",            NAVY),
    ("G", 12, "国家/地区",       NAVY),
    ("H", 35, "邮箱",            NAVY),
    ("I", 45, "内容方向（描述）", NAVY),
    ("J", 40, "个性化钩子 ✏️",   RED),   # 必填，红色提醒
    ("K", 14, "联系状态",        NAVY),
    ("L", 18, "跟进轮次",        NAVY),
    ("M", 25, "备注",            NAVY),
    ("N", 14, "收集日期",        NAVY),
]

for col_letter, width, header, bg in HEADERS:
    col_idx = ord(col_letter) - 64
    hdr(ws, 3, col_idx, header, bg)
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[3].height = 24

# ── 数据验证：联系状态下拉 ─────────────────────────────────────────
dv_status = DataValidation(
    type="list",
    formula1='"待发送,已发送,已回复,谈判中,已签约,无效"',
    allow_blank=True,
    showDropDown=False,
)
ws.add_data_validation(dv_status)

dv_platform = DataValidation(
    type="list",
    formula1='"YouTube,Twitter/X,Substack,Newsletter,播客,其他"',
    allow_blank=True,
    showDropDown=False,
)
ws.add_data_validation(dv_platform)

dv_followup = DataValidation(
    type="list",
    formula1='"Day0,Day7,Day14,完成"',
    allow_blank=True,
    showDropDown=False,
)
ws.add_data_validation(dv_followup)

# ── 示例数据行（第4行） ───────────────────────────────────────────
EXAMPLE_BG = "EBF3FB"
EXAMPLE = [
    (1,  "A", "（示例）Kalshi Creator",  EXAMPLE_BG),
    (2,  "B", "KOL Demo",               EXAMPLE_BG),
    (3,  "C", "YouTube",                EXAMPLE_BG),
    (4,  "D", "https://youtube.com/@demo", EXAMPLE_BG),
    (5,  "E", 85000,                    EXAMPLE_BG),
    (6,  "F", "Kalshi",                 EXAMPLE_BG),
    (7,  "G", "US",                     EXAMPLE_BG),
    (8,  "H", "creator@example.com",    EXAMPLE_BG),
    (9,  "I", "主要做 Kalshi 预测市场教程，覆盖选举/经济指标市场，受众以美国散户为主", EXAMPLE_BG),
    (10, "J", '你在上期视频里说 Kalshi 取消合作"挺突然的"——我们正好在找有这块受众的合作伙伴', LRED),
    (11, "K", "待发送",                  EXAMPLE_BG),
    (12, "L", "Day0",                   EXAMPLE_BG),
    (13, "M", "",                        EXAMPLE_BG),
    (14, "N", TODAY,                    EXAMPLE_BG),
]

for col_idx, _, value, bg in EXAMPLE:
    c = data_cell(ws, 4, col_idx, value, bg=bg, wrap=(col_idx in [9, 10]))
    if col_idx == 10:  # 个性化钩子列加红色提示
        c.font = Font(color=RED, size=9, name="Arial", italic=True)

# 应用数据验证范围
dv_status.add("K4:K1000")
dv_platform.add("C4:C1000")
dv_followup.add("L4:L1000")

ws.row_dimensions[4].height = 50

# ── 第5~20行：空白填写区 ──────────────────────────────────────────
for row_idx in range(5, 21):
    bg = WHITE if row_idx % 2 == 0 else GRAY
    for col_idx in range(1, 15):
        c = data_cell(ws, row_idx, col_idx, bg=bg)
        if col_idx == 10:
            c.fill = PatternFill("solid", fgColor="FFF0F0")  # 钩子列淡红底
    # 默认状态
    ws.cell(row=row_idx, column=11).value = "待发送"
    ws.cell(row=row_idx, column=12).value = "Day0"
    ws.cell(row=row_idx, column=14).value = TODAY
    ws.row_dimensions[row_idx].height = 40

# ── 冻结前3行 ────────────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── 第二个Sheet：搜索策略参考 ─────────────────────────────────────
ws2 = wb.create_sheet("搜索策略")
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 70

ws2.merge_cells("A1:B1")
ws2["A1"].value = "Kalshi 前KOL 搜索策略参考"
ws2["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws2["A1"].font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

search_data = [
    ("渠道", "搜索词 / 方法"),
    ("Twitter/X 搜索", '"kalshi affiliate" OR "kalshi partner" since:2025-01-01'),
    ("Twitter/X 搜索", '"kalshi" "referral" OR "ambassador"'),
    ("Twitter/X 搜索", '"lost kalshi" OR "kalshi cancelled" OR "kalshi ended"'),
    ("Google 搜索", 'site:youtube.com "kalshi" "prediction market" review 2025'),
    ("Google 搜索", '"kalshi ambassador" OR "kalshi creator program" site:youtube.com'),
    ("YouTube 搜索", "kalshi tutorial 2025 / kalshi trading how to"),
    ("Substack 搜索", "kalshi site:substack.com"),
    ("直接联系", "查看 Kalshi 官网历史 Blog/Press 页，找提到的 creator 名字"),
]

for r, (channel, query) in enumerate(search_data, 2):
    bg = GRAY if r % 2 == 0 else WHITE
    bold = (r == 2)
    c1 = ws2.cell(row=r, column=1, value=channel)
    c2 = ws2.cell(row=r, column=2, value=query)
    for c in [c1, c2]:
        c.fill = PatternFill("solid", fgColor=(LGOLD if bold else bg))
        c.font = Font(bold=bold, size=9, name="Arial")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = make_border()
    ws2.row_dimensions[r].height = 22

ws2.freeze_panes = "A3"

# ── 第三个Sheet：邮件模板参考 ────────────────────────────────────
ws3 = wb.create_sheet("邮件模板参考")
ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 80

ws3.merge_cells("A1:B1")
ws3["A1"].value = "Kalshi前KOL 邮件模板（参考，个性化钩子请手动填写）"
ws3["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws3["A1"].font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

template_rows = [
    ("主题行①", "Founding partner offer — 65% rev-share on futures"),
    ("主题行②", "We offer 65% rev-share on futures — not a typo"),
    ("主题行③", "Founding partner terms: up to 65% revenue share"),
    ("", ""),
    ("正文模板", (
        "Hi [Name],\n\n"
        "[个性化钩子 — 引用他们最近1条具体内容，证明你读过]\n\n"
        "Quick context: I run marketing at BYDFi MoonX — a prediction market aggregator "
        "that pulls live data from Polymarket, Kalshi, and on-chain markets into one feed.\n\n"
        "Here's why I'm reaching out specifically to you: we're opening founding partner slots, "
        "and the terms are different from standard:\n\n"
        "  Futures trading        → 65% rev-share to you\n"
        "  Meme/on-chain trading  → 50%\n"
        "  Prediction markets     → 40%\n\n"
        "Industry standard is 20~30%. We can offer this because we want the right partners, "
        "not the most partners.\n\n"
        "One question: what market are you covering next? I can pull the smart money flow "
        "data on that topic and send it to you before you publish.\n\n"
        "Kelly · Head of Marketing · BYDFi MoonX\n"
        "https://www.bydfi.com/en/moonx/markets/trending\n"
        "TG: @BDkelly"
    )),
]

for r, (label, content) in enumerate(template_rows, 2):
    bg = LGOLD if label.startswith("主题") else (WHITE if label == "" else "F0F8FF")
    c1 = ws3.cell(row=r, column=1, value=label)
    c2 = ws3.cell(row=r, column=2, value=content)
    for c in [c1, c2]:
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9, name="Arial", bold=(label.startswith("主题")))
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = make_border()
    ws3.row_dimensions[r].height = 18 if label.startswith("主题") or label == "" else 160

# ── 保存 ─────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"✅ 模板已生成：{OUTPUT.name}")
print(f"   路径：{OUTPUT}")
