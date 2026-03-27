#!/usr/bin/env python3
"""
生成更新版 Meme 币交易 KOL 名单 v2
保留已验证旧名单 + 补充新真实 KOL
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

# 已验证的已有 KOL（保留状态）
EXISTING_KOLS = [
    # name, handle, followers, focus, platform, region, contact, status
    ("Ansem", "@blknoiz06", "~600K", "Solana Meme币/pump.fun链上交易", "Axiom/GMGN/pump.fun", "美国", "Twitter DM", "DM已关闭"),
    ("Murad Mahmudov", "@MustStopMurad", "~500K", "Meme币周期/Solana生态", "通用Meme", "中亚/美国", "Twitter DM", "已发送DM"),
    ("Hsaka", "@HsakaTrades", "~200K", "链上交易/Solana技术分析", "通用", "英国", "Twitter DM", "已发送DM"),
    ("Kaleo", "@K_A_L_E_O", "~800K", "加密/Meme币周期时机", "通用", "匿名", "Twitter DM", "DM已关闭"),
    ("AltcoinSherpa", "@AltcoinSherpa", "~400K", "山寨币/Meme币综合", "通用", "匿名", "Twitter DM", "已发送DM"),
    ("Lookonchain", "@lookonchain", "~1.2M", "链上鲸鱼追踪/GMGN类工具", "GMGN", "匿名", "contact@lookonchain.com", "已发送邮件"),
    ("Frank DeGods", "@frankdegods", "~380K", "Solana NFT→Meme生态", "Axiom/pump.fun", "美国", "Twitter DM", "已发送DM"),
    ("Degenharambe", "@degenharambe", "~120K", "Banana Gun/Trojan/Solana小市值", "Banana Gun/Trojan", "匿名", "Twitter DM", "已发送DM"),
    ("Milk Road", "@MilkRoadDaily", "~400K", "加密新闻/链上工具趋势", "通用", "美国", "hello@milkroad.com", "已发送邮件"),
    ("WhalePanda", "@WhalePanda", "~250K", "加密老兵/Meme币周期评论", "通用", "匿名", "Twitter DM", "已发送DM"),
    ("CryptoNobler", "@CryptoNobler", "~80K", "GMGN/BullX Alpha分享", "GMGN/BullX", "匿名", "Twitter DM", "已发送DM"),
    ("SolanaFloor", "@SolanaFloor", "~120K", "Axiom/链上数据/Meme地板追踪", "Axiom", "匿名", "Twitter DM", "已发送DM"),
    ("CryptoGodJohn", "@CryptoGodJohn", "~300K", "Meme币狙击工具", "通用", "美国", "Twitter DM", "已发送DM"),
    ("Gainzy", "@gainzy222", "~180K", "Meme币交易/pump.fun", "pump.fun", "美国", "Twitter DM", "已发送DM"),
    ("Airdrop Alert", "@AirdropAlert", "~350K", "加密工具/空投/链上", "通用", "荷兰", "info@airdropalert.com", "已发送邮件"),
    ("TheCryptoDog", "@TheCryptoDog", "~700K", "加密市场评论/Meme币", "通用", "美国", "Twitter DM", "待联系"),
    ("KoroushAK", "@KoroushAK", "~500K", "加密交易策略/Meme币", "通用", "英国", "Twitter DM", "待联系"),
    ("Nathan Worsley", "@nathanworsley_", "~80K", "pump.fun/Solana Meme", "pump.fun", "匿名", "Twitter DM", "待联系"),
    ("Crypto Daku", "@crypto_daku", "~200K", "东南亚加密/Meme币", "通用", "东南亚", "Twitter DM", "DM已关闭"),
]

# 新增真实 KOL
NEW_KOLS = [
    # name, handle, followers, focus, platform, region, contact, status
    ("Ran Neuner", "@cryptomanran", "~950K", "市场分析/新兴项目/Crypto Banter主播", "通用", "南非", "info@cryptobanter.com", "待联系"),
    ("Miles Deutscher", "@milesdeutscher", "~600K", "山寨币/Meme币分析/早期趋势", "通用", "澳大利亚", "info@cryptobanter.com", "待联系"),
    ("Cobie", "@cobie", "~680K", "加密市场分析/UpOnly联合主播", "通用", "英国", "Twitter DM", "待联系"),
    ("Pentoshi", "@Pentosh1", "~480K", "链上分析/Solana/山寨币交易", "GMGN/Axiom", "匿名", "Twitter DM", "待联系"),
    ("EllioTrades", "@elliotrades", "~800K", "NFT/DeFi/Meme币/YouTube博主", "通用", "美国", "Twitter DM", "待联系"),
    ("Lark Davis (Crypto Lark)", "@TheCryptoLark", "~1M", "市场趋势/山寨币/交易教育", "通用", "新西兰", "Twitter DM", "待联系"),
    ("Crypto Rover", "@rovercrc", "~2.1M", "山寨币/Meme币炒作/零售受众", "通用", "匿名", "Twitter DM", "待联系"),
    ("CryptoWendyO", "@CryptoWendyO", "~700K", "技术分析/交易策略/Meme币", "通用", "美国", "Twitter DM", "待联系"),
    ("Andrew Kang", "@Rewkang", "~380K", "宏观加密/DeFi/Solana叙事", "通用", "美国", "Twitter DM", "待联系"),
    ("Loomdart", "@Loomdart", "~200K", "Meme文化/市场洞察/NFT", "通用", "匿名", "Twitter DM", "待联系"),
    ("Alon (pump.fun联创)", "@a1lon9", "~150K", "pump.fun生态/Solana代币发行Alpha", "pump.fun", "美国", "Twitter DM", "待联系"),
    ("Kmoney", "@Kmoney_69", "~100K", "Meme币倡导/MOG/PEPE/TRUMP", "通用", "匿名", "Twitter DM", "待联系"),
    ("0xVonGogh", "@0xVonGogh", "~80K", "Meme币全职交易/实战建议", "BullX/Axiom", "匿名", "Twitter DM", "待联系"),
    ("NotChaseColeman", "@NotChaseColeman", "~80K", "Meme币超级周期/GIGA倡导", "通用", "匿名", "Twitter DM", "待联系"),
    ("OnChain Wizard", "@OnChainWizard", "~80K", "DeFi套利/链上分析/聪明钱追踪", "GMGN", "匿名", "Twitter DM", "待联系"),
    ("Altcoin Daily", "@AltcoinDailyio", "~1.8M", "日常加密新闻/Meme币/Solana", "通用", "美国", "team@altcoindaily.co", "待联系"),
    ("Layah Heilpern", "@LayahHeilpern", "~700K", "比特币叙事/加密自由主义", "通用", "英国", "layahheilpern.com", "待联系"),
    ("Degen News", "@DegenNews", "~150K", "Solana NFT/Meme币突发新闻/链上", "通用", "匿名", "Twitter DM", "待联系"),
    ("Nansen", "@nansen_ai", "~270K", "链上聪明钱追踪/Meme交易者常引用", "GMGN类", "新加坡", "hello@nansen.ai", "待联系"),
    ("SmallCapScientist", "@SmallCapScientist", "~150K", "小市值/早期Meme币发现", "GMGN/BullX", "匿名", "Twitter DM", "待联系"),
    ("Algod", "@AlgodTrading", "~300K", "链上交易策略/Meme币Alpha", "通用", "匿名", "Twitter DM", "待联系"),
    ("Inversebrah", "@inversebrah", "~200K", "反向交易哲学/Meme币周期", "通用", "匿名", "Twitter DM", "待联系"),
    ("Daan Crypto", "@DaanCrypto", "~350K", "技术分析/Meme币趋势", "通用", "荷兰", "Twitter DM", "待联系"),
    ("ZachXBT", "@zachxbt", "~750K", "链上调查/钱包追踪/诈骗曝光", "GMGN类", "匿名", "Twitter DM", "待联系"),
    ("0xMert", "@0xMert_", "~200K", "Helius/Solana基础设施/Meme生态", "Solana", "匿名", "Twitter DM", "待联系"),
]

STATUS_COLORS = {
    "已发送DM": "C6EFCE",
    "已发送邮件": "C6EFCE",
    "DM已关闭": "FFE699",
    "待联系": "DDEBF7",
    "用户不存在": "FFC7CE",
}

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "KOL名单v2"

    # 大标题
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = f"Meme 币交易 KOL 名单 v2 — {datetime.now().strftime('%Y-%m-%d')} 更新"
    title.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title.font = Font(color="FFFFFF", bold=True, size=13)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 表头
    headers = ["#", "KOL 名字", "Twitter Handle", "粉丝数", "内容方向", "关联平台", "地区", "联系方式", "状态"]
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = header_fill
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 20

    # 分组标题
    def write_section_header(row, label):
        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        c.font = Font(bold=True, size=10, color="1F4E79")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 18

    def write_kols(start_row, kols, num_offset=0):
        for i, kol in enumerate(kols):
            row = start_row + i
            name, handle, followers, focus, platform, region, contact, status = kol
            ws.cell(row=row, column=1, value=num_offset + i + 1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=handle).font = Font(color="0070C0")
            ws.cell(row=row, column=4, value=followers).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=focus[:40])
            ws.cell(row=row, column=6, value=platform)
            ws.cell(row=row, column=7, value=region).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=8, value=contact)
            status_cell = ws.cell(row=row, column=9, value=status)
            status_cell.alignment = Alignment(horizontal="center")
            color = STATUS_COLORS.get(status, "FFFFFF")
            status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = thin_border()

    # 写已有名单
    write_section_header(3, f"▌ 已联系/跟进中 ({len(EXISTING_KOLS)} 个)")
    write_kols(4, EXISTING_KOLS, 0)

    # 写新名单
    new_start = 4 + len(EXISTING_KOLS) + 1
    write_section_header(new_start - 1, f"▌ 新增待联系 ({len(NEW_KOLS)} 个)")
    write_kols(new_start, NEW_KOLS, len(EXISTING_KOLS))

    # 统计行
    total_row = new_start + len(NEW_KOLS) + 1
    ws.merge_cells(f"A{total_row}:I{total_row}")
    summary = ws.cell(row=total_row, column=1)
    summary.value = f"共 {len(EXISTING_KOLS) + len(NEW_KOLS)} 个 KOL | 已联系: {sum(1 for k in EXISTING_KOLS if '已发送' in k[7])} | 待联系: {sum(1 for k in NEW_KOLS if k[7]=='待联系')} | 可发邮件: {sum(1 for k in EXISTING_KOLS+NEW_KOLS if '@' in k[6] or '.com' in k[6])}"
    summary.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    summary.font = Font(bold=True, size=9, color="375623")
    summary.alignment = Alignment(horizontal="center")

    # 列宽
    col_widths = [4, 22, 22, 10, 38, 18, 10, 28, 12]
    col_letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    for letter, w in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = w

    # 冻结前两行
    ws.freeze_panes = "A3"

    # 颜色图例 sheet
    ws2 = wb.create_sheet("图例说明")
    ws2["A1"] = "状态颜色说明"
    ws2["A1"].font = Font(bold=True, size=12)
    legend = [
        ("已发送DM / 已发送邮件", "C6EFCE", "已完成初步联系"),
        ("DM已关闭", "FFE699", "无法通过DM联系，可尝试邮件"),
        ("待联系", "DDEBF7", "新增，尚未联系"),
        ("用户不存在", "FFC7CE", "账号无效，需替换"),
    ]
    for i, (label, color, note) in enumerate(legend, 3):
        ws2.cell(row=i, column=1, value=label).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws2.cell(row=i, column=2, value=note)

    fname = "BYDFi_MoonX_Meme_KOL名单_v2.xlsx"
    wb.save(fname)
    print(f"✅ 已生成: {fname}")
    print(f"   已有名单: {len(EXISTING_KOLS)} 个（含状态）")
    print(f"   新增名单: {len(NEW_KOLS)} 个")
    print(f"   合计: {len(EXISTING_KOLS)+len(NEW_KOLS)} 个 KOL")
    print()
    print("📧 可发邮件的新增 KOL：")
    for k in NEW_KOLS:
        if "@" in k[6] or ".com" in k[6]:
            print(f"   {k[0]:25} {k[2]:20} {k[6]}")

if __name__ == "__main__":
    generate_excel()
