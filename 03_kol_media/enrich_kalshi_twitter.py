#!/usr/bin/env python3
"""
Kalshi KOL 补全脚本：自动查找 Twitter Handle + 从 Twitter Bio 提取邮箱
三层策略：
  1. YouTube brandingSettings.links（最准，官方挂的社交链接）
  2. SerpAPI Google 搜索 "[名字] site:twitter.com"（覆盖没挂链接的）
  3. twitter_scraper 拿到 handle 后获取 bio → 提取邮箱
"""

import os, re, time, logging, sys
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
import requests
sys.path.insert(0, str(Path(__file__).parent.parent))
import twitter_scraper
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

load_dotenv(Path(__file__).parent.parent / ".env", override=True)
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

YOUTUBE_API_KEY    = os.getenv("YOUTUBE_API_KEY")
SERPAPI_KEY        = os.getenv("SERPAPI_KEY")
OUTPUT_DIR         = Path(__file__).parent
LOG_DIR            = OUTPUT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "kalshi_enrich.log"),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)

EMAIL_BLOCKLIST = {
    "example.com", "domain.com", "sentry.io", "cloudflare.com",
    "youtube.com", "google.com", "twitter.com", "t.co",
}
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
TWITTER_RE = re.compile(
    r'(?:twitter\.com|x\.com)/(?!intent|share|home|search|hashtag|i/)([A-Za-z0-9_]{1,50})'
)


def extract_email(text: str) -> str:
    if not text:
        return ""
    t = re.sub(r'\s*[\(\[]at[\)\]]\s*', '@', text, flags=re.IGNORECASE)
    t = re.sub(r'\s*[\(\[]dot[\)\]]\s*', '.', t, flags=re.IGNORECASE)
    for email in EMAIL_RE.findall(t):
        domain = email.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST and "." in domain:
            return email.lower()
    return ""


# ── 第1层：YouTube brandingSettings.links ─────────────────────────

def get_youtube_social_links(channel_id: str) -> str:
    """从 YouTube API brandingSettings 获取频道挂的社交链接，提取 Twitter handle"""
    try:
        youtube = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)
        resp = youtube.channels().list(
            id=channel_id,
            part="brandingSettings",
        ).execute()
        items = resp.get("items", [])
        if not items:
            return ""
        links = items[0].get("brandingSettings", {}).get("channel", {}).get("links", [])
        for link in links:
            url = link.get("url", "")
            m = TWITTER_RE.search(url)
            if m:
                handle = m.group(1)
                if handle.lower() not in ("intent", "share", "home"):
                    return "@" + handle
        return ""
    except HttpError as e:
        logger.warning(f"YouTube brandingSettings 查询失败 ({channel_id}): {e}")
        return ""


# ── 第2层：SerpAPI Google 搜索 ────────────────────────────────────

def serpapi_find_twitter(name: str, channel_url: str) -> str:
    """用 SerpAPI 搜 '[名字] twitter' 找 handle"""
    if not SERPAPI_KEY:
        return ""
    try:
        # 搜索两个角度：名字 + twitter，以及频道URL + twitter
        queries = [
            f'"{name}" site:twitter.com OR site:x.com',
            f'"{name}" twitter prediction market',
        ]
        for query in queries:
            resp = requests.get(
                "https://serpapi.com/search.json",
                params={"q": query, "api_key": SERPAPI_KEY, "num": 5},
                timeout=10,
            )
            if resp.status_code != 200:
                continue
            data = resp.json()
            # 从 organic_results 的 link 里找 twitter.com/handle
            for result in data.get("organic_results", []):
                url = result.get("link", "")
                m = TWITTER_RE.search(url)
                if m:
                    handle = m.group(1)
                    if handle.lower() not in ("intent", "share", "home", "search"):
                        logger.info(f"    SerpAPI 找到: @{handle}")
                        return "@" + handle
                # 也搜标题里
                snippet = result.get("snippet", "") + result.get("title", "")
                m = TWITTER_RE.search(snippet)
                if m:
                    handle = m.group(1)
                    if handle.lower() not in ("intent", "share", "home"):
                        logger.info(f"    SerpAPI snippet 找到: @{handle}")
                        return "@" + handle
            time.sleep(1)
        return ""
    except Exception as e:
        logger.warning(f"SerpAPI 搜索失败 ({name}): {e}")
        return ""


# ── 第3层：twitter_scraper 拿 bio 提取邮箱 ───────────────────────


def get_twitter_bio_email(handle: str) -> tuple:
    """给定 @handle，返回 (bio文本, 邮箱)；失败时返回空"""
    try:
        username = handle.lstrip("@")
        user = twitter_scraper.get_user(username)
        if not user:
            return "", ""
        bio = user.get("description", "") or ""
        email = extract_email(bio)
        if not email:
            entities = user.get("entities", {})
            for url_obj in entities.get("description", {}).get("urls", []):
                expanded = url_obj.get("expanded_url", "")
                email = extract_email(expanded)
                if email:
                    break
        return bio[:200], email
    except Exception as e:
        logger.warning(f"Twitter bio 查询失败 (@{handle}): {e}")
        return "", ""


# ── 主流程 ────────────────────────────────────────────────────────

def enrich_kalshi_excel():
    # 找最新的 Kalshi 名单
    files = sorted(OUTPUT_DIR.glob("MoonX_Kalshi_KOL名单_*.xlsx"), reverse=True)
    if not files:
        logger.error("❌ 未找到 Kalshi KOL 名单 Excel")
        return

    excel_path = files[0]
    logger.info(f"📂 读取：{excel_path.name}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # 列索引（对应 build_excel 的顺序）
    # A=1:channel_id  B=2:name  C=3:channel_url  D=4:twitter
    # E=5:subs  F=6:tier  G=7:country  H=8:email  I=9:desc
    # J=10:钩子  K=11:状态  L=12:跟进  M=13:备注  N=14:日期
    COL_CHANNEL_ID = 1
    COL_NAME       = 2
    COL_CHANNEL_URL= 3
    COL_TWITTER    = 4
    COL_EMAIL      = 8

    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=4, values_only=False):
        channel_id  = str(row[COL_CHANNEL_ID - 1].value or "").strip()
        name        = str(row[COL_NAME       - 1].value or "").strip()
        channel_url = str(row[COL_CHANNEL_URL- 1].value or "").strip()
        twitter     = str(row[COL_TWITTER    - 1].value or "").strip()
        email       = str(row[COL_EMAIL      - 1].value or "").strip()

        # 跳过无效行、示例行、已有完整信息的行
        if not channel_id or name == "KOL Demo":
            continue
        if twitter and email:
            skipped += 1
            continue

        logger.info(f"\n🔍 处理: {name} ({channel_id})")

        # ── 层1：YouTube brandingSettings ──
        if not twitter and channel_id:
            logger.info("  层1: YouTube brandingSettings...")
            twitter = get_youtube_social_links(channel_id)
            if twitter:
                logger.info(f"  ✓ 层1 找到: {twitter}")

        # ── 层2：SerpAPI ──
        if not twitter:
            logger.info("  层2: SerpAPI Google 搜索...")
            twitter = serpapi_find_twitter(name, channel_url)

        # ── 层3：Twitter bio → 邮箱 ──
        if twitter and not email:
            logger.info(f"  层3: 查 Twitter bio ({twitter})...")
            bio, email = get_twitter_bio_email(twitter)
            if email:
                logger.info(f"  ✓ 邮箱: {email}")
            elif bio:
                logger.info(f"  bio 无邮箱: {bio[:60]}...")

        # 写回 Excel
        if twitter:
            row[COL_TWITTER - 1].value = twitter
        if email:
            row[COL_EMAIL - 1].value = email

        if twitter or email:
            updated += 1
            logger.info(f"  → 已更新: Twitter={twitter or '无'} | Email={email or '无'}")
        else:
            logger.info(f"  → 未找到任何联系方式")

        time.sleep(0.5)  # 避免触发限速

    wb.save(excel_path)
    logger.info(f"\n{'='*60}")
    logger.info(f"✅ 完成！更新 {updated} 条，跳过 {skipped} 条（已有完整信息）")
    logger.info(f"📊 Excel 已保存：{excel_path.name}")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    enrich_kalshi_excel()
