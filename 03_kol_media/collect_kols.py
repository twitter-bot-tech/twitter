#!/usr/bin/env python3
"""
自动收集 KOL — 加密货币行业全覆盖（预测市场 / Meme币 / DeFi / 链上数据 / 主流币等）
分级：A 级（100万+）/ B 级（10~100万）/ C 级（1~10万）/ D 级（1000~1万）
每日目标：30 个新 KOL

验证规则（2026-03-09 更新）：
1. 账号真实存在（API 返回有效 user_id）
2. 粉丝数 >= 1000
3. 近 30 天内有原创推文（僵尸号过滤）
联系渠道规则：
- A 级（100万+）→ 邮件优先，不发冷 DM
- B/C/D 级 → DM + 邮件双线
"""

import os
import re
import sys
import logging
import requests
from datetime import datetime, timezone, timedelta
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).parent.parent))
import twitter_scraper

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
EMAIL_BLOCKLIST = {"example.com", "domain.com", "email.com", "twitter.com", "t.co",
                   "sentry.io", "cloudflare.com", "wixpress.com", "squarespace.com"}

# 已知假邮箱（Apple示例、占位符等）
EMAIL_EXACT_BLOCKLIST = {
    "johnappleseed@gmail.com",
    "example@gmail.com",
    "test@gmail.com",
    "info@example.com",
    "contact@example.com",
}

def extract_email_from_text(text: str) -> str:
    """从文本中提取第一个有效邮箱。"""
    if not text:
        return ""
    for email in EMAIL_RE.findall(text):
        domain = email.split("@")[-1].lower()
        email_lower = email.lower()
        if domain not in EMAIL_BLOCKLIST and "." in domain and email_lower not in EMAIL_EXACT_BLOCKLIST:
            return email_lower
    return ""

def fetch_page(url: str, timeout: int = 8) -> str:
    """抓取页面文本，失败返回空字符串。"""
    if not url:
        return ""
    try:
        resp = requests.get(url, timeout=timeout, headers=HEADERS,
                            allow_redirects=True)
        if resp.status_code == 200:
            return resp.text
    except Exception:
        pass
    return ""

def find_email_linktree(url: str) -> str:
    """Linktree 页面：抓取后找邮箱或 mailto 链接。"""
    html = fetch_page(url)
    if not html:
        return ""
    # mailto: 链接
    mailto = re.findall(r'mailto:([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})', html)
    for m in mailto:
        domain = m.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST:
            return m.lower()
    return extract_email_from_text(html)

def find_email_youtube(username: str) -> str:
    """
    尝试 YouTube @username/about 页面找邮箱。
    YouTube 的 About 页面会在描述中显示邮箱（部分创作者）。
    """
    for handle in [username.lstrip("@"), username.lstrip("@").lower()]:
        html = fetch_page(f"https://www.youtube.com/@{handle}/about", timeout=10)
        if html:
            email = extract_email_from_text(html)
            if email:
                return email
            # YouTube 也会有 mailto 链接
            mailto = re.findall(
                r'mailto:([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})', html
            )
            for m in mailto:
                domain = m.split("@")[-1].lower()
                if domain not in EMAIL_BLOCKLIST:
                    return m.lower()
    return ""

def find_email_substack(username: str) -> str:
    """尝试 substack.com/@username 或 username.substack.com 找邮箱。"""
    urls = [
        f"https://substack.com/@{username.lstrip('@')}",
        f"https://{username.lstrip('@')}.substack.com",
    ]
    for url in urls:
        html = fetch_page(url)
        email = extract_email_from_text(html)
        if email:
            return email
    return ""

def find_email_github(username: str) -> str:
    """GitHub API 查用户公开邮箱（免费，无需 token）。"""
    handle = username.lstrip("@")
    try:
        resp = requests.get(
            f"https://api.github.com/users/{handle}",
            timeout=6,
            headers={**HEADERS, "Accept": "application/vnd.github+json"},
        )
        if resp.status_code == 200:
            data = resp.json()
            email = (data.get("email") or "").strip().lower()
            if email:
                domain = email.split("@")[-1]
                if domain not in EMAIL_BLOCKLIST:
                    return email
    except Exception:
        pass
    return ""

def find_email_mirror(username: str) -> str:
    """Mirror.xyz 个人主页找邮箱（加密原生写作者常用）。"""
    handle = username.lstrip("@")
    for url in [f"https://mirror.xyz/@{handle}", f"https://{handle}.mirror.xyz"]:
        html = fetch_page(url)
        if html:
            email = extract_email_from_text(html)
            if email:
                return email
            mailto = re.findall(r'mailto:([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})', html)
            for m in mailto:
                if m.split("@")[-1].lower() not in EMAIL_BLOCKLIST:
                    return m.lower()
    return ""

def find_email_contact_pages(base_url: str) -> str:
    """尝试官网的 /contact、/about、/contact-us 子页面找邮箱。"""
    if not base_url:
        return ""
    # 清理 base_url，只保留 scheme + netloc
    from urllib.parse import urlparse
    parsed = urlparse(base_url)
    if not parsed.scheme or not parsed.netloc:
        return ""
    root = f"{parsed.scheme}://{parsed.netloc}"
    for path in ["/contact", "/about", "/contact-us", "/links", "/work-with-me"]:
        html = fetch_page(root + path, timeout=6)
        if html:
            email = extract_email_from_text(html)
            if email:
                return email
            mailto = re.findall(r'mailto:([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})', html)
            for m in mailto:
                if m.split("@")[-1].lower() not in EMAIL_BLOCKLIST:
                    return m.lower()
    return ""

def find_email_website(url: str) -> str:
    """抓主页，Linktree 特殊处理，否则通用提取，再尝试子页面。"""
    if not url:
        return ""
    if "linktr.ee" in url or "linktree" in url:
        return find_email_linktree(url)
    html = fetch_page(url)
    if not html:
        return ""
    # 先找 mailto
    mailto = re.findall(
        r'mailto:([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})', html
    )
    for m in mailto:
        domain = m.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST:
            return m.lower()
    email = extract_email_from_text(html)
    if email:
        return email
    # 主页没找到，尝试 /contact 等子页面
    return find_email_contact_pages(url)

def auto_find_email(username: str, bio: str, website_url: str) -> str:
    """
    多渠道串联查找邮箱，找到即返回：
    1. Twitter bio
    2. Twitter 链接网站（含 Linktree + /contact 子页）
    3. GitHub 公开邮箱
    4. Mirror.xyz
    5. YouTube @username/about
    6. Substack
    """
    handle = username.lstrip("@")

    # 1. Twitter bio
    email = extract_email_from_text(bio)
    if email:
        return email

    # 2. Twitter 主页链接（含子页面）
    email = find_email_website(website_url)
    if email:
        return email

    # 3. GitHub
    email = find_email_github(handle)
    if email:
        return email

    # 4. Mirror.xyz
    email = find_email_mirror(handle)
    if email:
        return email

    # 5. YouTube
    email = find_email_youtube(handle)
    if email:
        return email

    # 6. Substack
    email = find_email_substack(handle)
    if email:
        return email

    return ""

load_dotenv(Path(__file__).parent.parent / ".env", override=True)

_log_dir = Path(__file__).parent / "logs"
_log_dir.mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(_log_dir / "kol_collection.log"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# KOL 分级阈值
TIER_A_MIN = 1_000_000   # A 级：100万粉+
TIER_B_MIN = 100_000     # B 级：10~100万粉
TIER_C_MIN = 10_000      # C 级：1~10万粉
TIER_D_MIN = 1_000       # D 级：1000~1万粉（微型 KOL，重点目标）
MIN_FOLLOWERS = 1_000    # 低于此数不收录
ACTIVE_DAYS = 30         # 近 N 天内必须有推文

# 每日收集目标
DAILY_TARGET = 30


def get_outreach_channel(tier: str, has_email: bool) -> str:
    """根据级别和邮箱确定联系渠道"""
    if tier == "A级(100万+)":
        return "邮件优先" if has_email else "等待中间人"
    return "DM+邮件" if has_email else "DM"


def is_recently_active(user_id: str) -> bool:
    """检查账号近 30 天内是否有发推（过滤僵尸号）"""
    try:
        tweets = twitter_scraper.get_user_tweets(user_id, limit=5, days=ACTIVE_DAYS)
        return bool(tweets)
    except Exception:
        return True

def get_tier(followers: int) -> str:
    if followers >= TIER_A_MIN:
        return "A级(100万+)"
    elif followers >= TIER_B_MIN:
        return "B级(10~100万)"
    elif followers >= TIER_C_MIN:
        return "C级(1~10万)"
    elif followers >= TIER_D_MIN:
        return "D级(1000~1万)"
    return "不达标"

# 预测市场 KOL 种子名单
PREDICTION_MARKET_KOLS = [
    ("Shayne Coplan", "@shayne_coplan", "Polymarket创始人"),
    ("Tarek Mansour", "@Tarek_Mansour", "Kalshi创始人"),
    ("Austin Chen", "@Austn_Chen", "预测市场研究"),
    ("Manifold Markets", "@ManifoldMarkets", "预测市场平台"),
    ("Metaculus", "@metaculus", "预测市场平台"),
    ("Scott Alexander", "@slatestarcodex", "理性主义/预测"),
    ("Zvi Mowshowitz", "@ZviMowshowitz", "预测市场分析"),
    ("Robin Hanson", "@robinhanson", "预测市场理论"),
    ("Nate Silver", "@NateSilver538", "数据预测/选举"),
    ("Superforecasters", "@GJOpen", "超级预测者社群"),
]

# Meme 币 KOL 种子名单
MEME_KOLS = [
    ("Murad Mahmudov", "@MustStopMurad", "Meme币/加密周期"),
    ("Hsaka", "@HsakaTrades", "加密交易/Meme币"),
    ("AltcoinSherpa", "@AltcoinSherpa", "山寨币/Meme币教育"),
    ("Frank DeGods", "@frankdegods", "Solana NFT/Meme币"),
    ("WhalePanda", "@WhalePanda", "链上数据/鲸鱼跟踪"),
    ("CryptoNobler", "@CryptoNobler", "Meme币发现"),
    ("CryptoGodJohn", "@CryptoGodJohn", "Meme币交易策略"),
    ("Crypto Rand", "@0xcryptorand", "链上数据/Meme币"),
    ("Koroush AK", "@KoroushAK", "加密交易策略"),
    ("Crypto Dog", "@TheCryptoDog", "加密市场评论"),
]

KNOWN_KOLS = PREDICTION_MARKET_KOLS + MEME_KOLS

def verify_kol(name: str, handle: str, category: str) -> dict:
    """
    验证并获取 KOL 信息。
    过滤条件：账号不存在 / 粉丝 < 1000 / 近 30 天无推文
    """
    username = handle.lstrip("@")
    try:
        user = twitter_scraper.get_user(username)
        if not user:
            logger.warning(f"❌ 未找到用户: {handle}")
            return None

        followers = user.get("followers_count", 0) or 0

        # 验证 1：粉丝数
        if followers < MIN_FOLLOWERS:
            logger.info(f"⏭ 跳过（粉丝不足 {MIN_FOLLOWERS:,}）: {handle} | {followers:,}")
            return None

        # 验证 2：近 30 天活跃度
        try:
            recent = twitter_scraper.get_user_tweets(user["id"], limit=5, days=ACTIVE_DAYS)
            if not recent:
                logger.info(f"⏭ 跳过（{ACTIVE_DAYS}天无原创推文）: {handle}")
                return None
        except Exception:
            pass  # 查不到时不过滤

        tier = get_tier(followers)
        description = user.get("description", "")
        verified = "✓" if user.get("verified", False) else ""

        entities = user.get("entities", {})
        expanded_url = ""
        url_entities = entities.get("url", {}).get("urls", [])
        if url_entities:
            expanded_url = url_entities[0].get("expanded_url", "")

        email = auto_find_email(handle, description, expanded_url)
        channel = get_outreach_channel(tier, bool(email))
        email_status = email if email else "需人工"
        logger.info(
            f"✅ [{tier}] {name:20} | {handle:20} | {followers:,} {verified} | 渠道: {channel} | 邮箱: {email_status}"
        )

        return {
            "name": name,
            "handle": handle,
            "followers": followers,
            "tier": tier,
            "description": description,
            "verified": "是" if user.get("verified", False) else "否",
            "category": category,
            "email": email,
            "outreach_channel": channel,
            "status": "待联系",
        }

    except Exception as e:
        logger.error(f"❌ 查询失败 {handle}: {e}")
        return None


def search_kols_by_keyword(keyword: str, max_results: int = 50, category: str = "搜索发现") -> list:
    """按关键词搜索 KOL，自动分级"""
    logger.info(f"\n🔍 搜索: '{keyword}'")
    try:
        query = f"{keyword} -is:retweet lang:en"
        tweets = twitter_scraper.search_tweets(query, limit=min(max_results, 100))

        if not tweets:
            logger.warning(f"  未找到相关推文")
            return []

        seen_ids = set()
        results = []
        for tweet in tweets:
            user = tweet.get("user", {})
            uid = user.get("id", "")
            if not uid or uid in seen_ids:
                continue
            seen_ids.add(uid)
            followers = user.get("followers_count", 0) or 0
            if followers >= MIN_FOLLOWERS:
                tier = get_tier(followers)
                desc = user.get("description", "")
                entities = user.get("entities", {})
                site_url = ""
                url_entities = entities.get("url", {}).get("urls", [])
                if url_entities:
                    site_url = url_entities[0].get("expanded_url", "")
                username = user.get("username", "")
                email = auto_find_email(username, desc, site_url)
                channel = get_outreach_channel(tier, bool(email))
                results.append({
                    "name": user.get("name", username),
                    "handle": f"@{username}",
                    "followers": followers,
                    "tier": tier,
                    "description": desc[:80],
                    "verified": "是" if user.get("verified", False) else "否",
                    "category": category,
                    "email": email,
                    "outreach_channel": channel,
                    "status": "待验证",
                })

        results.sort(key=lambda x: x["followers"], reverse=True)
        logger.info(f"  找到 {len(results)} 个 KOL（达标 {MIN_FOLLOWERS:,} 粉丝）")
        return results

    except Exception as e:
        logger.error(f"搜索失败: {e}")
        return []


def search_kols_by_engagement(seed_handles: list, max_per_seed: int = 100) -> list:
    """
    互动图谱法（Free tier 可用）：搜索提到/回复种子 KOL 的推文，
    从作者中筛选达标 KOL。
    逻辑：主动和大 KOL 互动的人，大概率在同一个行业圈子。
    注：替代 get_users_following（需要 Basic tier），改用 search_recent_tweets。
    """
    results = []
    for handle in seed_handles:
        username = handle.lstrip("@")
        logger.info(f"\n👥 互动图谱：搜索提到 @{username} 的账号...")
        try:
            query = f"@{username} -is:retweet lang:en"
            tweets = twitter_scraper.search_tweets(query, limit=min(max_per_seed, 100))
            if not tweets:
                logger.info(f"  未找到互动账号")
                continue

            seen_ids = set()
            count = 0
            for tweet in tweets:
                u = tweet.get("user", {})
                uid = u.get("id", "")
                uname = u.get("username", "")
                if not uid or uid in seen_ids or uname.lower() == username.lower():
                    continue
                seen_ids.add(uid)
                followers = u.get("followers_count", 0) or 0
                if followers < MIN_FOLLOWERS:
                    continue
                tier = get_tier(followers)
                desc = u.get("description", "")
                entities = u.get("entities", {})
                site_url = ""
                url_entities = entities.get("url", {}).get("urls", [])
                if url_entities:
                    site_url = url_entities[0].get("expanded_url", "")
                email = auto_find_email(uname, desc, site_url)
                channel = get_outreach_channel(tier, bool(email))
                results.append({
                    "name": u.get("name", uname),
                    "handle": f"@{uname}",
                    "followers": followers,
                    "tier": tier,
                    "description": desc[:80],
                    "verified": "是" if u.get("verified", False) else "否",
                    "category": f"互动图谱-{handle}",
                    "email": email,
                    "outreach_channel": channel,
                    "status": "待验证",
                })
                count += 1

            logger.info(f"  @{username} 互动圈找到 {count} 个达标 KOL")

        except Exception as e:
            logger.error(f"❌ 互动图谱 @{username} 失败: {e}")

    results.sort(key=lambda x: x["followers"], reverse=True)
    logger.info(f"\n  互动图谱法共找到 {len(results)} 个 KOL")
    return results


def load_existing_handles(directory: str) -> set:
    """从已有 xlsx 文件读取已收录的 handle，避免重复"""
    import glob
    handles = set()
    for f in glob.glob(f"{directory}/*.xlsx"):
        try:
            wb = openpyxl.load_workbook(f, read_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str) and cell.startswith("@"):
                            handles.add(cell.lower())
        except Exception:
            pass
    return handles


def generate_excel(kols: list, filename: str = "MoonX_KOL名单_自动收集.xlsx"):
    """生成 Excel 文件"""
    logger.info(f"\n📊 生成 Excel: {filename}")
    wb = Workbook()
    ws = wb.active
    ws.title = "KOL名单"

    # 标题
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "MoonX KOL 名单 — 自动收集（验证：账号存在 + 粉丝≥1000 + 30天活跃）"
    title.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title.font = Font(color="FFFFFF", bold=True, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头
    headers = [
        "#",
        "KOL 名字",
        "Twitter Handle",
        "粉丝数",
        "级别",
        "简介",
        "认证账号",
        "分类",
        "邮箱（自动）",
        "联系渠道",
        "联系状态",
        "备注",
    ]
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    # 级别颜色映射
    tier_colors = {
        "A级(100万+)":  "FFD700",   # 金色
        "B级(10~100万)": "87CEEB",  # 蓝色
        "C级(1~10万)":  "98FB98",   # 绿色
        "D级(1000~1万)": "FFE4B5",  # 淡橙色
    }

    # 数据
    for row_idx, kol in enumerate(kols, 3):
        ws.cell(row=row_idx, column=1, value=row_idx - 2)
        ws.cell(row=row_idx, column=2, value=kol.get("name", ""))
        ws.cell(row=row_idx, column=3, value=kol.get("handle", ""))
        ws.cell(row=row_idx, column=4, value=kol.get("followers", 0))

        tier = kol.get("tier", "")
        tier_cell = ws.cell(row=row_idx, column=5, value=tier)
        if tier in tier_colors:
            tier_cell.fill = PatternFill(
                start_color=tier_colors[tier], end_color=tier_colors[tier], fill_type="solid"
            )

        ws.cell(row=row_idx, column=6, value=kol.get("description", "")[:80])
        ws.cell(row=row_idx, column=7, value=kol.get("verified", "否"))
        ws.cell(row=row_idx, column=8, value=kol.get("category", ""))

        # 邮箱列：有值绿色，无值灰色
        email = kol.get("email", "")
        email_cell = ws.cell(row=row_idx, column=9, value=email if email else "—")
        if email:
            email_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            email_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # 联系渠道列（带颜色标注）
        channel = kol.get("outreach_channel", "DM")
        channel_cell = ws.cell(row=row_idx, column=10, value=channel)
        if channel == "邮件优先":
            channel_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        elif channel == "等待中间人":
            channel_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        elif channel == "DM+邮件":
            channel_cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        else:
            channel_cell.fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")

        status = kol.get("status", "待联系")
        status_cell = ws.cell(row=row_idx, column=11, value=status)
        if status in ("待联系", "待验证"):
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif status == "已联系":
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif status == "已合作":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ws.cell(row=row_idx, column=12, value="")  # 备注留空

    # 列宽
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 38
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 28
    ws.column_dimensions["J"].width = 14  # 联系渠道
    ws.column_dimensions["K"].width = 10  # 联系状态
    ws.column_dimensions["L"].width = 20  # 备注

    wb.save(filename)
    logger.info(f"✅ Excel 已生成: {filename}")
    return filename


def main():
    from datetime import date
    today = date.today().strftime("%Y-%m-%d")

    logger.info("=" * 60)
    logger.info(f"🚀 MoonX KOL 收集 — {today}")
    logger.info(f"📌 今日目标：{DAILY_TARGET} 个新 KOL")
    logger.info("=" * 60)

    kol_dir = str(Path(__file__).parent)
    existing_handles = load_existing_handles(kol_dir)
    logger.info(f"📂 已有 {len(existing_handles)} 个 handle，跳过重复")

    all_kols = []

    def add_if_new(kol):
        if kol and kol["handle"].lower() not in existing_handles:
            existing_handles.add(kol["handle"].lower())
            all_kols.append(kol)

    # 1. 验证种子 KOL
    logger.info("\n📋 验证种子 KOL...")
    for name, handle, category in KNOWN_KOLS:
        kol = verify_kol(name, handle, category)
        add_if_new(kol)

    # 2. 预测市场关键词（与 MoonX 最相关，优先搜）
    prediction_keywords = [
        ("prediction market crypto", "预测市场"),
        ("polymarket kalshi", "预测市场"),
        ("polymarket alternative", "预测市场"),
        ("crypto prediction market", "预测市场"),
        ("election prediction market", "预测市场"),
        ("prediction market alpha", "预测市场"),
    ]

    logger.info("\n🔮 搜索预测市场 KOL...")
    for keyword, category in prediction_keywords:
        found = search_kols_by_keyword(keyword, max_results=50, category=category)
        for kol in found:
            add_if_new(kol)

    # 3. Meme 币 / 链上数据
    meme_keywords = [
        ("meme coin trading solana", "Meme币"),
        ("pump.fun trading", "Meme币"),
        ("GMGN on-chain", "Meme币"),
        ("solana degen alpha", "Meme币"),
    ]

    logger.info("\n🐸 搜索 Meme 币 KOL...")
    for keyword, category in meme_keywords:
        found = search_kols_by_keyword(keyword, max_results=50, category=category)
        for kol in found:
            add_if_new(kol)

    # 4. 加密货币广域（DeFi / 主流币 / 链上数据 / 交易策略）
    crypto_broad_keywords = [
        ("on-chain data analysis crypto", "链上数据"),
        ("DeFi alpha calls", "DeFi"),
        ("BTC ETH analysis", "主流币"),
        ("crypto trading strategy", "加密交易"),
        ("altcoin picks 2026", "山寨币"),
        ("web3 market analysis", "Web3"),
        ("crypto portfolio tracker", "加密交易"),
        ("airdrop hunter crypto", "Airdrop"),
        ("NFT crypto trader", "NFT"),
        ("crypto KOL influencer", "加密KOL"),
    ]

    logger.info("\n🌐 搜索加密货币广域 KOL...")
    for keyword, category in crypto_broad_keywords:
        found = search_kols_by_keyword(keyword, max_results=50, category=category)
        for kol in found:
            add_if_new(kol)
        if len(all_kols) >= DAILY_TARGET * 3:
            break

    # 5. 粉丝图谱法：从已知 KOL 的 following 列表挖掘（量最大）
    # 选取活跃且领域相关的种子账号（预测市场 + 加密行业）
    graph_seeds = [
        "@robinhanson",       # 预测市场理论，关注大量预测市场 KOL
        "@ManifoldMarkets",   # 预测市场平台，关注同赛道竞品和 KOL
        "@metaculus",         # 预测市场平台
        "@coinbureau",        # 加密广域，关注量大且质量高
        "@WhalePanda",        # 链上数据，关注加密交易类 KOL
        "@AltcoinSherpa",     # 山寨币，关注中腰部 KOL 多
    ]
    logger.info("\n🕸️ 互动图谱法搜索（Free tier 可用版）...")
    graph_kols = search_kols_by_engagement(graph_seeds, max_per_seed=100)
    for kol in graph_kols:
        add_if_new(kol)

    # 6. 按粉丝数排序
    all_kols.sort(key=lambda x: x["followers"], reverse=True)

    # 5. 统计分级
    tier_counts = {}
    for kol in all_kols:
        t = kol.get("tier", "未分级")
        tier_counts[t] = tier_counts.get(t, 0) + 1

    logger.info(f"\n📊 收集结果（共 {len(all_kols)} 个新 KOL）：")
    for tier, count in sorted(tier_counts.items()):
        logger.info(f"  {tier}: {count} 个")

    target_status = "✅ 达标" if len(all_kols) >= DAILY_TARGET else f"⚠️ 未达标（目标 {DAILY_TARGET} 个）"
    logger.info(f"\n{target_status}")

    # 6. 生成 Excel
    filename = Path(kol_dir) / f"MoonX_KOL名单_{today}.xlsx"
    generate_excel(all_kols, str(filename))

    logger.info("\n🏆 Top 10 KOL：")
    for i, kol in enumerate(all_kols[:10], 1):
        print(
            f"  {i:2}. [{kol['tier']:12}] {kol['name']:20} {kol['handle']:22} | {kol['followers']:,} 粉丝"
        )


if __name__ == "__main__":
    main()
