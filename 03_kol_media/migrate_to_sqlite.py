#!/usr/bin/env python3
"""
KOL CRM — Excel → SQLite 数据迁移脚本
一次性运行，将所有历史 Excel 数据导入 kol_crm.db
"""
import os
import re
import sys
from pathlib import Path
from datetime import datetime

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

from kol_db import DB_PATH, DB_DIR, init_db, upsert_kol, upsert_media, mark_sent, record_reply, get_db

_ON_FLY = bool(os.getenv("FLY_APP_NAME"))
DATA_DIR = Path("/data") if _ON_FLY else Path(__file__).parent


def _extract_date(filename: str) -> str:
    m = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    return m.group(1) if m else ""


def migrate_kol_excel(verbose: bool = True):
    """迁移所有 YouTube KOL 名单 Excel"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    all_files = sorted(
        list(DATA_DIR.glob("MoonX_YouTube_KOL名单_*.xlsx")) +
        list(DATA_DIR.glob("MoonX_Kalshi_KOL名单_*.xlsx")),
        key=lambda f: f.name,
    )
    if not all_files:
        print(f"[migrate] 未找到 KOL Excel（在 {DATA_DIR}）")
        return 0

    total_inserted = total_updated = 0

    for f in all_files:
        collect_date = _extract_date(f.name)
        source = "kalshi" if "Kalshi" in f.name else "youtube"
        if verbose:
            print(f"\n[migrate] 读取 {f.name} ...")

        try:
            wb = load_workbook(f, read_only=True)
            ws = wb.active
        except Exception as e:
            print(f"  [ERROR] 打开失败: {e}")
            continue

        file_inserted = file_updated = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not (row and row[1]):  # name 为空跳过
                continue

            channel_id  = str(row[0] or "").strip()
            name        = str(row[1] or "").strip()
            channel_url = str(row[2] or "").strip() if len(row) > 2 else ""
            twitter     = str(row[3] or "").strip() if len(row) > 3 else ""
            try:
                subscribers = int(row[4] or 0) if len(row) > 4 else 0
            except (ValueError, TypeError):
                continue  # 跳过表头行
            tier        = str(row[5] or "").strip() if len(row) > 5 else ""
            country     = str(row[6] or "").strip() if len(row) > 6 else ""
            email       = str(row[7] or "").strip() if len(row) > 7 else ""
            description = str(row[8] or "").strip() if len(row) > 8 else ""
            status_raw  = str(row[9] or "").strip() if len(row) > 9 else ""

            if not name:
                continue

            # 映射状态
            if status_raw.startswith("已回复"):
                status = "已回复"
            elif status_raw == "已签约":
                status = "已签约"
            elif status_raw == "已发送":
                status = "已发送"
            else:
                status = "待发送"

            data = {
                "source":       source,
                "channel_id":   channel_id or None,
                "channel_url":  channel_url,
                "name":         name,
                "platform":     "YouTube",
                "subscribers":  subscribers,
                "tier":         tier,
                "email":        email,
                "description":  description,
                "twitter":      twitter,
                "country":      country,
                "collect_date": collect_date,
                "status":       status,
            }

            try:
                kol_id = upsert_kol(data)
                # 若已发送但 contacts 表无记录，补一条虚拟记录
                if status in ("已发送", "已回复", "已签约"):
                    with get_db() as conn:
                        cnt = conn.execute(
                            "SELECT COUNT(*) FROM contacts WHERE kol_id=?", (kol_id,)
                        ).fetchone()[0]
                        if cnt == 0:
                            sent_at = f"{collect_date} 22:00:00" if collect_date else "2026-01-01 22:00:00"
                            conn.execute(
                                "INSERT INTO contacts (kol_id, sent_at, template, status) VALUES (?,?,?,?)",
                                (kol_id, sent_at, "migrated", "sent")
                            )
                        # 若已回复但 replies 表无记录，补一条虚拟记录
                        if status in ("已回复", "已签约"):
                            rcnt = conn.execute(
                                "SELECT COUNT(*) FROM replies WHERE kol_id=?", (kol_id,)
                            ).fetchone()[0]
                            if rcnt == 0:
                                # 从状态里提取日期（如"已回复 2026-03-26"）
                                dm = re.search(r'(\d{4}-\d{2}-\d{2})', status_raw)
                                reply_date = dm.group(1) if dm else collect_date or "2026-01-01"
                                conn.execute(
                                    """INSERT INTO replies
                                       (kol_id, email_from, detected_at, intent, auto_response_sent)
                                       VALUES (?,?,?,?,?)""",
                                    (kol_id, email, f"{reply_date} 09:00:00", "migrated", 1)
                                )
                file_inserted += 1
            except Exception as e:
                print(f"  [WARN] {name}: {e}")

        wb.close()
        total_inserted += file_inserted
        if verbose:
            print(f"  导入 {file_inserted} 条")

    print(f"\n[migrate] KOL 迁移完成: {total_inserted} 条")
    return total_inserted


def migrate_media_excel(verbose: bool = True):
    """迁移所有媒体库 Excel"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0

    media_files = sorted(DATA_DIR.glob("MoonX_媒体库_*.xlsx"))
    if not media_files:
        print(f"[migrate] 未找到媒体库 Excel")
        return 0

    total = 0
    for f in media_files:
        collect_date = _extract_date(f.name)
        if verbose:
            print(f"\n[migrate] 读取 {f.name} ...")
        try:
            wb = load_workbook(f, read_only=True)
            ws = wb.active
        except Exception as e:
            print(f"  [ERROR] {e}")
            continue

        headers = [cell for cell in ws[2]]
        def _colidx(keyword):
            for i, h in enumerate(headers):
                if h and keyword in str(h.value or ""):
                    return i
            return None

        ci_name     = _colidx("媒体名称") or 0
        ci_type     = _colidx("类型")     or 1
        ci_website  = _colidx("网站")     or 2
        ci_email    = _colidx("邮箱")     or 5
        ci_priority = _colidx("优先级")   or 9
        ci_remark   = _colidx("备注")     or 10

        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not (row and row[ci_name]):
                continue
            name     = str(row[ci_name] or "").strip()
            typ      = str(row[ci_type] or "crypto").strip()
            website  = str(row[ci_website] or "").strip() if len(row) > ci_website else ""
            email    = str(row[ci_email] or "").strip() if len(row) > ci_email else ""
            priority = str(row[ci_priority] or "B").strip() if len(row) > ci_priority else "B"
            remark   = str(row[ci_remark] or "").strip() if len(row) > ci_remark else ""

            if not name:
                continue

            # 映射状态
            if "已回复" in remark:
                status = remark  # 保留原始状态带日期
            elif "已询价" in remark or "已发送" in remark:
                status = remark
            else:
                status = "待联系"

            try:
                media_id = upsert_media({
                    "name":         name,
                    "type":         typ,
                    "website":      website,
                    "email":        email,
                    "priority":     priority,
                    "notes":        remark,
                    "collect_date": collect_date,
                })
                # 更新状态
                if status != "待联系":
                    with get_db() as conn:
                        conn.execute(
                            "UPDATE media SET status=? WHERE id=?", (status, media_id)
                        )
                count += 1
            except Exception as e:
                print(f"  [WARN] {name}: {e}")

        wb.close()
        total += count
        if verbose:
            print(f"  导入 {count} 条")

    print(f"\n[migrate] 媒体 迁移完成: {total} 条")
    return total


def print_summary():
    with get_db() as conn:
        kol_total   = conn.execute("SELECT COUNT(*) FROM kols").fetchone()[0]
        kol_email   = conn.execute("SELECT COUNT(*) FROM kols WHERE email!=''").fetchone()[0]
        kol_sent    = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
        kol_replied = conn.execute("SELECT COUNT(*) FROM replies").fetchone()[0]
        media_total = conn.execute("SELECT COUNT(*) FROM media").fetchone()[0]
        status_rows = conn.execute(
            "SELECT status, COUNT(*) FROM kols GROUP BY status ORDER BY 2 DESC"
        ).fetchall()

    print("\n" + "=" * 50)
    print(f"DB: {DB_PATH}")
    print(f"KOL 总数:   {kol_total}  (有邮箱: {kol_email})")
    print(f"外联记录:   {kol_sent}")
    print(f"回复记录:   {kol_replied}")
    print(f"媒体总数:   {media_total}")
    print("\nKOL 状态分布:")
    for r in status_rows:
        print(f"  {r[0] or '(空)'}: {r[1]}")
    print("=" * 50)


if __name__ == "__main__":
    print(f"[migrate] 初始化数据库: {DB_PATH}")
    init_db()

    print("\n[1/2] 迁移 KOL Excel ...")
    migrate_kol_excel()

    print("\n[2/2] 迁移媒体 Excel ...")
    migrate_media_excel()

    print_summary()
