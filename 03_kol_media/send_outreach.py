#!/usr/bin/env python3
"""
GMGN 预测市场 — 自动外联邮件发送脚本
发件人：Kelly，市场总监
功能：读取 Excel 名单 → 匹配模板 → 批量发送 → 更新状态
"""

import smtplib
import time
import random
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from openpyxl import load_workbook
from dotenv import load_dotenv

# ── 加载配置 ──
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

SMTP_HOST    = os.getenv("BYDFI_SMTP_HOST")
SMTP_PORT    = int(os.getenv("BYDFI_SMTP_PORT", "465"))
SMTP_USER    = os.getenv("BYDFI_EMAIL", "kelly@bydfi.com")
SMTP_PASS    = os.getenv("BYDFI_EMAIL_PASSWORD")
SENDER_NAME  = os.getenv("SENDER_NAME", "Kelly")
SENDER_TITLE = os.getenv("SENDER_TITLE", "Head of Marketing")
PLATFORM     = os.getenv("PLATFORM_NAME", "BYDFi")
PLATFORM_URL = os.getenv("PLATFORM_URL", "https://www.bydfi.com/zh/moonx/markets/trending")
SENDER_EMAIL = SMTP_USER
SENDER_TG    = os.getenv("SENDER_TG", "@BDkelly")
REBATE_URL   = "https://www.bydfi.com/zh/moonx/account/my-rebate?type=my-rebate"

EXCEL_PATH = Path(__file__).parent / "GMGN_KOL媒体外联名单.xlsx"

# ══════════════════════════════════════════
# 邮件模板
# ══════════════════════════════════════════

def tpl_crypto_media(editor_name, media_name):
    subject = "Story tip: BYDFi's MoonX is doing GMGN-style smart money tracking for retail"
    body = f"""Hi {editor_name},

Thought this might be worth a look for {media_name}.

BYDFi just launched MoonX — an on-chain trading tool for Solana and BNB Chain that tracks smart wallet activity across 500K+ addresses and surfaces early meme coin signals before they hit mainstream crypto Twitter.

The angle that might interest your readers: it's essentially bringing the smart money tracking that previously required technical skill (querying on-chain data, running your own scripts) down to a one-click interface for regular traders.

A few data points:
• Filters 97% of low-quality tokens, surfaces only high-conviction moves
• Integrated with pump.fun, Raydium, PancakeSwap — catches new tokens within seconds of liquidity forming
• Copy-trading any wallet with one click

Happy to set up a demo or connect you with the founding team for a briefing. Can also share user data if useful for the story.

Platform: {PLATFORM_URL}

{SENDER_NAME}
{SENDER_TITLE} | {PLATFORM}
{SENDER_EMAIL} · {SENDER_TG}"""
    return subject, body


def tpl_finance_media(editor_name, media_name):
    subject = "How retail traders are following 'smart money' on-chain — story angle"
    body = f"""Hi {editor_name},

Reaching out with a story angle that might resonate with {media_name}'s audience.

There's a growing category of retail crypto traders who are using on-chain wallet tracking to follow sophisticated investors — essentially watching where "smart money" moves before it shows up in prices. BYDFi's MoonX is one of the tools making this accessible to non-technical users.

Why this might interest your readers now:
• Meme coin trading volume on Solana exceeded $10B in monthly volume in early 2026
• A new generation of tools (GMGN, MoonX) is making institutional-style flow analysis accessible to retail
• The behavior mirrors what quantitative funds do — but for tokens, not stocks

Happy to share more data or arrange an interview with the BYDFi team.

{SENDER_NAME}
{SENDER_TITLE} | {PLATFORM}
{SENDER_EMAIL} · {SENDER_TG}"""
    return subject, body


def tpl_followup(name, original_subject):
    subject = f"Re: {original_subject}"
    body = f"""Hi {name},

Just wanted to bump this up in case it got buried.

We're finalizing our launch partner list this week and wanted to make sure you had a chance to consider it before we close the early access window.

Even a 15-min call or quick email exchange works great.

Thanks either way!

{SENDER_NAME}
{SENDER_TITLE} | {PLATFORM}
{SENDER_EMAIL}"""
    return subject, body


# ══════════════════════════════════════════
# 发送函数
# ══════════════════════════════════════════

def send_email(to_email, subject, body, dry_run=True):
    """
    dry_run=True  → 只打印，不真实发送（测试用）
    dry_run=False → 真实发送
    """
    if dry_run:
        print(f"\n{'='*60}")
        print(f"[DRY RUN] 收件人: {to_email}")
        print(f"主题: {subject}")
        print(f"内容预览:\n{body[:200]}...")
        print(f"{'='*60}")
        return True

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"{SENDER_NAME} <{SMTP_USER}>"
        msg["To"]      = to_email

        msg.attach(MIMEText(body, "plain", "utf-8"))

        if SMTP_PORT == 465:
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_USER, to_email, msg.as_string())
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_USER, to_email, msg.as_string())

        print(f"  ✓ 已发送 → {to_email}")
        return True

    except Exception as e:
        print(f"  ✗ 发送失败 → {to_email}：{e}")
        return False


# ══════════════════════════════════════════
# 读取 Excel + 批量发送
# ══════════════════════════════════════════

def run(sheet_name, col_map, template_fn, dry_run=True):
    """
    sheet_name  : Excel Sheet 名称
    col_map     : {"name": 列号, "email": 列号, "status": 列号, ...}
    template_fn : 生成 (subject, body) 的函数
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]

    sent_count = 0
    skip_count = 0

    for row in ws.iter_rows(min_row=3):  # 跳过标题行
        name   = row[col_map["name"]  - 1].value or ""
        email  = row[col_map["email"] - 1].value or ""
        status = row[col_map["status"]- 1].value or ""

        if not email or "@" not in email:
            skip_count += 1
            continue

        if status in ("已发送", "已回复", "已签约"):
            skip_count += 1
            continue

        subject, body = template_fn(name, name)
        success = send_email(email, subject, body, dry_run=dry_run)

        if success and not dry_run:
            row[col_map["status"] - 1].value = "已发送"
            sent_count += 1
            # 同步写 SQLite CRM
            try:
                import sys as _sys
                _sys.path.insert(0, str(Path(__file__).parent))
                from kol_db import get_kols_by_email, mark_sent
                sent_map = get_kols_by_email()
                kol_row = sent_map.get(email.lower())
                if not kol_row:
                    # 也检查 status!='已发送' 的 KOL
                    import sqlite3 as _sq
                    _db = Path(__file__).parent / "kol_crm.db"
                    with _sq.connect(str(_db)) as _c:
                        _c.row_factory = _sq.Row
                        _r = _c.execute(
                            "SELECT id, name, status FROM kols WHERE LOWER(email)=?",
                            (email.lower(),)
                        ).fetchone()
                        kol_row = dict(_r) if _r else None
                if kol_row:
                    mark_sent(kol_row["id"], subject=subject, template=sheet_name)
                    print(f"  ✓ SQLite 已同步 (kol_id={kol_row['id']})")
            except Exception as _e:
                print(f"  ⚠ SQLite 同步失败（不影响发送）: {_e}")
            # 随机等待 30-90 秒，避免被 Gmail 限速
            wait = random.randint(30, 90)
            print(f"  等待 {wait} 秒后发下一封...")
            time.sleep(wait)
        elif dry_run:
            sent_count += 1

    if not dry_run:
        wb.save(EXCEL_PATH)

    print(f"\n[{sheet_name}] 完成：发送 {sent_count} 封，跳过 {skip_count} 条")


# ══════════════════════════════════════════
# 主程序
# ══════════════════════════════════════════

if __name__ == "__main__":
    import sys

    # 默认 dry_run=True，加 --send 参数才真实发送
    dry_run = "--send" not in sys.argv

    if dry_run:
        print("=" * 60)
        print("⚠️  当前为测试模式（DRY RUN），不会真实发送邮件")
        print("   确认无误后运行：python3 send_outreach.py --send")
        print("=" * 60)
    else:
        print("=" * 60)
        print("🚀 真实发送模式已开启，开始发送邮件...")
        print("=" * 60)

    # 发送加密媒体邮件（Sheet: 加密媒体，列：名称=1, 邮箱=5, 状态=7）
    print("\n📧 加密媒体...")
    run(
        sheet_name="加密媒体",
        col_map={"name": 1, "email": 5, "status": 7},
        template_fn=tpl_crypto_media,
        dry_run=dry_run
    )

    # 发送财经媒体邮件（Sheet: 财经股票媒体）
    print("\n📧 财经股票媒体...")
    run(
        sheet_name="财经股票媒体",
        col_map={"name": 1, "email": 5, "status": 7},
        template_fn=tpl_finance_media,
        dry_run=dry_run
    )

    print("\n✅ 全部完成！")
    if not dry_run:
        print(f"📊 Excel 状态已更新：{EXCEL_PATH}")
