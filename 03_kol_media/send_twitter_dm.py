#!/usr/bin/env python3
"""
BYDFi MoonX — 自动发送 Twitter DM 给 Meme 币 KOL
"""

import os, time, random, logging
from pathlib import Path
from dotenv import load_dotenv
import tweepy
from openpyxl import load_workbook

load_dotenv(Path(__file__).parent.parent / ".env", override=True)
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

# ── Twitter 认证 ──
client = tweepy.Client(
    bearer_token=os.getenv("TWITTER_BEARER_TOKEN"),
    consumer_key=os.getenv("TWITTER_API_KEY"),
    consumer_secret=os.getenv("TWITTER_API_SECRET"),
    access_token=os.getenv("TWITTER_ACCESS_TOKEN"),
    access_token_secret=os.getenv("TWITTER_ACCESS_TOKEN_SECRET"),
)

MOONX_URL    = "https://www.bydfi.com/en/moonx/markets/trending"
REBATE_URL   = "https://www.bydfi.com/zh/moonx/account/my-rebate?type=my-rebate"
SENDER_TG    = os.getenv("SENDER_TG", "@BDkelly")
SENDER_EMAIL = os.getenv("BYDFI_EMAIL", "kelly@bydfi.com")

# 读取所有 KOL 名单文件（按日期排序，从旧到新）
import glob
_all_files = sorted(glob.glob(str(Path(__file__).parent / "MoonX_KOL名单_*.xlsx")))

# ── 日志 ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(Path(__file__).parent / "logs" / "twitter_dm.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ── DM 内容（5个模版随机轮换，降低封号风险）──
DM_TEMPLATES = [
    # 模版 1 — 直接介绍产品
    (
        "Hey {name} 👋\n\n"
        "Love your content — you clearly know how to spot moves early.\n\n"
        "I'm Kelly from BYDFi. We just launched MoonX ({url}) — "
        "tracks smart wallet activity across Solana and BNB Chain in real time. "
        "Filters 97% of noise, surfaces only tokens where smart money is actually moving.\n\n"
        "Onboarding early KOL partners now:\n"
        "✅ Free full access\n"
        "✅ Personal referral link with rev-share ({rebate})\n"
        "✅ Custom alpha data for your content\n\n"
        "Interested? Let's chat 🙏\n"
        "TG: {tg}"
    ),
    # 模版 2 — 强调收益
    (
        "Hey {name} — big fan 🔥\n\n"
        "Building the KOL partner program for BYDFi MoonX and your audience is exactly who we're looking for.\n\n"
        "MoonX tracks smart wallet moves on Solana + BNB Chain — pump.fun, Raydium, PancakeSwap — "
        "one-click copy trading on any wallet.\n\n"
        "What we offer partners:\n"
        "💰 Rev share on every trader you refer ({rebate})\n"
        "📊 Early access to smart money dashboard\n"
        "🎯 Alpha data feeds for your content\n\n"
        "Worth a quick chat? Hit me up.\n"
        "— Kelly | TG: {tg}"
    ),
    # 模版 3 — 工具角度
    (
        "Hi {name} 👋\n\n"
        "Your on-chain content is some of the best out there — thought this might be relevant.\n\n"
        "I'm Kelly from BYDFi. We shipped MoonX ({url}) — "
        "monitors 500K+ wallets across Solana and BNB Chain, shows you where high-conviction "
        "money is moving before the price does. Integrated with pump.fun and major DEXes.\n\n"
        "Quietly onboarding a few KOL partners before full launch:\n"
        "→ Free access\n"
        "→ Personal referral dashboard with rev-share\n"
        "→ Raw data feeds for your content\n\n"
        "Open to exploring? TG: {tg} 🙏"
    ),
    # 模版 4 — 简短
    (
        "Hey {name}! 👋\n\n"
        "I run partnerships at BYDFi — we just launched MoonX, a smart money tracker for "
        "Solana and BNB Chain meme coins ({url}).\n\n"
        "Your audience would find this useful. Early KOL partners get:\n"
        "• Free full access\n"
        "• Rev share on referrals\n"
        "• Alpha data for content\n\n"
        "Keen to chat? TG: {tg} or reply here 🙌"
    ),
    # 模版 5 — 社区角度
    (
        "What's up {name} 🤙\n\n"
        "Been following your calls — solid stuff.\n\n"
        "I'm Kelly from BYDFi. We built MoonX ({url}) for traders like your community — "
        "tracks smart wallet moves on pump.fun and major Solana/BNB DEXes in real time. "
        "One-click copy trading on any wallet.\n\n"
        "Looking for a few KOL partners to grow together:\n"
        "✅ Free access\n"
        "✅ Rev share referral link ({rebate})\n"
        "✅ Alpha data to level up your content\n\n"
        "Let me know if curious!\n"
        "Kelly | BYDFi | TG: {tg}"
    ),
]

def get_dm_text(name):
    template = random.choice(DM_TEMPLATES)
    return template.format(name=name, url=MOONX_URL, tg=SENDER_TG, email=SENDER_EMAIL, rebate=REBATE_URL)

def get_username(handle):
    """去掉 @ 符号"""
    return handle.lstrip("@")

def send_dm(username, name, dry_run=True):
    try:
        # 1. 根据用户名查找 user_id
        user = client.get_user(username=username)
        if not user.data:
            logger.warning(f"  ✗ 找不到用户: @{username}")
            return False
        user_id = user.data.id
        dm_text = get_dm_text(name)

        if dry_run:
            logger.info(f"  [DRY RUN] @{username} (ID:{user_id})")
            logger.info(f"  内容预览: {dm_text[:100]}...")
            return True

        # 2. 发送 DM
        client.create_direct_message(participant_id=user_id, text=dm_text)
        logger.info(f"  ✓ DM 已发送 → @{username} ({name})")
        return True

    except tweepy.errors.Forbidden as e:
        logger.error(f"  ✗ 无法发送给 @{username}（对方可能关闭了 DM）: {e}")
        return False
    except tweepy.errors.TooManyRequests:
        logger.warning("  ⚠ Twitter 限速，等待 15 分钟...")
        time.sleep(900)
        return False
    except Exception as e:
        logger.error(f"  ✗ 失败 @{username}: {e}")
        return False

DAILY_LIMIT = 10  # 每天最多发送条数，避免被封

def run(dry_run=True):
    if not _all_files:
        logger.error("❌ 未找到任何 KOL 名单文件")
        return

    # 跨文件去重：记录已处理过的 handle
    seen_handles = set()
    sent, skip = 0, 0

    for file_path in reversed(_all_files):  # 从最新文件开始
        if sent >= DAILY_LIMIT:
            break

        fp = Path(file_path)
        logger.info(f"📂 读取名单: {fp.name}")
        wb = load_workbook(fp)
        sheet_name = "KOL名单" if "KOL名单" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        is_new_format = "KOL名单" in wb.sheetnames
        file_sent = 0

        for row in ws.iter_rows(min_row=3):
            if sent >= DAILY_LIMIT:
                break

            if is_new_format:
                name       = str(row[1].value or "").strip()
                handle     = str(row[2].value or "").strip()
                tier       = str(row[4].value or "").strip()
                channel    = str(row[9].value or "").strip()
                status     = str(row[10].value or "").strip()
                status_col = 10
            else:
                name       = str(row[0].value or "").strip()
                handle     = str(row[1].value or "").strip()
                contact    = str(row[4].value or "").strip()
                status     = str(row[7].value or "").strip()
                tier       = ""
                channel    = "DM"
                status_col = 7

            if not handle or not handle.startswith("@"):
                skip += 1
                continue
            if status in ("已发送DM", "已回复", "已签约"):
                skip += 1
                continue
            if channel in ("邮件优先", "等待中间人"):
                logger.info(f"  ⏭ 跳过 A级 KOL（{channel}）: {handle}")
                skip += 1
                continue
            if not is_new_format and "Twitter DM" not in contact and "@" not in contact:
                skip += 1
                continue
            if handle.lower() in seen_handles:
                skip += 1
                continue

            seen_handles.add(handle.lower())
            username = get_username(handle)
            logger.info(f"▶ 发送给: {name} ({handle}) [{tier}]")

            success = send_dm(username, name, dry_run=dry_run)
            if success and not dry_run:
                row[status_col].value = "已发送DM"
                file_sent += 1
                sent += 1
                if sent >= DAILY_LIMIT:
                    logger.info(f"  ⚠ 已达每日上限 {DAILY_LIMIT} 条，停止发送")
                    break
                wait = random.randint(90, 200)
                logger.info(f"  ⏱ 等待 {wait} 秒...")
                time.sleep(wait)
            elif dry_run:
                sent += 1

        if not dry_run and file_sent > 0:
            wb.save(fp)
            logger.info(f"  Excel 已更新: {fp.name}")

    logger.info(f"\n{'='*50}")
    logger.info(f"完成！发送 {sent} 条DM，跳过 {skip} 条")

if __name__ == "__main__":
    import sys
    dry_run = "--send" not in sys.argv

    if dry_run:
        logger.info("="*50)
        logger.info("⚠️  测试模式 — 不会真实发送 DM")
        logger.info("   真实发送：python3 send_twitter_dm.py --send")
        logger.info("="*50)
    else:
        logger.info("="*50)
        logger.info("🚀 开始发送 Twitter DM...")
        logger.info("="*50)

    run(dry_run=dry_run)
