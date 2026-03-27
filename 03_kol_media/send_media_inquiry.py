#!/usr/bin/env python3
"""
媒体询价邮件批量发送
- 读取 MoonX_媒体库_*.xlsx 中优先级 A、有邮箱、未发送的媒体
- 加密/科技类媒体 → MEDIA_INQUIRY_CRYPTO_EN
- 财经类媒体      → MEDIA_INQUIRY_FINANCE_EN
- 发送后在 Excel 联系状态列写入 "已询价 YYYY-MM-DD"
"""

import os
import re
import time
import random
import logging
import smtplib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook

from outreach_templates import MEDIA_INQUIRY_CRYPTO_EN, MEDIA_INQUIRY_FINANCE_EN

load_dotenv(Path(__file__).parent.parent / ".env", override=True)
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

GMAIL        = os.getenv("GMAIL_ADDRESS")
APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD")
SENDER_EMAIL = GMAIL or "ppmworker@gmail.com"

TODAY    = datetime.now().strftime("%Y-%m-%d")
_DIR     = Path(__file__).parent
_ON_FLY  = bool(os.getenv("FLY_APP_NAME"))
DATA_DIR = Path("/data") if _ON_FLY else _DIR
LOG_DIR  = DATA_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_DIR / "media_inquiry.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


def find_latest_media_excel():
    files = sorted(DATA_DIR.glob("MoonX_媒体库_*.xlsx"), reverse=True)
    return files[0] if files else None


def build_email(media: dict) -> tuple[str, str]:
    """根据媒体类型选择模板，返回 (subject, body)。"""
    typ = media.get("type", "crypto")
    template = MEDIA_INQUIRY_FINANCE_EN if typ == "finance" else MEDIA_INQUIRY_CRYPTO_EN
    full = template.format(
        media_name=media["name"],
        contact_name="Team",
        sender_email=SENDER_EMAIL,
    )
    lines = full.strip().splitlines()
    subject = lines[0].replace("Subject: ", "").strip()
    body = "\n".join(lines[2:]).strip()
    return subject, body


def send_email(to_email: str, subject: str, body: str) -> bool:
    try:
        msg = MIMEMultipart("alternative")
        msg["From"]    = SENDER_EMAIL
        msg["To"]      = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL, APP_PASSWORD)
            server.sendmail(SENDER_EMAIL, to_email, msg.as_string())
        return True
    except Exception as e:
        logger.error(f"  发送失败: {e}")
        return False


def run():
    excel_path = find_latest_media_excel()
    if not excel_path:
        logger.error("未找到媒体库 Excel，请先运行 collect_media.py")
        return

    logger.info("=" * 60)
    logger.info(f"📰 媒体询价邮件发送 — {TODAY}")
    logger.info(f"   读取: {excel_path.name}")
    logger.info("=" * 60)

    wb = load_workbook(excel_path)
    ws = wb.active

    # 找各列位置（第2行表头）
    headers = [cell.value for cell in ws[2]]
    def col(name):
        for i, h in enumerate(headers):
            if h and name in str(h):
                return i + 1
        return None

    col_name     = col("媒体名称") or 1
    col_type     = col("类型")     or 2
    col_email    = col("联系邮箱") or 6
    col_priority = col("优先级")   or 10
    col_status   = col("备注")     or 11

    sent, skipped = 0, 0
    to_send = []

    for row in ws.iter_rows(min_row=3):
        name     = row[col_name - 1].value
        typ      = row[col_type - 1].value or "crypto"
        email    = row[col_email - 1].value or ""
        priority = row[col_priority - 1].value or ""
        status   = row[col_status - 1].value or ""

        if not name or priority != "A":
            continue
        if not EMAIL_RE.match(str(email).strip()):
            logger.info(f"  跳过 {name}：邮箱无效（{email}）")
            skipped += 1
            continue
        if "已询价" in str(status):
            logger.info(f"  跳过 {name}：已发送过")
            skipped += 1
            continue

        to_send.append({
            "row":   row,
            "name":  name,
            "type":  typ,
            "email": email.strip(),
            "status_cell": row[col_status - 1],
        })

    logger.info(f"待发送: {len(to_send)} 家  跳过: {skipped} 家\n")

    for i, m in enumerate(to_send, 1):
        subject, body = build_email(m)
        logger.info(f"[{i}/{len(to_send)}] {m['name']} <{m['email']}>")
        ok = send_email(m["email"], subject, body)
        if ok:
            m["status_cell"].value = f"已询价 {TODAY}"
            sent += 1
            logger.info(f"  ✓ 已发送")
        delay = random.randint(60, 120)
        if i < len(to_send):
            logger.info(f"  ⏱ 等待 {delay} 秒...")
            time.sleep(delay)

    wb.save(excel_path)
    logger.info(f"\n{'='*60}")
    logger.info(f"完成！发送 {sent} 封，跳过 {skipped} 家")
    logger.info(f"Excel 状态已更新: {excel_path.name}")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    run()
