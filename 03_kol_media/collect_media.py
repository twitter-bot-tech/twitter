#!/usr/bin/env python3
"""
BYDFi MoonX — 媒体库收集脚本
覆盖范围：欧美加密媒体 + 欧美财经媒体 + 东南亚加密媒体
收集字段：媒体名、类型、地区、官网、流量级别、联系邮箱、广告入口、媒体包、报价、优先级
每周一自动运行，输出 MoonX_媒体库_YYYY-MM-DD.xlsx
"""

import os
import re
import time
import logging
import requests
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

load_dotenv(Path(__file__).parent.parent / ".env", override=True)
load_dotenv(Path(__file__).parent.parent / ".env.outreach", override=True)

TODAY = datetime.now().strftime("%Y-%m-%d")
_SCRIPT_DIR = Path(__file__).parent
_ON_FLY = bool(os.getenv("FLY_APP_NAME"))
OUTPUT_DIR = Path("/data") if _ON_FLY else _SCRIPT_DIR
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR = OUTPUT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "media_collect.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}
REQUEST_TIMEOUT = 12

# ── 种子媒体列表 ─────────────────────────────────────────────────────────────
# traffic_tier: S(>500万/月) A(100-500万) B(20-100万) C(<20万)
# type: crypto / finance / tech
MEDIA_SEEDS = [
    # ── 欧美 加密媒体 ──────────────────────────────────────────────────────
    {"name": "CoinDesk",         "type": "crypto",   "region": "US",  "traffic_tier": "S",
     "url": "https://www.coindesk.com",       "ad_path": "/advertise"},
    {"name": "CoinTelegraph",    "type": "crypto",   "region": "US",  "traffic_tier": "S",
     "url": "https://cointelegraph.com",      "ad_path": "/advertise"},
    {"name": "Decrypt",          "type": "crypto",   "region": "US",  "traffic_tier": "A",
     "url": "https://decrypt.co",             "ad_path": "/advertise"},
    {"name": "The Block",        "type": "crypto",   "region": "US",  "traffic_tier": "A",
     "url": "https://www.theblock.co",        "ad_path": "/about"},
    {"name": "Blockworks",       "type": "crypto",   "region": "US",  "traffic_tier": "A",
     "url": "https://blockworks.co",          "ad_path": "/advertise"},
    {"name": "BeInCrypto",       "type": "crypto",   "region": "EU",  "traffic_tier": "A",
     "url": "https://beincrypto.com",         "ad_path": "/advertise"},
    {"name": "CryptoSlate",      "type": "crypto",   "region": "US",  "traffic_tier": "B",
     "url": "https://cryptoslate.com",        "ad_path": "/advertise"},
    {"name": "Bitcoin Magazine",  "type": "crypto",  "region": "US",  "traffic_tier": "B",
     "url": "https://bitcoinmagazine.com",    "ad_path": "/advertise"},
    {"name": "DL News",          "type": "crypto",   "region": "EU",  "traffic_tier": "B",
     "url": "https://www.dlnews.com",         "ad_path": "/about"},
    {"name": "Protos",           "type": "crypto",   "region": "EU",  "traffic_tier": "C",
     "url": "https://protos.com",             "ad_path": "/about"},
    {"name": "Bankless",         "type": "crypto",   "region": "US",  "traffic_tier": "B",
     "url": "https://www.bankless.com",       "ad_path": "/advertise"},
    {"name": "Unchained",        "type": "crypto",   "region": "US",  "traffic_tier": "B",
     "url": "https://unchainedcrypto.com",    "ad_path": "/about"},
    {"name": "Crypto Briefing",  "type": "crypto",   "region": "US",  "traffic_tier": "B",
     "url": "https://cryptobriefing.com",     "ad_path": "/advertise"},
    {"name": "Messari",          "type": "crypto",   "region": "US",  "traffic_tier": "A",
     "url": "https://messari.io",             "ad_path": "/advertise"},
    {"name": "The Defiant",      "type": "crypto",   "region": "US",  "traffic_tier": "B",
     "url": "https://thedefiant.io",          "ad_path": "/about"},
    {"name": "Milk Road",        "type": "crypto",   "region": "US",  "traffic_tier": "B",
     "url": "https://milkroad.com",           "ad_path": "/advertise"},
    {"name": "Pomp Letter",      "type": "crypto",   "region": "US",  "traffic_tier": "B",
     "url": "https://pomp.substack.com",      "ad_path": ""},
    # ── 欧美 财经媒体 ──────────────────────────────────────────────────────
    {"name": "Bloomberg Crypto", "type": "finance",  "region": "US",  "traffic_tier": "S",
     "url": "https://www.bloomberg.com",      "ad_path": "/company/advertising"},
    {"name": "Forbes Crypto",    "type": "finance",  "region": "US",  "traffic_tier": "S",
     "url": "https://www.forbes.com",         "ad_path": "/fdc/advertise.html"},
    {"name": "Reuters Finance",  "type": "finance",  "region": "US",  "traffic_tier": "S",
     "url": "https://www.reuters.com",        "ad_path": "/advertise"},
    {"name": "Axios Markets",    "type": "finance",  "region": "US",  "traffic_tier": "A",
     "url": "https://www.axios.com",          "ad_path": "/advertise"},
    {"name": "Business Insider", "type": "finance",  "region": "US",  "traffic_tier": "S",
     "url": "https://www.businessinsider.com","ad_path": "/advertise"},
    {"name": "Investopedia",     "type": "finance",  "region": "US",  "traffic_tier": "S",
     "url": "https://www.investopedia.com",   "ad_path": "/advertise"},
    {"name": "MarketWatch",      "type": "finance",  "region": "US",  "traffic_tier": "S",
     "url": "https://www.marketwatch.com",    "ad_path": "/advertising"},
    {"name": "Seeking Alpha",    "type": "finance",  "region": "US",  "traffic_tier": "A",
     "url": "https://seekingalpha.com",       "ad_path": "/advertise"},
    {"name": "The Motley Fool",  "type": "finance",  "region": "US",  "traffic_tier": "A",
     "url": "https://www.fool.com",           "ad_path": "/advertising"},
    {"name": "Financial Times",  "type": "finance",  "region": "EU",  "traffic_tier": "S",
     "url": "https://www.ft.com",             "ad_path": "/advertising"},
    # ── 东南亚 加密媒体 ────────────────────────────────────────────────────
    {"name": "Coinvestasi",      "type": "crypto",   "region": "ID",  "traffic_tier": "B",
     "url": "https://coinvestasi.com",        "ad_path": "/advertise"},
    {"name": "Blocktempo",       "type": "crypto",   "region": "TW",  "traffic_tier": "B",
     "url": "https://www.blocktempo.com",     "ad_path": "/about"},
    {"name": "CoinCU",           "type": "crypto",   "region": "VN",  "traffic_tier": "C",
     "url": "https://coincu.com",             "ad_path": "/advertise"},
    {"name": "Siamblockchain",   "type": "crypto",   "region": "TH",  "traffic_tier": "C",
     "url": "https://www.siamblockchain.com", "ad_path": "/contact"},
    {"name": "Cryptonews ID",    "type": "crypto",   "region": "ID",  "traffic_tier": "C",
     "url": "https://id.cryptonews.com",      "ad_path": "/advertise"},
    {"name": "BitcoinWorld",     "type": "crypto",   "region": "IN",  "traffic_tier": "C",
     "url": "https://bitcoinworld.co.in",     "ad_path": "/contact"},
    {"name": "Forkast News",     "type": "crypto",   "region": "SG",  "traffic_tier": "B",
     "url": "https://forkast.news",           "ad_path": "/advertise"},
    {"name": "KrASIA",           "type": "tech",     "region": "SG",  "traffic_tier": "B",
     "url": "https://kr.asia",                "ad_path": "/about"},
    {"name": "Tech in Asia",     "type": "tech",     "region": "SG",  "traffic_tier": "A",
     "url": "https://www.techinasia.com",     "ad_path": "/advertise"},
    {"name": "e27",              "type": "tech",     "region": "SG",  "traffic_tier": "B",
     "url": "https://e27.co",                 "ad_path": "/advertise"},
    {"name": "BeInCrypto ID",    "type": "crypto",   "region": "ID",  "traffic_tier": "B",
     "url": "https://id.beincrypto.com",      "ad_path": "/advertise"},
    {"name": "BeInCrypto TH",    "type": "crypto",   "region": "TH",  "traffic_tier": "B",
     "url": "https://th.beincrypto.com",      "ad_path": "/advertise"},
    {"name": "BeInCrypto VN",    "type": "crypto",   "region": "VN",  "traffic_tier": "B",
     "url": "https://beincrypto.com/vn",      "ad_path": "/advertise"},
    # ── 付费广告平台 ───────────────────────────────────────────────────────
    {"name": "Reddit Ads",       "type": "paid_ad",  "region": "US",  "traffic_tier": "S",
     "url": "https://ads.reddit.com",         "ad_path": ""},
]

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
EMAIL_BLOCKLIST = {
    "example.com", "domain.com", "sentry.io", "cloudflare.com",
    "wixpress.com", "squarespace.com", "google.com", "w3.org",
}
MEDIA_KIT_PATTERNS = re.compile(
    r'href=["\']([^"\']*(?:media[_-]?kit|advertis|partner|sponsor|rate[_-]?card)[^"\']*)["\']',
    re.IGNORECASE,
)


def _get(url: str, timeout: int = REQUEST_TIMEOUT) -> str:
    """GET 请求，失败返回空字符串。"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        if resp.status_code == 200:
            return resp.text
    except Exception:
        pass
    return ""


def extract_emails(html: str) -> list[str]:
    found = []
    for email in EMAIL_RE.findall(html):
        domain = email.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST and "." in domain:
            found.append(email.lower())
    return list(dict.fromkeys(found))  # 去重保序


def find_contact_email(base_url: str) -> str:
    """依次尝试 /advertise、/contact、/about 页面提取联系邮箱。"""
    for path in ["/advertise", "/contact", "/about", "/contact-us", "/work-with-us"]:
        html = _get(base_url.rstrip("/") + path)
        if not html:
            continue
        emails = extract_emails(html)
        # 优先广告/媒体相关邮箱
        for kw in ["advertis", "partner", "media", "sponsor", "press", "biz"]:
            for e in emails:
                if kw in e:
                    return e
        if emails:
            return emails[0]
        time.sleep(0.5)
    return ""


def find_media_kit_url(base_url: str, ad_path: str) -> str:
    """在广告页面寻找 media kit / rate card 链接。"""
    target = base_url.rstrip("/") + ad_path if ad_path else base_url
    html = _get(target)
    if not html:
        return ""
    matches = MEDIA_KIT_PATTERNS.findall(html)
    if matches:
        link = matches[0]
        if link.startswith("http"):
            return link
        return base_url.rstrip("/") + "/" + link.lstrip("/")
    # 检查是否有 PDF 链接（常见的 media kit 格式）
    pdf_match = re.search(r'href=["\']([^"\']*\.pdf)["\']', html, re.IGNORECASE)
    if pdf_match:
        link = pdf_match.group(1)
        return link if link.startswith("http") else base_url.rstrip("/") + "/" + link.lstrip("/")
    return ""


def score_priority(media: dict) -> str:
    """
    A：S/A 流量 + 有邮箱
    B：B 流量 + 有邮箱，或 S/A 流量无邮箱
    C：其余
    """
    tier = media.get("traffic_tier", "C")
    has_email = bool(media.get("email"))
    has_kit = bool(media.get("media_kit_url"))
    if tier in ("S", "A") and has_email:
        return "A"
    if (tier in ("S", "A")) or (tier == "B" and has_email):
        return "B" if not has_kit else "A"
    return "C"


def load_existing_media() -> set[str]:
    """从已有媒体库 Excel 中读取已收录的媒体名，去重用。"""
    seen = set()
    for fp in sorted(OUTPUT_DIR.glob("MoonX_媒体库_*.xlsx")):
        try:
            wb = load_workbook(fp, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and row[0]:
                    seen.add(str(row[0]).strip())
            wb.close()
        except Exception:
            pass
    return seen


def build_excel(media_list: list[dict], output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "媒体库"

    orange       = "FF6B00"
    dark_orange  = "E55A00"
    white        = "FFFFFF"
    light_orange = "FFF3E0"

    title_fill  = PatternFill("solid", fgColor=orange)
    header_fill = PatternFill("solid", fgColor=dark_orange)
    alt_fill    = PatternFill("solid", fgColor=light_orange)
    title_font  = Font(bold=True, color=white, size=13)
    header_font = Font(bold=True, color=white, size=10)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    priority_colors = {"A": "FFD700", "B": "C6EFCE", "C": "DDEBF7"}

    # 标题行
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = f"BYDFi MoonX — 媒体库（{TODAY}）共 {len(media_list)} 家"
    c.fill = title_fill
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # 表头
    headers = [
        "媒体名称", "类型", "地区", "流量级别", "官网",
        "联系邮箱", "广告入口", "媒体包链接", "报价", "优先级", "备注", "收录日期",
    ]
    col_widths = [22, 10, 8, 10, 38, 35, 38, 42, 20, 10, 20, 14]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[chr(64 + col_idx)].width = w
    ws.row_dimensions[2].height = 22

    for row_idx, m in enumerate(media_list, 3):
        priority = m.get("priority", "C")
        row_fill = PatternFill("solid", fgColor=priority_colors.get(priority, "FFFFFF"))
        if row_idx % 2 == 0 and priority == "C":
            row_fill = alt_fill
        row_data = [
            m["name"],
            m["type"],
            m["region"],
            m["traffic_tier"],
            m["url"],
            m.get("email", ""),
            m["url"].rstrip("/") + m.get("ad_path", "") if m.get("ad_path") else "",
            m.get("media_kit_url", ""),
            m.get("pricing", "待询价"),
            priority,
            "",
            TODAY,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (5, 6, 7, 8)))
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A3"
    wb.save(output_path)
    logger.info(f"✅ Excel 已保存: {output_path.name}（{len(media_list)} 家媒体）")


def run():
    logger.info("=" * 60)
    logger.info(f"📰 媒体库收集开始 — {TODAY}")
    logger.info(f"   种子媒体: {len(MEDIA_SEEDS)} 家")
    logger.info("=" * 60)

    existing = load_existing_media()
    logger.info(f"📋 已有记录: {len(existing)} 家（自动去重）")

    results: list[dict] = []
    new_count = 0

    for i, seed in enumerate(MEDIA_SEEDS, 1):
        name = seed["name"]
        logger.info(f"[{i}/{len(MEDIA_SEEDS)}] {name} ({seed['region']})")

        # 已存在则跳过采集，但仍写入 Excel 保持完整库
        if name in existing:
            seed["email"] = ""
            seed["media_kit_url"] = ""
            seed["pricing"] = "待询价"
            seed["priority"] = score_priority(seed)
            results.append(seed)
            logger.info(f"   ↩ 已有记录，跳过采集")
            continue

        # 采集联系邮箱
        email = find_contact_email(seed["url"])
        seed["email"] = email
        if email:
            logger.info(f"   ✉ 邮箱: {email}")

        # 查找媒体包
        kit_url = find_media_kit_url(seed["url"], seed.get("ad_path", ""))
        seed["media_kit_url"] = kit_url
        if kit_url:
            logger.info(f"   📎 媒体包: {kit_url}")

        seed["pricing"] = "待询价"
        seed["priority"] = score_priority(seed)
        results.append(seed)
        new_count += 1
        time.sleep(1.5)

    # 按优先级排序
    priority_order = {"A": 0, "B": 1, "C": 2}
    results.sort(key=lambda x: (priority_order.get(x["priority"], 3), x["name"]))

    # 统计
    counts = {"A": 0, "B": 0, "C": 0}
    emails_found = sum(1 for m in results if m.get("email"))
    kits_found = sum(1 for m in results if m.get("media_kit_url"))
    for m in results:
        counts[m["priority"]] = counts.get(m["priority"], 0) + 1

    logger.info(f"\n📊 收集完成:")
    logger.info(f"   新增: {new_count} 家  总计: {len(results)} 家")
    logger.info(f"   优先级 A: {counts['A']}  B: {counts['B']}  C: {counts['C']}")
    logger.info(f"   找到邮箱: {emails_found}  找到媒体包: {kits_found}")

    output_path = OUTPUT_DIR / f"MoonX_媒体库_{TODAY}.xlsx"
    build_excel(results, output_path)

    logger.info(f"\n{'='*60}")
    logger.info(f"完成！输出: {output_path.name}")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    import sys as _sys; _sys.path.insert(0, str(Path(__file__).parent.parent))
    from scripts.pid_lock import acquire_lock; acquire_lock()
    run()
