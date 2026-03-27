#!/usr/bin/env python3
"""
MoonX KOL 名单 → Google Sheets 同步脚本

前置条件：
1. Google Cloud Console 创建 Service Account，开启 Google Sheets API + Google Drive API
2. 下载 JSON key，粘贴到 .env 的 GOOGLE_SERVICE_ACCOUNT_JSON（整个 JSON 一行）
3. 创建 Google Sheet，把 Service Account 邮箱加为编辑者
4. 把 Sheet ID 填入 .env 的 GOOGLE_SHEET_ID_KOL

用法：
    python3 sync_kol_to_sheets.py                  # 同步今天的 Excel
    python3 sync_kol_to_sheets.py --file xxx.xlsx  # 同步指定文件
"""

import os, json, argparse
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials

load_dotenv(Path(__file__).parent.parent / ".env", override=True)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

TODAY = datetime.now().strftime("%Y-%m-%d")
SCRIPT_DIR = Path(__file__).parent
TOKEN_FILE = SCRIPT_DIR / "token.json"


def get_service():
    if not TOKEN_FILE.exists():
        raise FileNotFoundError("token.json 不存在，请先运行 python3 auth_google.py 完成授权")
    creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    return build("sheets", "v4", credentials=creds)


def find_today_excel(target_file=None):
    if target_file:
        p = Path(target_file)
        if not p.exists():
            raise FileNotFoundError(f"文件不存在: {target_file}")
        return p
    for pattern in [
        f"MoonX_Web3_KOL_{TODAY}.xlsx",
        f"MoonX_Browser_KOL名单_{TODAY}.xlsx",
        f"MoonX_KOL名单_{TODAY}.xlsx",
        f"MoonX_YouTube_KOL名单_{TODAY}.xlsx",
    ]:
        p = SCRIPT_DIR / pattern
        if p.exists():
            return p
    raise FileNotFoundError(f"未找到今日 KOL 文件（{TODAY}），请指定 --file")


def read_excel(path: Path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    wb.close()
    return rows


def get_or_create_sheet(service, spreadsheet_id: str, sheet_name: str) -> int:
    """获取 sheet tab，不存在则创建，返回 sheetId"""
    meta = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    for s in meta["sheets"]:
        if s["properties"]["title"] == sheet_name:
            return s["properties"]["sheetId"]
    # 创建新 tab
    resp = service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": sheet_name}}}]},
    ).execute()
    return resp["replies"][0]["addSheet"]["properties"]["sheetId"]


def format_sheet(service, spreadsheet_id: str, sheet_id: int, num_rows: int, num_cols: int):
    """设置标题行橙色、冻结首行、自动列宽"""
    requests = [
        # 标题行背景橙色
        {
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 1.0, "green": 0.42, "blue": 0.0},
                        "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}, "fontSize": 11},
                        "horizontalAlignment": "CENTER",
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
            }
        },
        # 冻结首行
        {
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }
        },
        # 自动列宽
        {
            "autoResizeDimensions": {
                "dimensions": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": num_cols,
                }
            }
        },
        # 交替行色（数据行）
        {
            "addConditionalFormatRule": {
                "rule": {
                    "ranges": [{"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": num_rows + 1}],
                    "booleanRule": {
                        "condition": {"type": "CUSTOM_FORMULA", "values": [{"userEnteredValue": "=ISEVEN(ROW())"}]},
                        "format": {"backgroundColor": {"red": 1.0, "green": 0.95, "blue": 0.88}},
                    },
                },
                "index": 0,
            }
        },
    ]
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id, body={"requests": requests}
    ).execute()


def sync(target_file=None):
    spreadsheet_id = os.getenv("GOOGLE_SHEET_ID_KOL", "").strip()
    if not spreadsheet_id:
        raise ValueError("GOOGLE_SHEET_ID_KOL 未配置，请在 .env 填入 Google Sheet ID")

    print(f"[1/4] 连接 Google Sheets API...")
    service = get_service()

    print(f"[2/4] 读取 Excel...")
    excel_path = find_today_excel(target_file)
    rows = read_excel(excel_path)
    print(f"      {excel_path.name}，共 {len(rows)-1} 条数据")

    sheet_name = f"KOL_{TODAY}"
    print(f"[3/4] 写入 Sheet tab: {sheet_name}")
    sheet_id = get_or_create_sheet(service, spreadsheet_id, sheet_name)

    # 清空并写入
    range_name = f"{sheet_name}!A1"
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id, range=f"{sheet_name}!A:Z"
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=range_name,
        valueInputOption="RAW",
        body={"values": rows},
    ).execute()

    print(f"[4/4] 格式化...")
    format_sheet(service, spreadsheet_id, sheet_id, len(rows) - 1, len(rows[0]) if rows else 8)

    sheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={sheet_id}"
    print(f"\n✅ 同步完成！")
    print(f"   写入行数: {len(rows)-1} 条 KOL")
    print(f"   Sheet URL: {sheet_url}")
    return sheet_url


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="同步 KOL 名单到 Google Sheets")
    parser.add_argument("--file", help="指定 Excel 文件路径，默认读取今日文件")
    args = parser.parse_args()
    sync(args.file)
