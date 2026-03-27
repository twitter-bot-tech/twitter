#!/usr/bin/env python3
"""
MoonX Web3 KOL Scraper — 多渠道收集

渠道：Twitter/X, YouTube, Telegram(tgstat), CoinTelegraph, CoinDesk, Lunarcrush

用法：
    python3 web3_kol_scraper.py            # headless 模式
    python3 web3_kol_scraper.py --headed   # 有头模式（可见浏览器）

首次运行时会弹出浏览器让你登录 Twitter，登录后状态自动保存到 twitter_state.json
之后运行直接复用 session，无需再次登录。
"""

import re, time, json, logging, argparse, requests, urllib3
from datetime import datetime
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

try:
    from bs4 import BeautifulSoup
    BS4_OK = True
except ImportError:
    BS4_OK = False

TODAY      = datetime.now().strftime("%Y-%m-%d")
SCRIPT_DIR = Path(__file__).parent
LOG_DIR    = SCRIPT_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
STATE_FILE = SCRIPT_DIR / "twitter_state.json"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "web3_kol.log"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# ── Twitter 搜索关键词 ──────────────────────────────────
TWITTER_QUERIES = [
    "crypto KOL influencer",
    "bitcoin ethereum trader",
    "solana meme alpha",
    "defi web3 influencer",
    "crypto analyst calls",
    "altcoin gems 2026",
    "on-chain alpha hunter",
    "pump fun memecoin",
    "bnb chain defi",
    "crypto education english",
]

# ── YouTube 搜索关键词 ──────────────────────────────────
YOUTUBE_QUERIES = [
    "crypto trading analysis 2026",
    "bitcoin market update",
    "web3 defi tutorial",
    "altcoin gems crypto",
    "crypto KOL education",
]

MAX_PER_QUERY  = 30
MIN_FOLLOWERS  = 5000

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
EMAIL_BLOCKLIST = {
    "example.com", "domain.com", "twitter.com", "t.co", "x.com",
    "sentry.io", "cloudflare.com", "wixpress.com", "youtu.be",
    "youtube.com",
    # gmail.com 保留（很多 KOL 用 Gmail 作业务联系邮箱）
}

REGION_MAP = {
    "北美":    ["usa", "united states", "canada", "america", "new york", "california", "toronto"],
    "欧洲":    ["uk", "london", "europe", "germany", "france", "spain", "netherlands", "italy"],
    "亚太":    ["japan", "korea", "singapore", "australia", "india", "china", "hong kong", "taiwan"],
    "拉美":    ["brazil", "mexico", "argentina", "colombia", "latam", "latin"],
    "中东/非洲": ["dubai", "uae", "nigeria", "south africa", "turkey", "arab", "africa"],
}

LANG_MAP = {
    "中文":       ["中", "华", "币", "链", "加密", "区块", "比特", "以太"],
    "Spanish":    ["cripto", "mercado", "inversión", "análisis"],
    "German":     ["krypto", "bitcoin", "analyse", "markt"],
    "French":     ["crypto", "analyse", "marché", "investissement"],
    "Portuguese": ["cripto", "mercado", "análise", "investimento"],
    "Turkish":    ["kripto", "bitcoin", "borsa", "analiz"],
}

SOURCE_COLORS = {
    "Twitter":      "E8F4FD",
    "YouTube":      "FDECEA",
    "Telegram":     "E8F0FE",
    "CoinTelegraph": "E6F4EA",
    "CoinDesk":     "FFF8E1",
    "Lunarcrush":   "F3E5F5",
}


# ─────────────────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────────────────

TG_RE = re.compile(r'(?:t\.me|telegram\.me)/([a-zA-Z0-9_]{4,})')


def extract_email(text: str) -> str:
    if not text:
        return ""
    for email in EMAIL_RE.findall(text):
        domain = email.split("@")[-1].lower()
        if domain not in EMAIL_BLOCKLIST:
            return email.lower()
    return ""


def extract_telegram(text: str) -> str:
    """从 bio/页面文本提取 Telegram 链接"""
    if not text:
        return ""
    m = TG_RE.search(text)
    if m:
        handle = m.group(1)
        # 过滤掉非个人账号（频道邀请链接等）
        if not handle.startswith("+") and len(handle) >= 4:
            return f"https://t.me/{handle}"
    return ""


def _scrape_email_from_url(ctx, url: str) -> str:
    """访问 URL（含 Linktree/beacons 等），多层尝试提取邮箱"""
    if not url:
        return ""
    skip_domains = {"instagram.com", "t.me", "telegram.me", "discord.gg",
                    "discord.com", "youtube.com", "youtu.be", "tiktok.com"}
    try:
        from urllib.parse import urlparse
        if urlparse(url).netloc.lstrip("www.") in skip_domains:
            return ""
    except Exception:
        pass

    wp = None
    try:
        wp = ctx.new_page()
        wp.goto(url, wait_until="domcontentloaded", timeout=12000)
        wp.wait_for_timeout(1500)
        final_url = wp.url

        # Linktree / beacons.ai — 找 mailto: 链接
        if "linktr.ee" in final_url or "beacons.ai" in final_url or "linktree" in final_url:
            hrefs = wp.evaluate(
                "() => Array.from(document.querySelectorAll('a[href^=\"mailto:\"]')).map(a=>a.href)"
            )
            for href in hrefs:
                email = href.replace("mailto:", "").split("?")[0].strip()
                if email and "@" in email:
                    domain = email.split("@")[-1].lower()
                    if domain not in EMAIL_BLOCKLIST:
                        wp.close()
                        return email.lower()

        # 尝试主页正文
        body = wp.inner_text("body") if wp.query_selector("body") else ""
        email = extract_email(body)
        if email:
            wp.close()
            return email

        # 尝试 /contact 页
        base = final_url.rstrip("/").rsplit("/", 1)[0] if "/" in final_url[8:] else final_url
        for suffix in ["/contact", "/about", "/contact-us"]:
            try:
                wp.goto(base + suffix, wait_until="domcontentloaded", timeout=6000)
                wp.wait_for_timeout(800)
                body2 = wp.inner_text("body") if wp.query_selector("body") else ""
                email = extract_email(body2)
                if email:
                    wp.close()
                    return email
            except Exception:
                break
    except Exception:
        pass
    finally:
        try:
            if wp:
                wp.close()
        except Exception:
            pass
    return ""


def parse_follower_count(text: str) -> int:
    if not text:
        return 0
    text = text.replace(",", "").strip()
    try:
        if text.upper().endswith("K"):
            return int(float(text[:-1]) * 1_000)
        elif text.upper().endswith("M"):
            return int(float(text[:-1]) * 1_000_000)
        elif text.upper().endswith("B"):
            return int(float(text[:-1]) * 1_000_000_000)
        return int(re.sub(r"[^\d]", "", text) or "0")
    except Exception:
        return 0


def classify_tier(followers: int) -> str:
    if followers >= 1_000_000:
        return "A级(100万+)"
    elif followers >= 500_000:
        return "A级(50万+)"
    elif followers >= 100_000:
        return "B级(10万+)"
    elif followers >= 10_000:
        return "C级(1万+)"
    elif followers > 0:
        return "D级(1万以下)"
    return "媒体KOL"


def infer_region(bio: str, location: str) -> str:
    text = (bio + " " + location).lower()
    for region, keywords in REGION_MAP.items():
        if any(kw in text for kw in keywords):
            return region
    return "其他"


def infer_language(bio: str) -> str:
    for lang, keywords in LANG_MAP.items():
        if any(kw in bio for kw in keywords):
            return lang
    return "English"


def make_kol(source, handle, url, followers, bio, location="", website="", email="", name="", notes="", telegram="") -> dict:
    """统一 KOL 记录格式"""
    bio_clean = bio[:200] if bio else ""
    email = email or extract_email(bio_clean)
    telegram = telegram or extract_telegram(bio_clean)
    return {
        "source":        source,
        "region":        infer_region(bio_clean, location),
        "country":       location or "未知",
        "language":      infer_language(bio_clean),
        "handle":        handle,
        "name":          name or handle,
        "url":           url,
        "followers":     followers,
        "tier":          classify_tier(followers),
        "bio":           bio_clean,
        "website":       website,
        "email":         email,
        "telegram":      telegram,
        "exchanges":     "",
        "priority":      "",
        "notes":         notes or bio_clean[:80],
        "contacted":     "否",
        "date":          TODAY,
    }


def load_existing_handles() -> set:
    seen = set()
    for fp in SCRIPT_DIR.glob("MoonX_Web3_KOL_*.xlsx"):
        try:
            wb = load_workbook(fp, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and row[3]:
                    seen.add(str(row[3]).lstrip("@").lower())
            wb.close()
        except Exception:
            pass
    return seen


# ─────────────────────────────────────────────────────────
# Twitter/X 爬取
# ─────────────────────────────────────────────────────────

def refresh_state_from_chrome() -> bool:
    """从本机 Chrome 浏览器直接读取 x.com cookies，构造 twitter_state.json"""
    try:
        from pycookiecheat import chrome_cookies
        cookies = chrome_cookies("https://x.com")
        if not cookies.get("auth_token"):
            return False
        playwright_cookies = []
        for name, value in cookies.items():
            for domain in (".x.com", ".twitter.com"):
                playwright_cookies.append({
                    "name": name, "value": value,
                    "domain": domain, "path": "/",
                    "expires": -1,
                    "httpOnly": name in ("auth_token", "kdt"),
                    "secure": True, "sameSite": "None",
                })
        STATE_FILE.write_text(json.dumps({"cookies": playwright_cookies, "origins": []}, indent=2))
        logger.info(f"已从 Chrome 提取 Twitter session（{len(cookies)} 个 cookies）")
        return True
    except Exception as e:
        logger.warning(f"Chrome cookie 提取失败: {e}")
        return False


def login_and_save_state(browser):
    ctx = browser.new_context()
    page = ctx.new_page()
    page.goto("https://x.com/login", wait_until="domcontentloaded")
    logger.info("请在弹出的浏览器中登录 Twitter/X，登录完成后按 Enter 继续...")
    input("登录完成后按 Enter：")
    ctx.storage_state(path=str(STATE_FILE))
    logger.info(f"登录状态已保存到 {STATE_FILE}")
    page.close()
    ctx.close()


def twitter_scrape_profile(page, handle: str):
    url = f"https://x.com/{handle.lstrip('@')}"
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=20000)
        page.wait_for_timeout(2000)

        name = ""
        try:
            el = page.query_selector('[data-testid="UserName"] span span')
            if el:
                name = el.inner_text().strip()
        except Exception:
            pass

        followers = 0
        try:
            el = page.query_selector('a[href$="/verified_followers"] span span')
            if not el:
                el = page.query_selector('a[href$="/followers"] span span')
            if el:
                followers = parse_follower_count(el.inner_text().strip())
        except Exception:
            pass

        if followers < MIN_FOLLOWERS:
            return None

        bio, location, website = "", "", ""
        try:
            el = page.query_selector('[data-testid="UserDescription"]')
            if el:
                bio = el.inner_text().strip()
        except Exception:
            pass
        try:
            el = page.query_selector('[data-testid="UserLocation"]')
            if el:
                location = el.inner_text().strip()
        except Exception:
            pass
        try:
            el = page.query_selector('[data-testid="UserUrl"] a')
            if el:
                website = el.get_attribute("href") or ""
        except Exception:
            pass

        email = extract_email(bio)
        if not email and website and "twitter.com" not in website and "x.com" not in website:
            email = _scrape_email_from_url(page.context, website)

        return make_kol(
            source="Twitter",
            handle="@" + handle.lstrip("@"),
            url=f"https://x.com/{handle.lstrip('@')}",
            followers=followers,
            bio=bio,
            location=location,
            website=website,
            email=email,
            name=name or handle,
        )
    except Exception as e:
        logger.warning(f"抓取 @{handle} 失败: {e}")
        return None


def twitter_search_handles(page, query: str, max_results: int = 30) -> list:
    handles = []
    try:
        encoded = query.replace(" ", "%20")
        page.goto(
            f"https://x.com/search?q={encoded}&src=typed_query&f=user",
            wait_until="domcontentloaded", timeout=20000,
        )
        page.wait_for_timeout(3000)
        for _ in range(4):
            page.keyboard.press("End")
            page.wait_for_timeout(1500)

        cells = page.query_selector_all('[data-testid="UserCell"]')
        for cell in cells[:max_results]:
            try:
                link = cell.query_selector('a[href^="/"]')
                if link:
                    href = link.get_attribute("href") or ""
                    h = href.strip("/").split("/")[0]
                    if h and not h.startswith("i/") and h not in ("home", "explore", "notifications"):
                        handles.append(h)
            except Exception:
                pass

        seen, unique = set(), []
        for h in handles:
            if h.lower() not in seen:
                seen.add(h.lower())
                unique.append(h)
        return unique
    except Exception as e:
        logger.warning(f"Twitter搜索 '{query}' 失败: {e}")
        return []


def scrape_twitter(pw, existing: set) -> list:
    logger.info("=" * 40)
    logger.info("渠道 1/6: Twitter/X")

    # 每次运行前自动从本机 Chrome 刷新 session（无需手动登录）
    refresh_state_from_chrome()

    browser = pw.chromium.launch(headless=True, args=["--no-sandbox"])
    if STATE_FILE.exists():
        ctx = browser.new_context(storage_state=str(STATE_FILE))
        logger.info("已加载 Twitter 登录状态")
    else:
        logger.warning("未找到 twitter_state.json，且 Chrome cookie 提取失败，跳过 Twitter 渠道")
        browser.close()
        return []

    page = ctx.new_page()
    page.goto("https://x.com/home", wait_until="domcontentloaded", timeout=20000)
    page.wait_for_timeout(2000)
    if "login" in page.url or page.query_selector('[data-testid="loginButton"]'):
        logger.warning("Twitter 登录失效，尝试重新提取 Chrome cookie...")
        if refresh_state_from_chrome():
            ctx = browser.new_context(storage_state=str(STATE_FILE))
            page = ctx.new_page()
        else:
            browser.close()
            return []
    logger.info("Twitter 登录状态正常")

    discovered = []
    for query in TWITTER_QUERIES:
        logger.info(f"  搜索: {query}")
        handles = twitter_search_handles(page, query, MAX_PER_QUERY)
        new = [h for h in handles
               if h.lower() not in existing
               and h.lower() not in [x.lower() for x in discovered]]
        logger.info(f"  发现 {len(handles)} 个，新增 {len(new)} 个")
        discovered.extend(new)
        time.sleep(1)

    seen, unique = set(), []
    for h in discovered:
        if h.lower() not in seen:
            seen.add(h.lower())
            unique.append(h)

    kols = []
    for i, handle in enumerate(unique, 1):
        logger.info(f"  [{i}/{len(unique)}] @{handle}")
        kol = twitter_scrape_profile(page, handle)
        if kol:
            kols.append(kol)
            logger.info(f"    {kol['name']} | {kol['followers']:,}粉 | {kol['tier']}")
        time.sleep(1.5)

    page.close()
    ctx.close()
    logger.info(f"Twitter: 收集 {len(kols)} 个 KOL")
    return kols


# ─────────────────────────────────────────────────────────
# YouTube 爬取
# ─────────────────────────────────────────────────────────

def scrape_youtube(browser) -> list:
    logger.info("=" * 40)
    logger.info("渠道 2/6: YouTube")
    if not BS4_OK:
        logger.warning("bs4 未安装，跳过 YouTube")
        return []

    kols = []
    seen = set()
    ctx = browser.new_context()
    page = ctx.new_page()

    for query in YOUTUBE_QUERIES:
        try:
            encoded = requests.utils.quote(query)
            # sp=EgIQAg%3D%3D = 频道过滤
            page.goto(
                f"https://www.youtube.com/results?search_query={encoded}&sp=EgIQAg%253D%253D",
                wait_until="domcontentloaded", timeout=20000,
            )
            page.wait_for_timeout(3000)
            # 滚动加载更多
            for _ in range(3):
                page.keyboard.press("End")
                page.wait_for_timeout(1000)

            channels = page.query_selector_all("ytd-channel-renderer")
            logger.info(f"  YouTube '{query}': {len(channels)} 个频道")

            for ch in channels[:15]:
                try:
                    name_el = ch.query_selector("#channel-title, #text.ytd-channel-name, .ytd-channel-name")
                    subs_el = ch.query_selector("#subscribers")
                    link_el = ch.query_selector("a#main-link, a.channel-link")
                    desc_el = ch.query_selector("#description-text, #description")

                    name = name_el.inner_text().strip() if name_el else ""
                    subs_text = subs_el.inner_text().strip() if subs_el else ""
                    href = link_el.get_attribute("href") if link_el else ""
                    bio = desc_el.inner_text().strip() if desc_el else ""

                    if not href or href in seen or not name:
                        continue
                    seen.add(href)

                    channel_url = "https://www.youtube.com" + href if href.startswith("/") else href
                    subs_clean = re.sub(r"[^\d\.KMB]", "", subs_text.upper())
                    followers = parse_follower_count(subs_clean)
                    if followers < MIN_FOLLOWERS:
                        continue

                    # 尝试从 About 页提取邮箱
                    email = extract_email(bio)
                    if not email:
                        try:
                            about_page = ctx.new_page()
                            about_page.goto(channel_url + "/about", wait_until="domcontentloaded", timeout=10000)
                            about_page.wait_for_timeout(1500)
                            # YouTube About 页邮箱通常在 "View email address" 按钮后
                            about_text = about_page.inner_text("body") if about_page.query_selector("body") else ""
                            email = extract_email(about_text)
                            about_page.close()
                        except Exception:
                            pass

                    kols.append(make_kol(
                        source="YouTube",
                        handle=name,
                        url=channel_url,
                        followers=followers,
                        bio=bio,
                        website=channel_url,
                        email=email,
                        notes=f"YouTube频道 | {subs_text}",
                    ))
                except Exception as e:
                    logger.debug(f"YouTube channel parse: {e}")
        except Exception as e:
            logger.warning(f"YouTube搜索 '{query}' 失败: {e}")
        time.sleep(1)

    page.close()
    ctx.close()
    logger.info(f"YouTube: 收集 {len(kols)} 个 KOL")
    return kols


# ─────────────────────────────────────────────────────────
# Telegram — tgstat.com
# ─────────────────────────────────────────────────────────

TGSTAT_URLS = [
    "https://tgstat.com/en/crypto",
    "https://tgstat.com/en/crypto/top/subscribers",
    "https://telemetr.io/en/channels?category=cryptocurrency&sort=subscribers",
]


def scrape_telegram() -> list:
    logger.info("=" * 40)
    logger.info("渠道 3/6: Telegram (tgstat / telemetr)")
    if not BS4_OK:
        logger.warning("bs4 未安装，跳过 Telegram")
        return []

    kols = []
    seen = set()

    for url in TGSTAT_URLS:
        try:
            resp = requests.get(url, headers=HTTP_HEADERS, timeout=15, verify=False)
            if resp.status_code != 200:
                logger.warning(f"  tgstat {url} → {resp.status_code}")
                continue

            soup = BeautifulSoup(resp.text, "html.parser")
            items = soup.select(".peer-item-row, .channel-item, [class*='peer-item']")
            logger.info(f"  {url}: {len(items)} 个频道")

            for item in items[:30]:
                try:
                    name_el = item.select_one(".font-16.text-dark, .font-16.text-truncate")
                    desc_el = item.select_one(".font-14.text-muted, .line-clamp-2")
                    subs_el = item.select_one(".font-12 b, .font-12.text-truncate b")
                    link_el = item.select_one("a.text-body[href*='tgstat.com/channel']")

                    name = name_el.get_text(strip=True) if name_el else ""
                    bio  = desc_el.get_text(strip=True) if desc_el else ""
                    subs_text = subs_el.get_text(strip=True) if subs_el else ""
                    tgstat_href = link_el.get("href", "") if link_el else ""

                    if not name or name in seen:
                        continue
                    seen.add(name)

                    followers = parse_follower_count(re.sub(r"[^\d]", "", subs_text))
                    # 从 tgstat URL 提取 Telegram handle
                    tg_handle = re.search(r'/channel/@?([\w]+)', tgstat_href)
                    handle_str = tg_handle.group(1) if tg_handle else name
                    tg_url = f"https://t.me/{handle_str}"

                    kols.append(make_kol(
                        source="Telegram",
                        handle=name,
                        url=tg_url,
                        followers=followers,
                        bio=bio,
                        notes=f"Telegram频道 | {subs_text}订阅",
                    ))
                except Exception as e:
                    logger.debug(f"tgstat item parse: {e}")
        except Exception as e:
            logger.warning(f"tgstat {url} 失败: {e}")
        time.sleep(1)

    logger.info(f"Telegram: 收集 {len(kols)} 个 KOL")
    return kols


# ─────────────────────────────────────────────────────────
# CoinTelegraph 作者
# ─────────────────────────────────────────────────────────

def scrape_cointelegraph() -> list:
    logger.info("=" * 40)
    logger.info("渠道 4/6: CoinTelegraph 作者")
    if not BS4_OK:
        logger.warning("bs4 未安装，跳过 CoinTelegraph")
        return []

    kols = []
    seen = set()

    urls = [
        "https://cointelegraph.com/authors",
        "https://cointelegraph.com/magazine/contributors/",
    ]

    for url in urls:
        try:
            resp = requests.get(url, headers=HTTP_HEADERS, timeout=15)
            if resp.status_code != 200:
                continue

            soup = BeautifulSoup(resp.text, "html.parser")

            # 找到所有作者链接
            author_links = set()
            for a in soup.select("a[href*='/authors/'], a[href*='/author/']"):
                href = a.get("href", "")
                if not href.startswith("http"):
                    href = "https://cointelegraph.com" + href
                author_links.add(href)

            logger.info(f"  CoinTelegraph {url}: {len(author_links)} 个作者链接")

            for author_url in list(author_links)[:40]:
                if author_url in seen:
                    continue
                seen.add(author_url)
                try:
                    r2 = requests.get(author_url, headers=HTTP_HEADERS, timeout=10)
                    if r2.status_code != 200:
                        continue
                    s2 = BeautifulSoup(r2.text, "html.parser")

                    name_el = (s2.select_one("h1")
                               or s2.select_one("[class*='authorName']")
                               or s2.select_one("[class*='author-name']"))
                    bio_el  = (s2.select_one("[class*='authorBio']")
                               or s2.select_one("[class*='author-bio']")
                               or s2.select_one(".post-meta__author-info p"))

                    name = name_el.get_text(strip=True) if name_el else ""
                    bio  = bio_el.get_text(strip=True) if bio_el else ""

                    if not name or len(name) > 60:
                        continue

                    kols.append(make_kol(
                        source="CoinTelegraph",
                        handle=name,
                        url=author_url,
                        followers=0,
                        bio=bio,
                        website=author_url,
                        notes="CoinTelegraph媒体作者",
                    ))
                    time.sleep(0.3)
                except Exception:
                    pass
        except Exception as e:
            logger.warning(f"CoinTelegraph {url} 失败: {e}")

    logger.info(f"CoinTelegraph: 收集 {len(kols)} 个 KOL")
    return kols


# ─────────────────────────────────────────────────────────
# CoinDesk 作者
# ─────────────────────────────────────────────────────────

def scrape_coindesk() -> list:
    logger.info("=" * 40)
    logger.info("渠道 5/6: CoinDesk 作者")
    if not BS4_OK:
        logger.warning("bs4 未安装，跳过 CoinDesk")
        return []

    kols = []
    seen = set()

    # 通过 RSS 提取近期文章作者，再访问作者页
    # 爬多个页面收集作者 slug
    seed_pages = [
        "https://www.coindesk.com/",
        "https://www.coindesk.com/news/",
        "https://www.coindesk.com/markets/",
        "https://www.coindesk.com/tech/",
        "https://www.coindesk.com/policy/",
    ]
    author_slugs = set()
    for page_url in seed_pages:
        try:
            resp = requests.get(page_url, headers=HTTP_HEADERS, timeout=15, verify=False)
            if resp.status_code == 200:
                found = re.findall(r'/author/([a-z0-9\-]+)', resp.text)
                author_slugs.update(found)
                logger.info(f"  CoinDesk {page_url}: +{len(found)} 个作者")
        except Exception as e:
            logger.debug(f"CoinDesk {page_url}: {e}")

    author_urls = {f"https://www.coindesk.com/author/{slug}" for slug in author_slugs
                   if slug and slug not in ("tag", "category", "type")}
    logger.info(f"  CoinDesk 总计: {len(author_urls)} 个作者URL")

    logger.info(f"  CoinDesk 总计: {len(author_urls)} 个作者URL")

    for author_url in list(author_urls)[:40]:
        author_url = author_url.rstrip("/")
        if author_url in seen:
            continue
        seen.add(author_url)
        try:
            r2 = requests.get(author_url, headers=HTTP_HEADERS, timeout=10, verify=False)
            if r2.status_code != 200:
                continue
            s2 = BeautifulSoup(r2.text, "html.parser")

            name_el = (s2.select_one("h1")
                       or s2.select_one("[class*='authorName']")
                       or s2.select_one("[class*='author-name']")
                       or s2.select_one("[class*='AuthorName']"))
            bio_el  = (s2.select_one("[class*='authorBio']")
                       or s2.select_one("[class*='author-bio']")
                       or s2.select_one("[class*='AuthorBio']"))

            name = name_el.get_text(strip=True) if name_el else ""
            bio  = bio_el.get_text(strip=True) if bio_el else ""

            # 从 URL 提取名字备用
            if not name:
                slug = author_url.rstrip("/").split("/")[-1]
                name = slug.replace("-", " ").title()

            if not name or len(name) > 60:
                continue

            kols.append(make_kol(
                source="CoinDesk",
                handle=name,
                url=author_url,
                followers=0,
                bio=bio,
                website=author_url,
                notes="CoinDesk媒体作者",
            ))
            time.sleep(0.3)
        except Exception:
            pass

    logger.info(f"CoinDesk: 收集 {len(kols)} 个 KOL")
    return kols


# ─────────────────────────────────────────────────────────
# Lunarcrush 影响力榜
# ─────────────────────────────────────────────────────────

def scrape_lunarcrush(browser=None) -> list:
    logger.info("=" * 40)
    logger.info("渠道 6/6: Lunarcrush")
    if not BS4_OK:
        logger.warning("bs4 未安装，跳过 Lunarcrush")
        return []

    kols = []

    # 尝试公开 API（无需登录的端点）
    api_urls = [
        "https://lunarcrush.com/api4/public/creators/v1?limit=50&order_by=followers",
        "https://lunarcrush.com/api4/public/influencers/v1?limit=50",
    ]
    for api_url in api_urls:
        try:
            resp = requests.get(api_url, headers=HTTP_HEADERS, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                items = data.get("data", [])
                logger.info(f"  Lunarcrush API: {len(items)} 条")
                for item in items:
                    name = item.get("name", "") or item.get("display_name", "")
                    handle = item.get("screen_name", "") or item.get("username", "")
                    followers = item.get("followers", 0) or item.get("twitter_followers", 0)
                    bio = item.get("description", "") or item.get("bio", "")
                    url = f"https://x.com/{handle}" if handle else ""

                    if not handle or followers < MIN_FOLLOWERS:
                        continue

                    kols.append(make_kol(
                        source="Lunarcrush",
                        handle="@" + handle.lstrip("@"),
                        url=url,
                        followers=followers,
                        bio=bio,
                        name=name,
                        notes=f"Lunarcrush影响力榜 | {followers:,}粉",
                    ))
                break  # API 成功就不用继续
        except Exception as e:
            logger.debug(f"Lunarcrush API {api_url}: {e}")

    # API 失败时用 Playwright 加载 JS 页
    if not kols and browser:
        try:
            ctx = browser.new_context()
            page = ctx.new_page()
            page.goto("https://lunarcrush.com/discover/creators", wait_until="domcontentloaded", timeout=20000)
            page.wait_for_timeout(4000)
            for _ in range(3):
                page.keyboard.press("End")
                page.wait_for_timeout(1000)

            data = page.evaluate("""() => {
                const items = [];
                document.querySelectorAll('[class*="creator"], [class*="influencer"], [class*="Creator"]').forEach(el => {
                    const name = el.querySelector('[class*="name"], h3, h4')?.innerText?.trim() || '';
                    const handle = el.querySelector('[class*="handle"], [class*="username"]')?.innerText?.trim()?.replace('@','') || '';
                    const fol = el.querySelector('[class*="follower"], [class*="subscriber"]')?.innerText?.trim() || '';
                    const link = el.querySelector('a[href]')?.getAttribute('href') || '';
                    if (name) items.push({name, handle, fol, link});
                });
                return items;
            }""")

            for item in data:
                name = item.get("name", "")
                handle = item.get("handle", "") or name
                subs_text = item.get("fol", "")
                followers = parse_follower_count(re.sub(r"[^\d\.KMk]", "", subs_text))
                if not name or followers < MIN_FOLLOWERS:
                    continue
                kols.append(make_kol(
                    source="Lunarcrush",
                    handle="@" + handle.lstrip("@"),
                    url=f"https://x.com/{handle}",
                    followers=followers,
                    bio="",
                    name=name,
                    notes="Lunarcrush影响力榜",
                ))
            page.close()
            ctx.close()
        except Exception as e:
            logger.warning(f"Lunarcrush Playwright爬取失败: {e}")

    logger.info(f"Lunarcrush: 收集 {len(kols)} 个 KOL")
    return kols


# ─────────────────────────────────────────────────────────
# Excel 输出
# ─────────────────────────────────────────────────────────

def build_excel(kols: list, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "KOL总表"

    orange    = PatternFill("solid", fgColor="FF6B00")
    dark_blue = PatternFill("solid", fgColor="1F3864")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="FFFFFF", size=13)
    center    = Alignment(horizontal="center", vertical="center")
    thin      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # 标题行
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = f"MoonX Web3 KOL 多渠道名单（{TODAY}）"
    c.fill, c.font, c.alignment = dark_blue, title_font, center
    ws.row_dimensions[1].height = 28

    headers = [
        "来源渠道", "区域", "国家/地区", "语言", "用户名", "主页链接",
        "粉丝数", "分级", "Bio摘要", "官网", "邮箱", "Telegram",
        "已合作交易所", "BYDFi合作优先级", "备注", "是否联繫",
        "BD负责人", "联繫进度", "更新日期",
    ]
    col_widths = [14, 10, 12, 10, 22, 35, 10, 14, 45, 30, 28, 30, 18, 14, 30, 10, 12, 12, 12]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = orange, hdr_font, center, thin
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w
    ws.row_dimensions[2].height = 22

    tier_colors = {
        "A级(100万+)":  "FFD700",
        "A级(50万+)":   "FFC200",
        "B级(10万+)":   "C6EFCE",
        "C级(1万+)":    "DDEBF7",
        "D级(1万以下)": "F2F2F2",
        "媒体KOL":       "EDE7F6",
    }

    for i, kol in enumerate(kols, 3):
        tier_color = tier_colors.get(kol["tier"], "FFFFFF")
        source_color = SOURCE_COLORS.get(kol["source"], "FFFFFF")
        # 偶数行用渠道色，奇数行用 tier 色（混合区分）
        fill_hex = source_color if i % 2 == 0 else tier_color
        fill = PatternFill("solid", fgColor=fill_hex)

        row_data = [
            kol["source"],
            kol["region"], kol["country"], kol["language"],
            kol["handle"], kol["url"],
            kol["followers"], kol["tier"],
            kol["bio"], kol["website"], kol["email"], kol.get("telegram", ""),
            kol["exchanges"], kol["priority"], kol["notes"],
            kol["contacted"], "", "", kol["date"],
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill, cell.border = fill, thin
            cell.alignment = Alignment(vertical="center", wrap_text=(j in (9, 14)))
        ws.row_dimensions[i].height = 18

    ws.freeze_panes = "A3"
    wb.save(output_path)
    logger.info(f"已保存: {output_path.name}（{len(kols)} 条）")


# ─────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────

def run(headed: bool = False):
    existing = load_existing_handles()
    logger.info("=" * 60)
    logger.info(f"MoonX Web3 KOL 多渠道爬取 — {TODAY}")
    logger.info(f"已有记录: {len(existing)} 个（去重）")
    logger.info(f"模式: {'有头（可见）' if headed else 'Headless'}")
    logger.info("渠道: Twitter | YouTube | Telegram | CoinTelegraph | CoinDesk | Lunarcrush")
    logger.info("=" * 60)

    all_kols = []

    with sync_playwright() as pw:
        # ── 渠道 1: Twitter（优先 CDP，fallback 独立浏览器）──
        try:
            kols = scrape_twitter(pw, existing)
            all_kols.extend(kols)
        except Exception as e:
            logger.error(f"Twitter 渠道失败: {e}")

        # ── 渠道 2 & 6: YouTube + Lunarcrush（独立浏览器）──
        try:
            launch_opts = {"headless": not headed, "args": ["--no-sandbox"]}
            browser = pw.chromium.launch(**launch_opts)

            try:
                kols = scrape_youtube(browser)
                all_kols.extend(kols)
            except Exception as e:
                logger.error(f"YouTube 渠道失败: {e}")

            try:
                kols = scrape_lunarcrush(browser)
                all_kols.extend(kols)
            except Exception as e:
                logger.error(f"Lunarcrush 渠道失败: {e}")

            browser.close()
        except Exception as e:
            logger.error(f"Playwright 浏览器启动失败: {e}")

    # ── 渠道 3-5: requests-based ──
    for fn, name in [
        (scrape_telegram,      "Telegram"),
        (scrape_cointelegraph, "CoinTelegraph"),
        (scrape_coindesk,      "CoinDesk"),
    ]:
        try:
            kols = fn()
            all_kols.extend(kols)
        except Exception as e:
            logger.error(f"{name} 渠道失败: {e}")

    if not all_kols:
        logger.info("没有收集到 KOL，结束")
        return

    # 全局去重（按 url）
    seen_urls, unique_kols = set(), []
    for k in all_kols:
        key = k["url"].lower().rstrip("/")
        if key and key not in seen_urls:
            seen_urls.add(key)
            unique_kols.append(k)

    unique_kols.sort(key=lambda x: x["followers"], reverse=True)

    # 统计
    from collections import Counter
    src_count  = Counter(k["source"] for k in unique_kols)
    tier_count = Counter(k["tier"] for k in unique_kols)
    with_email    = sum(1 for k in unique_kols if k.get("email"))
    with_telegram = sum(1 for k in unique_kols if k.get("telegram"))

    logger.info(f"\n渠道来源: {dict(src_count)}")
    logger.info(f"分级统计: {dict(tier_count)}")
    logger.info(f"有邮箱: {with_email} / {len(unique_kols)}")
    logger.info(f"有Telegram: {with_telegram} / {len(unique_kols)}")

    # 保存 Excel
    out = SCRIPT_DIR / f"MoonX_Web3_KOL_{TODAY}.xlsx"
    build_excel(unique_kols, out)

    # 保存待 DM 名单（无邮箱的 Twitter KOL）
    no_email = [
        {"handle": k["handle"], "name": k["name"], "followers": k["followers"],
         "tier": k["tier"], "source": k["source"]}
        for k in unique_kols if not k["email"] and k["source"] == "Twitter"
    ]
    if no_email:
        dm_out = SCRIPT_DIR / f"MoonX_待DM_{TODAY}.json"
        with open(dm_out, "w") as f:
            json.dump(no_email, f, ensure_ascii=False, indent=2)
        logger.info(f"待 DM 名单: {dm_out.name}（{len(no_email)} 条）")

    logger.info(f"\n完成！共收集 {len(unique_kols)} 个 KOL（6个渠道）")
    logger.info(f"文件: {out.name}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--headed", action="store_true", help="有头模式（可见浏览器）")
    args = parser.parse_args()
    run(headed=args.headed)
